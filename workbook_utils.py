import html
import json
import os
import re
import unicodedata
import urllib.request
from datetime import datetime
from typing import Iterable
from urllib.parse import urlparse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_DIR_NAME = "data"
GOOGLE_SHEET_FILENAME_PREFIX = "google_sheet_"
LEGACY_GOOGLE_SHEET_FILE_ID = "data/google_sheet_main.xlsx"
GOOGLE_SHEET_LABEL = "Google Sheet chính"
GOOGLE_SHEET_REGISTRY_FILENAME = "google_sheet_sources.json"
REPORT_COLUMNS = ["NGÀY AIR", "TÊN KÊNH", "LINK AIR", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"]
SUMMARY_SHEET_NAME = "Tổng kết"
SUMMARY_SHEET_TITLE_PREFIX = "Tổng kết "
LAST_UPDATE_COLUMN = "Cập nhật lần cuối"
SUMMARY_COLUMNS = ["Stt", "ĐỐI TÁC", "TỔNG LINK", "TỔNG LƯỢT XEM", "TỔNG TIM", "TỔNG BÌNH LUẬN", "TỔNG LƯỢT LƯU", "TỔNG CHIA SẺ", LAST_UPDATE_COLUMN]
SUMMARY_TOTAL_LABEL = "TỔNG"
# Cam cảnh báo đối tác chỉ có đúng 1 link — cần chú ý khi báo cáo/nghiệm thu.
SINGLE_LINK_FILL_COLOR = "FFC000"
METRIC_COLUMNS = ["LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"]
SUMMARY_METRIC_COLUMNS = ["TỔNG LƯỢT XEM", "TỔNG TIM", "TỔNG BÌNH LUẬN", "TỔNG LƯỢT LƯU", "TỔNG CHIA SẺ"]
PARTNER_HEADING_MARKERS = ("DANH SÁCH", "DANH SACH", "BỘ ẢNH", "BO ANH")
CHANNEL_OVERRIDE_FILENAME = "channel_name_overrides.json"
DEFAULT_CHANNEL_OVERRIDES = {}
RESULT_SHEET_PREFIX = "report seeding tiktok"
RESULT_SHEET_TIMESTAMP_RE = re.compile(
    r"^(?:\d{2}-\d{2}-\d{4}-\d{2}-\d{2}|\d{2}-\d{2}-\d{4}-\d{2}:\d{2})(?:-\d+)?$"
)
DISPLAY_DATETIME_FORMAT = "%d/%m/%Y-%H:%M"
FILENAME_DATETIME_FORMAT = "%d-%m-%Y-%H-%M"
SHEET_DATETIME_FORMAT = "%d-%m-%Y-%H:%M"


def ensure_data_dir(base_dir):
    data_dir = os.path.join(base_dir, DATA_DIR_NAME)
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def is_internal_workbook_filename(filename):
    """Skip Excel temp files and underscore-prefixed internal/test workbooks."""
    name = os.path.basename(str(filename or ""))
    return not name or name.startswith("~$") or name.startswith("_")


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_key(value):
    text = str(value or "").replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text.strip()).casefold()


def clean_text(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def is_numeric_channel_garbage(value):
    text = clean_text(value)
    if not text:
        return False
    normalized = text.replace(",", "").strip()
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", normalized))


def normalize_channel_display(value):
    text = clean_text(value)
    channel_fixes = {
        "B?o Quy?n": "Bảo Quyên",
    }
    return channel_fixes.get(text, text)


def is_generated_username_channel(value, link=""):
    text = clean_text(value)
    if not text:
        return False
    compact = re.sub(r"\s+", "", text.lstrip("@"))
    if re.fullmatch(r"user\d{4,}.*", compact, flags=re.IGNORECASE):
        return True
    return False


def is_generic_tiktok_channel_name(value):
    key = normalize_key(value)
    if key in {
        "tiktok",
        "make your day",
        "tiktok make your day",
        "tiktok - make your day",
        "screen time breaks",
        "screen time break",
    }:
        return True
    if "tiktok" in key and "make your day" in key:
        return True
    if "screen time" in key:
        return True
    return False


def display_channel_name_from_file(link, raw_channel):
    normalized_channel = normalize_channel_display(raw_channel)
    if not normalized_channel:
        return ""
    if (
        is_numeric_channel_garbage(normalized_channel)
        or is_generated_username_channel(normalized_channel, link)
        or is_generic_tiktok_channel_name(normalized_channel)
    ):
        return "Lỗi"
    return normalized_channel


def resolve_channel_name(link, raw_channel, channel_overrides=None):
    username = extract_tiktok_username(link)
    normalized_channel = display_channel_name_from_file(link, raw_channel)
    if normalized_channel:
        return normalized_channel
    if username and channel_overrides:
        override_name = channel_overrides.get(username.casefold())
        if override_name:
            return override_name
    return ""


def clean_preview_value(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = clean_text(value)
    if text.endswith(" 00:00:00"):
        return text[:10]
    return text


def parse_google_spreadsheet_id(source_url):
    parsed = urlparse(source_url.strip())
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        raise ValueError("URL Google Sheet không hợp lệ")
    return match.group(1)


def extract_tiktok_username(url):
    match = re.search(r"tiktok\.com/@([^/?]+)", str(url or ""), re.IGNORECASE)
    return match.group(1).strip() if match else ""


def google_sheet_file_id(spreadsheet_id):
    safe_id = re.sub(r"[^a-zA-Z0-9_-]+", "", spreadsheet_id)
    if not safe_id:
        raise ValueError("Google Sheet ID không hợp lệ")
    return f"{DATA_DIR_NAME}/{GOOGLE_SHEET_FILENAME_PREFIX}{safe_id}.xlsx"


def safe_workbook_filename(name, *, max_length=80):
    filename = re.sub(r'[\\/:*?"<>|]+', "-", clean_text(name))
    filename = re.sub(r"\s+", " ", filename).strip(" .")
    return filename[:max_length] if filename else "google-sheet"


def fetch_google_spreadsheet_title(source_url):
    spreadsheet_id = parse_google_spreadsheet_id(source_url)
    view_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    request = urllib.request.Request(view_url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            page_html = response.read().decode("utf-8", errors="ignore")
    except Exception:
        return f"Google Sheet {spreadsheet_id[:12]}"

    for pattern in (
        r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:title["\']',
        r"<title[^>]*>([^<]+)</title>",
    ):
        match = re.search(pattern, page_html, flags=re.IGNORECASE)
        if not match:
            continue
        title = clean_text(html.unescape(match.group(1)))
        title = re.sub(r"\s*-\s*Google Sheets\s*$", "", title, flags=re.IGNORECASE)
        title = re.sub(r"\s*-\s*Google Trang tính\s*$", "", title, flags=re.IGNORECASE)
        if title:
            return title
    return f"Google Sheet {spreadsheet_id[:12]}"


def format_display_datetime(moment=None):
    value = moment or datetime.now()
    return value.strftime(DISPLAY_DATETIME_FORMAT)


def format_filename_datetime(moment=None):
    value = moment or datetime.now()
    return value.strftime(FILENAME_DATETIME_FORMAT)


def format_excel_sheet_datetime(moment=None):
    value = moment or datetime.now()
    return value.strftime(SHEET_DATETIME_FORMAT)


def parse_filename_datetime_stamp(stamp):
    text = clean_text(stamp)
    match = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})-(\d{2})-(\d{2})", text)
    if not match:
        return text
    day, month, year, hour, minute = match.groups()
    return f"{day}/{month}/{year}-{hour}:{minute}"


def google_sheet_sync_timestamp_display(moment=None):
    return format_display_datetime(moment)


def google_sheet_sync_timestamp_filename(moment=None):
    return format_filename_datetime(moment)


def google_sheet_sync_label(title, timestamp_display=None):
    sheet_title = clean_text(title)
    stamp = clean_text(timestamp_display or google_sheet_sync_timestamp_display())
    return f"{sheet_title} {stamp}".strip()


def google_sheet_file_id_from_title(title, timestamp):
    safe_title = safe_workbook_filename(title)
    safe_stamp = safe_workbook_filename(timestamp, max_length=20)
    return f"{DATA_DIR_NAME}/{safe_title} {safe_stamp}.xlsx"


def google_sheet_filename_to_label(filename):
    base = clean_text(str(filename or "")).removesuffix(".xlsx")
    match = re.fullmatch(r"(.+?) (\d{2})-(\d{2})-(\d{4})-(\d{2})-(\d{2})", base)
    if match:
        title = match.group(1)
        stamp = base[len(title) + 1 :]
        return f"{title} {parse_filename_datetime_stamp(stamp)}"
    legacy = re.fullmatch(r"(.+?)-(\d{2})-(\d{2})-(\d{4})-(\d{2})-(\d{2})", base)
    if legacy:
        title = legacy.group(1)
        stamp = "-".join(legacy.groups()[1:])
        return f"{title} {parse_filename_datetime_stamp(stamp)}"
    return base


def timestamped_google_sheet_file_id(timestamp):
    return google_sheet_file_id_from_title("Report Seeding Tiktok", timestamp)


def google_sheet_registry_path(base_dir):
    return os.path.join(ensure_data_dir(base_dir), GOOGLE_SHEET_REGISTRY_FILENAME)


def load_google_sheet_registry(base_dir):
    path = google_sheet_registry_path(base_dir)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as file_obj:
            data = json.load(file_obj)
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def save_google_sheet_registry(base_dir, registry):
    path = google_sheet_registry_path(base_dir)
    with open(path, "w", encoding="utf-8") as file_obj:
        json.dump(registry, file_obj, ensure_ascii=False, indent=2)


def register_google_sheet_source(base_dir, file_id, source_url, title=""):
    registry = load_google_sheet_registry(base_dir)
    registry[file_id] = {
        "url": clean_text(source_url),
        "spreadsheetId": parse_google_spreadsheet_id(source_url),
        "title": clean_text(title),
    }
    save_google_sheet_registry(base_dir, registry)
    return registry[file_id]


def google_sheet_source_for_file(base_dir, file_id):
    registry = load_google_sheet_registry(base_dir)
    source = registry.get(file_id)
    if isinstance(source, dict):
        return source

    normalized_file_id = str(file_id or "").replace("\\", "/")
    for registered_file_id, registered_source in registry.items():
        if str(registered_file_id).replace("\\", "/") == normalized_file_id and isinstance(registered_source, dict):
            return registered_source

    return {}


def base_dir_for_file(file_path):
    absolute_path = os.path.abspath(file_path)
    parent = os.path.dirname(absolute_path)
    if os.path.basename(parent).casefold() == DATA_DIR_NAME:
        return os.path.dirname(parent)
    return parent


def channel_override_path(file_path):
    return os.path.join(ensure_data_dir(base_dir_for_file(file_path)), CHANNEL_OVERRIDE_FILENAME)


def load_channel_overrides(file_path):
    path = channel_override_path(file_path)
    overrides = {
        str(key).casefold(): clean_text(value)
        for key, value in DEFAULT_CHANNEL_OVERRIDES.items()
        if clean_text(value)
    }
    if not os.path.exists(path):
        return overrides
    try:
        with open(path, "r", encoding="utf-8") as file_obj:
            data = json.load(file_obj)
    except (OSError, json.JSONDecodeError):
        return overrides
    if not isinstance(data, dict):
        return overrides
    overrides.update({str(key).casefold(): clean_text(value) for key, value in data.items() if clean_text(value)})
    return overrides


def safe_join(base_dir, relative_path):
    """Resolve a relative path under base_dir; reject traversal outside base."""
    if not relative_path:
        return ""
    normalized = str(relative_path).replace("\\", "/").lstrip("/")
    if ".." in normalized.split("/"):
        return ""
    candidate = os.path.normpath(os.path.join(base_dir, normalized.replace("/", os.sep)))
    base_abs = os.path.abspath(base_dir)
    candidate_abs = os.path.abspath(candidate)
    if candidate_abs != base_abs and not candidate_abs.startswith(base_abs + os.sep):
        return ""
    return candidate


def metric_number(value):
    if pd.isna(value) or value == "" or value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(float(value))
    raw_text = str(value).strip()
    if not raw_text:
        return 0
    if re.fullmatch(r"\d{1,3}([.,]\d{3})+", raw_text):
        return int(re.sub(r"[.,]", "", raw_text))
    if re.fullmatch(r"\d+\.0+", raw_text):
        return int(float(raw_text))
    compact = re.sub(r"[,\s]", "", raw_text)
    try:
        return int(float(compact))
    except ValueError:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def format_metric(value):
    if pd.isna(value) or value == "" or value is None:
        return 0
    if isinstance(value, (int, float)):
        number = float(value)
        return int(number) if number.is_integer() else number
    try:
        number = float(value)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return value


def to_number(value):
    return metric_number(value)


def google_sheet_export_url(source_url):
    spreadsheet_id = parse_google_spreadsheet_id(source_url)
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def download_google_sheet(source_url, destination_path):
    export_url = google_sheet_export_url(source_url)
    request = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        content = response.read()
    with open(destination_path, "wb") as file_obj:
        file_obj.write(content)
    return destination_path


def load_excel_file(file_path):
    return pd.ExcelFile(file_path, engine="openpyxl")


def workbook_sheet_names(file_path):
    workbook = load_excel_file(file_path)
    try:
        return list(workbook.sheet_names)
    finally:
        workbook.close()


def read_sheet_frame(file_path, sheet_name):
    return pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")


def summary_sheet_title_for_data_sheet(data_sheet_name):
    """Build Excel tab title for a data sheet's partner summary (e.g. Tháng 6 → Tổng kết tháng 6)."""
    label = clean_text(data_sheet_name).casefold()
    if not label:
        return SUMMARY_SHEET_NAME
    return f"{SUMMARY_SHEET_TITLE_PREFIX}{label}"[:31]


def data_sheet_name_for_summary_title(workbook_sheet_names, summary_sheet_name):
    """Resolve the source data sheet name from a per-sheet summary tab title."""
    text = clean_text(summary_sheet_name)
    if normalize_key(text) == normalize_key(SUMMARY_SHEET_NAME):
        return ""
    prefix_folded = SUMMARY_SHEET_TITLE_PREFIX.casefold()
    if not text.casefold().startswith(prefix_folded):
        return ""
    suffix = text[len(SUMMARY_SHEET_TITLE_PREFIX) :].strip()
    if not suffix:
        return ""
    for candidate in workbook_sheet_names:
        if is_summary_sheet_name(candidate) or is_result_sheet_name(candidate):
            continue
        if clean_text(candidate).casefold() == suffix:
            return candidate
    return ""


def is_summary_sheet_name(sheet_name):
    sheet_key = normalize_key(sheet_name)
    if sheet_key == normalize_key(SUMMARY_SHEET_NAME):
        return True
    return sheet_key.startswith(normalize_key(SUMMARY_SHEET_TITLE_PREFIX))


def is_result_sheet_timestamp_name(sheet_name):
    return bool(RESULT_SHEET_TIMESTAMP_RE.fullmatch(clean_text(sheet_name)))


def is_result_sheet_name(sheet_name):
    if normalize_key(sheet_name).startswith(RESULT_SHEET_PREFIX):
        return True
    return is_result_sheet_timestamp_name(sheet_name)


def is_total_label(value):
    return normalize_key(clean_text(value)) == "tong"


def is_tiktok_link(value):
    return "tiktok.com" in clean_text(value).casefold()


def normalize_tiktok_url(value):
    """Ensure TikTok URLs have a scheme so browsers can navigate them."""
    text = clean_text(value)
    if not text:
        return ""
    lower = text.casefold()
    if "tiktok.com" not in lower and "vt.tiktok.com" not in lower:
        return text
    if lower.startswith("http://") or lower.startswith("https://"):
        return text
    if text.startswith("//"):
        return f"https:{text}"
    return f"https://{text.lstrip('/')}"


def is_scrapable_tiktok_url(value):
    """True when the cell value is a TikTok link row worth scraping (not TỔNG, not junk)."""
    url = normalize_tiktok_url(value)
    if not url or is_total_label(url):
        return False
    lower = url.casefold()
    return "tiktok.com" in lower or "vt.tiktok.com" in lower


def fill_preview_total_row(frame, link_column, metric_columns):
    if not link_column or link_column not in frame.columns:
        return frame

    total_positions = [
        index
        for index, value in enumerate(frame[link_column].tolist())
        if is_total_label(value)
    ]
    if not total_positions:
        return frame

    total_position = total_positions[-1]
    if total_position <= 0:
        return frame

    source_frame = frame.iloc[:total_position]
    for column in metric_columns:
        if not column or column not in frame.columns:
            continue
        total_value = pd.to_numeric(source_frame[column], errors="coerce").fillna(0).sum()
        frame.iat[total_position, frame.columns.get_loc(column)] = int(total_value)
    return frame


def fill_missing_dates_from_previous(frame, date_column, link_column):
    if not date_column or date_column not in frame.columns:
        return frame

    last_date = ""
    for index, row in frame.iterrows():
        raw_date = row.get(date_column, "")
        cleaned_date = clean_text(raw_date)
        link = clean_text(row.get(link_column, "")) if link_column else ""

        if cleaned_date:
            last_date = raw_date
            continue
        if last_date and ("tiktok.com" in link or "vt.tiktok.com" in link):
            frame.at[index, date_column] = last_date
    return frame


def read_sheet_preview(file_path, sheet_name=None, limit=None):
    workbook = load_excel_file(file_path)
    try:
        sheets = list(workbook.sheet_names)
        current_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else "")
        if not current_sheet:
            return {"sheets": [], "currentSheet": "", "columns": [], "data": [], "message": "Workbook không có sheet nào."}

        frame = workbook.parse(current_sheet)
        link_column = find_link_column_name(frame)
        channel_column = find_column_name(frame, ["TÊN KÊNH", "Tên Kênh"])
        date_column = find_column_name(frame, ["NGÀY AIR", "Ngày"])
        partner_columns = dataframe_partner_columns(frame)
        metric_columns = [
            find_column_name(frame, ["LƯỢT XEM"]),
            find_column_name(frame, ["TIM"]),
            find_column_name(frame, ["BÌNH LUẬN"]),
            find_column_name(frame, ["LƯỢT LƯU"]),
            find_column_name(frame, ["CHIA SẺ"]),
        ]
        frame = fill_missing_dates_from_previous(frame, date_column, link_column)
        for column in frame.select_dtypes(include=["datetime"]).columns:
            frame[column] = frame[column].dt.strftime(DISPLAY_DATETIME_FORMAT)
        frame = fill_preview_total_row(frame, link_column, metric_columns)

        preview_frame = frame
        if limit and link_column and len(frame.index) > limit:
            total_positions = [
                index
                for index, value in enumerate(frame[link_column].tolist())
                if is_total_label(value)
            ]
            if total_positions and total_positions[-1] >= limit:
                preview_frame = pd.concat(
                    [frame.head(max(limit - 1, 0)), frame.iloc[[total_positions[-1]]]],
                    ignore_index=True,
                )
            else:
                preview_frame = frame.head(limit)
        elif limit:
            preview_frame = frame.head(limit)

        data = []
        for record in preview_frame.to_dict(orient="records"):
            preview_row = {}
            for column, value in record.items():
                cleaned = clean_preview_value(value)
                preview_row[column] = cleaned
            if link_column and channel_column:
                preview_row[channel_column] = display_channel_name_from_file(
                    record.get(link_column, ""),
                    preview_row.get(channel_column, ""),
                )
            if link_column and partner_columns and is_tiktok_link(record.get(link_column, "")):
                row_partners = extract_row_partners(record, partner_columns)
                if len(row_partners) == 1:
                    preview_row["_singlePartner"] = True
            data.append(preview_row)

        return {
            "sheets": sheets,
            "currentSheet": current_sheet,
            "columns": frame.columns.tolist(),
            "data": data,
            "totalRows": int(len(frame.index)),
            "shownRows": int(len(preview_frame.index)),
        }
    finally:
        workbook.close()


def read_summary_dashboard(file_path, data_sheet_name=None):
    workbook = load_excel_file(file_path)
    try:
        sheet_names = list(workbook.sheet_names)
        requested_data_sheet = clean_text(data_sheet_name)
        if requested_data_sheet:
            summary_sheet = summary_sheet_title_for_data_sheet(requested_data_sheet)
            if summary_sheet not in sheet_names:
                summary_sheet = ""
        else:
            summary_sheet = next((sheet for sheet in sheet_names if is_summary_sheet_name(sheet)), "")
        if not summary_sheet:
            label = requested_data_sheet or "sheet này"
            return {
                "sheet": "",
                "dataSheet": requested_data_sheet,
                "columns": SUMMARY_COLUMNS,
                "rows": [],
                "totals": {},
                "message": f"Workbook chưa có tổng kết cho {label}. Quét sheet đó để tạo.",
            }

        resolved_data_sheet = requested_data_sheet or data_sheet_name_for_summary_title(
            sheet_names, summary_sheet
        )
        frame = workbook.parse(summary_sheet).fillna("")
    finally:
        workbook.close()

    columns = frame.columns.tolist()
    if LAST_UPDATE_COLUMN not in columns:
        columns.append(LAST_UPDATE_COLUMN)
    rows = []
    partner_column = find_column_name(frame, ["ĐỐI TÁC", "Đối tác"])
    total_link_column = find_column_name(frame, ["TỔNG LINK", "Tổng link"])
    summary_metric_map = {
        "views": find_column_name(frame, ["TỔNG LƯỢT XEM", "LƯỢT XEM"]),
        "likes": find_column_name(frame, ["TỔNG TIM", "TIM"]),
        "comments": find_column_name(frame, ["TỔNG BÌNH LUẬN", "BÌNH LUẬN"]),
        "saves": find_column_name(frame, ["TỔNG LƯỢT LƯU", "LƯỢT LƯU"]),
        "shares": find_column_name(frame, ["TỔNG CHIA SẺ", "CHIA SẺ"]),
    }

    numeric_summary_keys = {
        normalize_key(item)
        for item in SUMMARY_COLUMNS
        if normalize_key(item) not in {"doi tac", normalize_key(LAST_UPDATE_COLUMN)}
    }
    totals_row = None
    for record in frame.to_dict(orient="records"):
        partner_name = clean_text(record.get(partner_column or "", ""))
        if not partner_name:
            continue
        if is_total_label(partner_name):
            totals_row = {}
            for column in columns:
                value = record.get(column, "")
                totals_row[column] = to_number(value) if normalize_key(column) in numeric_summary_keys else clean_text(value)
            totals_row.setdefault(LAST_UPDATE_COLUMN, "")
            continue
        row = {}
        for column in columns:
            value = record.get(column, "")
            row[column] = to_number(value) if normalize_key(column) in numeric_summary_keys else clean_text(value)
        row.setdefault(LAST_UPDATE_COLUMN, "")
        rows.append(row)

    totals = {
        "partners": len(rows),
        "links": 0,
        "views": 0,
        "likes": 0,
        "comments": 0,
        "saves": 0,
        "shares": 0,
    }
    for row in rows:
        totals["links"] += to_number(row.get(total_link_column, 0)) if total_link_column else 0
        for key, column in summary_metric_map.items():
            if column:
                totals[key] += to_number(row.get(column, 0))

    if totals_row:
        totals = {
            "partners": len(rows),
            "links": to_number(totals_row.get(total_link_column, 0)) if total_link_column else totals["links"],
            "views": to_number(totals_row.get(summary_metric_map["views"], 0)) if summary_metric_map["views"] else totals["views"],
            "likes": to_number(totals_row.get(summary_metric_map["likes"], 0)) if summary_metric_map["likes"] else totals["likes"],
            "comments": to_number(totals_row.get(summary_metric_map["comments"], 0)) if summary_metric_map["comments"] else totals["comments"],
            "saves": to_number(totals_row.get(summary_metric_map["saves"], 0)) if summary_metric_map["saves"] else totals["saves"],
            "shares": to_number(totals_row.get(summary_metric_map["shares"], 0)) if summary_metric_map["shares"] else totals["shares"],
        }

    return {
        "sheet": summary_sheet,
        "dataSheet": resolved_data_sheet,
        "columns": columns,
        "rows": rows,
        "totals": totals,
    }


def dataframe_partner_columns(frame):
    columns = list(frame.columns)
    start_index = None
    for index, column in enumerate(columns):
        if normalize_key(column).startswith("doi tac"):
            start_index = index
            break

    if start_index is None:
        return []

    partner_columns = [columns[start_index]]
    for column in columns[start_index + 1:]:
        normalized = normalize_key(column)
        raw_normalized = normalize_header(column)
        if normalized.startswith("doi tac") or raw_normalized.startswith("unnamed:") or raw_normalized == "":
            partner_columns.append(column)
        else:
            break
    return partner_columns


def split_partner_value(value):
    text = clean_text(value)
    if not text:
        return []

    raw_lines = [line.strip() for line in text.replace("\r", "\n").split("\n") if line.strip()]
    if len(raw_lines) <= 1 and not any(marker in text.upper() for marker in PARTNER_HEADING_MARKERS) and not text.lstrip().startswith(("-", "*", "•")):
        cleaned = _normalize_partner_token(text)
        return [cleaned] if cleaned else []

    partners = []
    for line in raw_lines:
        cleaned = _normalize_partner_token(line)
        if not cleaned:
            continue
        upper_line = cleaned.upper()
        if any(marker in upper_line for marker in PARTNER_HEADING_MARKERS):
            continue
        partners.append(cleaned)
    return unique_preserve_order(partners)


def _normalize_partner_token(text):
    cleaned = clean_text(text)
    cleaned = unicodedata.normalize("NFC", cleaned)
    cleaned = re.sub(r"[\u200B-\u200D\uFEFF\u00AD]", "", cleaned)
    cleaned = re.sub(r"^[\-\*\u2022]+\s*", "", cleaned)
    cleaned = re.sub(r"^\d+[\.\)]\s*", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -:\t")
    return cleaned


def partner_dedup_key(value):
    text = unicodedata.normalize("NFC", str(value or ""))
    text = re.sub(r"[\u200B-\u200D\uFEFF\u00AD]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def extract_row_partners(row, partner_columns: Iterable):
    partners = []
    for column in partner_columns:
        for partner in split_partner_value(row.get(column, "")):
            partners.append(partner)
    return unique_preserve_order(partners)


def unique_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        key = partner_dedup_key(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def find_column_name(frame, aliases):
    alias_set = {normalize_key(alias) for alias in aliases}
    for column in frame.columns:
        column_key = normalize_key(column)
        if column_key in alias_set:
            return column
    return None


def find_link_column_name(frame):
    column = find_column_name(frame, ["LINK AIR", "Link", "URL"])
    if column:
        return column
    for candidate in frame.columns:
        key = normalize_key(candidate)
        if "link" in key or "url" in key:
            return candidate
    return None


def find_data_sheet_names(file_path):
    workbook = load_excel_file(file_path)
    try:
        return find_data_sheet_names_in_workbook(workbook)
    finally:
        workbook.close()


def find_data_sheet_names_in_workbook(workbook):
    sheets = workbook.sheet_names
    data_sheets = [sheet for sheet in sheets if not is_summary_sheet_name(sheet) and not is_result_sheet_name(sheet)]
    return data_sheets


def build_workbook_rows(file_path, selected_partner=None, sheet_name=None):
    workbook = load_excel_file(file_path)
    try:
        selected_key = selected_partner.casefold() if selected_partner else None
        requested_sheet = clean_text(sheet_name)
        rows = []
        data_sheets = find_data_sheet_names_in_workbook(workbook)
        if requested_sheet:
            data_sheets = [requested_sheet] if requested_sheet in workbook.sheet_names else []

        for sheet_name in data_sheets:
            frame = workbook.parse(sheet_name)
            partner_columns = dataframe_partner_columns(frame)
            date_column = find_column_name(frame, ["NGÀY AIR", "Ngày"])
            channel_column = find_column_name(frame, ["TÊN KÊNH", "Tên Kênh"])
            link_column = find_link_column_name(frame)
            metric_columns = {
                "LƯỢT XEM": find_column_name(frame, ["LƯỢT XEM"]),
                "TIM": find_column_name(frame, ["TIM"]),
                "BÌNH LUẬN": find_column_name(frame, ["BÌNH LUẬN"]),
                "LƯỢT LƯU": find_column_name(frame, ["LƯỢT LƯU"]),
                "CHIA SẺ": find_column_name(frame, ["CHIA SẺ"]),
            }

            if not link_column:
                continue
            frame = fill_missing_dates_from_previous(frame, date_column, link_column)

            for _, row in frame.iterrows():
                partners = extract_row_partners(row, partner_columns) if partner_columns else []
                if selected_key and selected_key not in {partner.casefold() for partner in partners}:
                    continue
                if selected_key and not partners:
                    continue

                link = normalize_tiktok_url(row.get(link_column, ""))
                if not link or is_total_label(link) or not is_tiktok_link(link):
                    continue

                rows.append({
                    "sheet_name": sheet_name,
                    "NGÀY AIR": row.get(date_column, "") if date_column else "",
                    "TÊN KÊNH": display_channel_name_from_file(
                        link,
                        clean_text(row.get(channel_column, "")) if channel_column else "",
                    ),
                    "LINK AIR": link,
                    "LƯỢT XEM": row.get(metric_columns["LƯỢT XEM"], "") if metric_columns["LƯỢT XEM"] else "",
                    "TIM": row.get(metric_columns["TIM"], "") if metric_columns["TIM"] else "",
                    "BÌNH LUẬN": row.get(metric_columns["BÌNH LUẬN"], "") if metric_columns["BÌNH LUẬN"] else "",
                    "LƯỢT LƯU": row.get(metric_columns["LƯỢT LƯU"], "") if metric_columns["LƯỢT LƯU"] else "",
                    "CHIA SẺ": row.get(metric_columns["CHIA SẺ"], "") if metric_columns["CHIA SẺ"] else "",
                    "partners": partners,
                })

        return rows
    finally:
        workbook.close()


def is_exportable_report_row(row, *, apply_min_views=True, min_views=100):
    if is_failed_channel_name(row.get("TÊN KÊNH", "")):
        return False
    if apply_min_views:
        threshold = max(int(min_views or 0), 0)
        if to_number(row.get("LƯỢT XEM", 0)) < threshold:
            return False
    return True


def list_workbook_partners_with_link_counts(
    file_path,
    sheet_name=None,
    *,
    apply_min_views=True,
    min_views=100,
):
    partner_stats = {}
    workbook = load_excel_file(file_path)
    try:
        data_sheets = find_data_sheet_names_in_workbook(workbook)
        requested_sheet = clean_text(sheet_name)
        if requested_sheet:
            data_sheets = [requested_sheet] if requested_sheet in data_sheets else []
        for current_sheet in data_sheets:
            frame = workbook.parse(current_sheet)
            partner_columns = dataframe_partner_columns(frame)
            if not partner_columns:
                continue
            for _, row in frame.iterrows():
                for partner in extract_row_partners(row, partner_columns):
                    key = partner_dedup_key(partner)
                    if key not in partner_stats:
                        partner_stats[key] = {"name": partner, "linkCount": 0, "rawLinkCount": 0}
    finally:
        workbook.close()

    if not partner_stats:
        return []

    for row in build_workbook_rows(file_path, sheet_name=sheet_name):
        row_partners = row.get("partners") or []
        for partner in row_partners:
            key = partner_dedup_key(partner)
            if key not in partner_stats:
                continue
            partner_stats[key]["rawLinkCount"] += 1
            if is_exportable_report_row(row, apply_min_views=apply_min_views, min_views=min_views):
                partner_stats[key]["linkCount"] += 1

    return sorted(partner_stats.values(), key=lambda item: item["name"].casefold())


def list_workbook_partners(file_path, sheet_name=None):
    return [item["name"] for item in list_workbook_partners_with_link_counts(file_path, sheet_name=sheet_name)]


def worksheet_headers(worksheet):
    max_column = worksheet.max_column or 0
    if max_column < 1:
        return []
    return [worksheet.cell(row=1, column=index).value for index in range(1, max_column + 1)]


def worksheet_find_column_index(worksheet, aliases):
    alias_set = {normalize_key(alias) for alias in aliases}
    for index, header in enumerate(worksheet_headers(worksheet), start=1):
        if normalize_key(header) in alias_set:
            return index
    return None


def worksheet_find_link_column_index(worksheet):
    column_index = worksheet_find_column_index(worksheet, ["LINK AIR", "Link", "URL"])
    if column_index:
        return column_index

    for index, header in enumerate(worksheet_headers(worksheet), start=1):
        key = normalize_key(header)
        if "link" in key or "url" in key:
            return index
    return None


def worksheet_find_last_update_column_index(worksheet):
    return worksheet_find_column_index(worksheet, [LAST_UPDATE_COLUMN, "Ngày cập nhật", "Ngay cap nhat"])


def worksheet_has_link_column(worksheet):
    if worksheet_find_link_column_index(worksheet):
        return True

    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0
    for row_index in range(2, min(max_row, 25) + 1):
        for column_index in range(1, max_column + 1):
            value = clean_text(worksheet.cell(row=row_index, column=column_index).value)
            if "tiktok.com" in value or "vt.tiktok.com" in value:
                return True
    return False


def workbook_data_sheet_names(workbook):
    return [
        sheet for sheet in workbook.sheetnames
        if not is_summary_sheet_name(sheet) and not is_result_sheet_name(sheet)
    ]


def result_sheet_display_name(timestamp_text=None):
    stamp = clean_text(timestamp_text) or format_excel_sheet_datetime()
    return stamp[:31]


def worksheet_partner_column_indexes(worksheet):
    headers = worksheet_headers(worksheet)
    start_index = None
    for index, header in enumerate(headers, start=1):
        if normalize_key(header).startswith("doi tac"):
            start_index = index
            break

    if start_index is None:
        return []

    indexes = [start_index]
    for index in range(start_index + 1, len(headers) + 1):
        header_key = normalize_key(headers[index - 1])
        raw_header_key = normalize_header(headers[index - 1])
        if header_key.startswith("doi tac") or not raw_header_key or raw_header_key.startswith("unnamed:"):
            indexes.append(index)
        else:
            break
    return indexes


def worksheet_row_partners(worksheet, row_index, partner_columns):
    values = {column_index: worksheet.cell(row=row_index, column=column_index).value for column_index in partner_columns}
    partners = []
    for value in values.values():
        partners.extend(split_partner_value(value))
    return unique_preserve_order(partners)


def is_failed_channel_name(value):
    text = clean_text(value)
    key = normalize_key(text)
    normalized_for_error = text.casefold().replace("\u2019", "'").replace("\u2018", "'")
    error_phrases = (
        "couldn't find this account",
        "couldnt find this account",
        "video currently unavailable",
        "video unavailable",
        "page not available",
        "page unavailable",
    )
    return (
        not text
        or key in {"loi", "l?i"}
        or text.casefold().startswith("error:")
        or any(phrase in normalized_for_error for phrase in error_phrases)
        or is_numeric_channel_garbage(text)
        or is_generated_username_channel(text)
        or is_generic_tiktok_channel_name(text)
    )


def read_existing_summary_updates(worksheet):
    updates = {}
    if not worksheet or (worksheet.max_row or 0) < 2:
        return updates
    partner_column = worksheet_find_column_index(worksheet, ["ĐỐI TÁC", "Đối tác"])
    update_column = worksheet_find_column_index(worksheet, [LAST_UPDATE_COLUMN, "Ngày cập nhật", "Ngay cap nhat"])
    if not partner_column or not update_column:
        return updates
    for row_index in range(2, (worksheet.max_row or 0) + 1):
        partner = clean_text(worksheet.cell(row=row_index, column=partner_column).value)
        update_value = clean_text(worksheet.cell(row=row_index, column=update_column).value)
        if partner and update_value:
            updates[partner.casefold()] = update_value
    return updates


def normalize_selected_partner_keys(selected_partner=None, selected_partners=None):
    values = []
    if isinstance(selected_partners, (list, tuple, set)):
        values.extend(selected_partners)
    elif selected_partners:
        values.append(selected_partners)
    if isinstance(selected_partner, (list, tuple, set)):
        values.extend(selected_partner)
    elif selected_partner:
        values.append(selected_partner)
    return {clean_text(value).casefold() for value in values if clean_text(value)}


def build_partner_summary_rows(
    workbook,
    summary_update_time="",
    selected_partner=None,
    selected_partners=None,
    previous_updates=None,
    data_sheet_name=None,
):
    selected_keys = normalize_selected_partner_keys(selected_partner, selected_partners)
    previous_updates = previous_updates or {}
    summary = {}
    data_sheets = workbook_data_sheet_names(workbook)
    requested_sheet = clean_text(data_sheet_name)
    if requested_sheet:
        data_sheets = [requested_sheet] if requested_sheet in workbook.sheetnames else []
    for sheet_name in data_sheets:
        worksheet = workbook[sheet_name]
        link_column = worksheet_find_link_column_index(worksheet)
        if not link_column:
            continue

        partner_columns = worksheet_partner_column_indexes(worksheet)
        if not partner_columns:
            continue

        metric_columns = {
            header: worksheet_find_column_index(worksheet, [header])
            for header in METRIC_COLUMNS
        }
        last_update_column = worksheet_find_last_update_column_index(worksheet)

        for row_index in range(2, (worksheet.max_row or 0) + 1):
            link = clean_text(worksheet.cell(row=row_index, column=link_column).value)
            if not link or ("tiktok.com" not in link and "vt.tiktok.com" not in link):
                continue

            partners = worksheet_row_partners(worksheet, row_index, partner_columns)
            if not partners:
                continue

            for partner in partners:
                bucket = summary.setdefault(
                    partner,
                    {
                        "ĐỐI TÁC": partner,
                        "TỔNG LINK": 0,
                        "TỔNG LƯỢT XEM": 0,
                        "TỔNG TIM": 0,
                        "TỔNG BÌNH LUẬN": 0,
                        "TỔNG LƯỢT LƯU": 0,
                        "TỔNG CHIA SẺ": 0,
                        LAST_UPDATE_COLUMN: previous_updates.get(partner.casefold(), ""),
                    },
                )
                bucket["TỔNG LINK"] += 1
                if last_update_column:
                    update_value = clean_text(worksheet.cell(row=row_index, column=last_update_column).value)
                    if update_value and update_value > clean_text(bucket.get(LAST_UPDATE_COLUMN, "")):
                        bucket[LAST_UPDATE_COLUMN] = update_value
                for metric, column_index in metric_columns.items():
                    if not column_index:
                        continue
                    value = to_number(worksheet.cell(row=row_index, column=column_index).value)
                    if metric == "LƯỢT XEM":
                        bucket["TỔNG LƯỢT XEM"] += value
                    elif metric == "TIM":
                        bucket["TỔNG TIM"] += value
                    elif metric == "BÌNH LUẬN":
                        bucket["TỔNG BÌNH LUẬN"] += value
                    elif metric == "LƯỢT LƯU":
                        bucket["TỔNG LƯỢT LƯU"] += value
                    elif metric == "CHIA SẺ":
                        bucket["TỔNG CHIA SẺ"] += value

    result = []
    if summary_update_time:
        for partner, row in summary.items():
            if not selected_keys or partner.casefold() in selected_keys:
                row[LAST_UPDATE_COLUMN] = summary_update_time

    for index, name in enumerate(sorted(summary, key=lambda value: value.casefold()), start=1):
        row = summary[name]
        row["Stt"] = index
        result.append(row)
    return result


def _summary_row_metric_sums(partner_rows):
    sums = {}
    for row in partner_rows:
        for column, value in row.items():
            column_key = normalize_key(column)
            if column_key in {"stt", normalize_key(LAST_UPDATE_COLUMN)} or column_key.startswith("doi tac"):
                continue
            sums[column_key] = sums.get(column_key, 0) + to_number(value)
    return sums


def build_summary_totals_row(partner_rows):
    metric_sums = _summary_row_metric_sums(partner_rows)
    totals = {
        "Stt": "",
        "ĐỐI TÁC": SUMMARY_TOTAL_LABEL,
        "TỔNG LINK": 0,
        "TỔNG LƯỢT XEM": 0,
        "TỔNG TIM": 0,
        "TỔNG BÌNH LUẬN": 0,
        "TỔNG LƯỢT LƯU": 0,
        "TỔNG CHIA SẺ": 0,
        LAST_UPDATE_COLUMN: "",
    }
    for header, value in totals.items():
        header_key = normalize_key(header)
        if header_key in metric_sums:
            totals[header] = metric_sums[header_key]
    return totals


def build_summary_totals_row_aligned(columns, partner_rows):
    metric_sums = _summary_row_metric_sums(partner_rows)
    aligned = {}
    for column in columns:
        column_key = normalize_key(column)
        if column_key == "stt":
            aligned[column] = ""
        elif column_key.startswith("doi tac"):
            aligned[column] = SUMMARY_TOTAL_LABEL
        elif column_key == normalize_key(LAST_UPDATE_COLUMN):
            aligned[column] = ""
        else:
            aligned[column] = metric_sums.get(column_key, 0)
    return aligned


def rebuild_summary_sheet(
    workbook,
    summary_update_time="",
    selected_partner=None,
    selected_partners=None,
    data_sheet_name=None,
):
    source_sheet = clean_text(data_sheet_name)
    if not source_sheet or source_sheet not in workbook.sheetnames:
        return 0

    summary_sheet = summary_sheet_title_for_data_sheet(source_sheet)
    existing_worksheet = workbook[summary_sheet] if summary_sheet in workbook.sheetnames else None
    previous_updates = read_existing_summary_updates(existing_worksheet)
    insert_index = workbook.sheetnames.index(source_sheet) + 1
    if summary_sheet in workbook.sheetnames:
        worksheet = workbook[summary_sheet]
        worksheet.delete_rows(1, max(worksheet.max_row or 1, 1))
        if worksheet.title != summary_sheet:
            worksheet.title = summary_sheet
        current_index = workbook.sheetnames.index(summary_sheet)
        if current_index != insert_index:
            workbook.move_sheet(worksheet, offset=insert_index - current_index)
    else:
        worksheet = workbook.create_sheet(summary_sheet, insert_index)

    rows = build_partner_summary_rows(
        workbook,
        summary_update_time=summary_update_time,
        selected_partner=selected_partner,
        selected_partners=selected_partners,
        previous_updates=previous_updates,
        data_sheet_name=source_sheet,
    )
    header_fill = PatternFill("solid", fgColor="0B5ED7")
    for column_index, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(SUMMARY_COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row.get(header, ""))
            if header == "ĐỐI TÁC":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif header == LAST_UPDATE_COLUMN:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    last_row = max(len(rows) + 1, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{last_row}"
    widths = [8, 38, 12, 16, 12, 16, 16, 14, 20]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    return len(rows)


def _cell_has_single_link_fill(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    return str(fill.fgColor.rgb or "").upper().endswith(SINGLE_LINK_FILL_COLOR)


def highlight_single_partner_link_rows(workbook, data_sheet_name):
    """Bôi cam các dòng có link TikTok mà dòng đó chỉ gắn đúng 1 đối tác (link độc quyền).

    Dọn sạch highlight cam cũ (chỉ xóa đúng màu SINGLE_LINK_FILL_COLOR) trên các
    dòng không còn thỏa điều kiện (ví dụ dòng vừa được gắn thêm đối tác thứ 2),
    để lần quét sau không để sót màu cam.
    """
    source_sheet = clean_text(data_sheet_name)
    if not source_sheet or source_sheet not in workbook.sheetnames:
        return 0

    worksheet = workbook[source_sheet]
    link_column = worksheet_find_link_column_index(worksheet)
    if not link_column:
        return 0

    partner_columns = worksheet_partner_column_indexes(worksheet)
    if not partner_columns:
        return 0

    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0
    if max_row < 2 or max_column < 1:
        return 0

    highlight_fill = PatternFill("solid", fgColor=SINGLE_LINK_FILL_COLOR)
    clear_fill = PatternFill(fill_type=None)
    highlighted_count = 0
    for row_index in range(2, max_row + 1):
        link = clean_text(worksheet.cell(row=row_index, column=link_column).value)
        should_highlight = False
        if link and ("tiktok.com" in link or "vt.tiktok.com" in link):
            partners = worksheet_row_partners(worksheet, row_index, partner_columns)
            should_highlight = len(partners) == 1
        if should_highlight:
            highlighted_count += 1
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if should_highlight:
                cell.fill = highlight_fill
            elif _cell_has_single_link_fill(cell):
                cell.fill = clear_fill

    return highlighted_count


def workbook_file_entries(base_dir):
    data_dir = ensure_data_dir(base_dir)
    entries = []

    for filename in os.listdir(base_dir):
        if is_internal_workbook_filename(filename):
            continue
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            entries.append({
                "id": filename.replace("\\", "/"),
                "label": filename,
                "source": "local",
            })

    registry = load_google_sheet_registry(base_dir)
    for filename in os.listdir(data_dir):
        if is_internal_workbook_filename(filename):
            continue
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            relative_id = f"{DATA_DIR_NAME}/{filename}".replace("\\", "/")
            source = registry.get(relative_id, {})
            if isinstance(source, dict) and clean_text(source.get("title")):
                label = clean_text(source["title"])
            else:
                label = google_sheet_filename_to_label(filename)
            if relative_id == LEGACY_GOOGLE_SHEET_FILE_ID:
                label = GOOGLE_SHEET_LABEL
            entries.append({
                "id": relative_id,
                "label": label,
                "source": "google",
            })

    return sorted(entries, key=lambda entry: (entry["source"] != "google", entry["label"].casefold()))
