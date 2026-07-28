import asyncio
import html
import json
import os
import random
import re
import threading
import time
import unicodedata
import urllib.error
import urllib.request
from datetime import datetime

import openpyxl
from playwright.async_api import async_playwright

from workbook_utils import (
    clean_text,
    highlight_single_partner_link_rows,
    highlight_video_link_rows,
    is_generated_username_channel,
    is_generic_tiktok_channel_name,
    is_numeric_channel_garbage,
    is_scrapable_tiktok_url,
    should_highlight_video_link,
    load_channel_overrides,
    metric_number,
    format_display_datetime,
    format_excel_sheet_datetime,
    normalize_tiktok_url,
    rebuild_summary_sheet,
    resolve_channel_name,
    result_sheet_display_name,
    summary_sheet_title_for_data_sheet,
    TTBD_RESOLVED_URL_HEADER,
    TTBD_SCAN_STATUS_HEADER,
    TTBD_SOURCE_URL_HEADER,
    workbook_data_sheet_names,
    worksheet_partner_column_indexes,
    worksheet_row_partners,
)
from proxy_utils import (
    assign_worker_proxy,
    get_session_proxies,
    playwright_proxy_settings,
    proxy_label,
    release_thread_proxy,
    resolve_proxy_configs,
    set_session_proxies,
    urlopen_request,
)


USER_AGENTS = [
    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPad; CPU OS 16_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Mobile/15E148 Safari/604.1",
]
DEFAULT_USER_AGENT = USER_AGENTS[0]
BROWSER_LOCALE = "vi-VN"
BROWSER_TIMEZONE = "Asia/Ho_Chi_Minh"
BROWSER_GEOLOCATION = {"latitude": 11.9404, "longitude": 108.4583}

METRIC_HEADERS = {
    "views": "LƯỢT XEM",
    "likes": "TIM",
    "comments": "BÌNH LUẬN",
    "saves": "LƯỢT LƯU",
    "shares": "CHIA SẺ",
}

METRIC_KEYS = {
    "views": "Views",
    "likes": "Likes",
    "comments": "Comments",
    "saves": "Saves",
    "shares": "Shares",
}

LAST_UPDATE_HEADER = "Cập nhật lần cuối"

COUNT_FIELD_MAP = {
    "Views": "playCount",
    "Likes": "diggCount",
    "Comments": "commentCount",
    "Saves": "collectCount",
    "Shares": "shareCount",
}

COUNT_PATTERNS = {
    metric: rf'"{field}"\s*:\s*"?(\d+)"?'
    for metric, field in COUNT_FIELD_MAP.items()
}

BLOCKED_RESOURCE_TYPES = {"image", "media", "font", "texttrack"}
UNIVERSAL_DETAIL_KEY_MARKERS = (
    "video-detail",
    "photo-detail",
    "photomode",
    "image-detail",
    "reflow.video.detail",
    "reflow.photo.detail",
)
EMBEDDED_STATE_SCRIPT_PATTERNS = (
    r"<script\b[^>]*\bid\s*=\s*['\"]__UNIVERSAL_DATA_FOR_REHYDRATION__['\"][^>]*>(.*?)</script\s*>",
    r"<script\b[^>]*\bid\s*=\s*['\"]SIGI_STATE['\"][^>]*>(.*?)</script\s*>",
    r"<script\b[^>]*\bid\s*=\s*['\"]api-data['\"][^>]*>(.*?)</script\s*>",
)

MAX_CONCURRENT_REQUESTS = 20
DIRECT_MAX_WORKERS = 10
MAX_WORKERS_PER_PROXY = 10
REQUEST_SHELL_RETRY_DELAY = 0.35
REQUEST_SHELL_RETRY_DELAY_PROXY = 0.12
REQUEST_CANDIDATE_LIMIT_DIRECT = 4
REQUEST_CANDIDATE_LIMIT_PROXY = 3
REQUEST_METRIC_HINT_PATTERN = re.compile(r'"(?:playCount|diggCount)"\s*:\s*"?(\d+)"?')
_request_semaphore = threading.Semaphore(MAX_CONCURRENT_REQUESTS)
_request_block_lock = threading.Lock()
_request_block_until = 0.0
_network_fail_streak = 0
NETWORK_FAIL_STREAK_PAUSE = 3
NETWORK_FAIL_PAUSE_SECONDS = 5.0


def session_uses_proxy():
    return bool(get_session_proxies())


def clamp_worker_count(worker_count, proxy_count=0):
    """Chỉ bó số luồng vào khoảng hợp lệ (1..MAX_WORKERS).

    Không tự giảm luồng dù không có proxy hay ít proxy — luôn chạy đúng số
    luồng người dùng chọn. proxy_count hiện chỉ dùng cho log/cảnh báo, không
    còn ảnh hưởng tới số luồng thực chạy.
    """
    return clamp_int(worker_count, DEFAULT_WORKERS, 1, MAX_WORKERS)


def configure_request_concurrency(worker_count, proxy_count=0):
    global _request_semaphore, _network_fail_streak, _request_block_until
    limit = clamp_worker_count(worker_count, proxy_count=proxy_count)
    _request_semaphore = threading.Semaphore(limit)
    with _request_block_lock:
        _network_fail_streak = 0
        _request_block_until = 0.0


def is_request_rate_limited_status(status):
    text = str(status or "")
    return "HTTP 403" in text or "HTTP 429" in text


def is_transient_network_status(status):
    """DNS / connection reset / timeout / SSL đứt giữa chừng — lỗi tạm, nên retry đủ số lần."""
    text = str(status or "").casefold()
    markers = (
        "getaddrinfo failed",
        "errno 11001",
        "winerror 10054",
        "connection was forcibly closed",
        "connection reset",
        "timed out",
        "timeout",
        "temporarily unavailable",
        "name or service not known",
        "nodename nor servname",
        "network is unreachable",
        "connection aborted",
        "broken pipe",
        "unexpected_eof_while_reading",
        "eof occurred in violation of protocol",
        "ssl:",
        "sslerror",
        "wrong version number",
        "connection refused",
        "remote end closed connection",
        "incomplete read",
    )
    return any(marker in text for marker in markers)


def wait_if_request_blocked():
    with _request_block_lock:
        until = _request_block_until
    remaining = until - time.time()
    if remaining > 0:
        time.sleep(remaining)


def note_request_rate_limit(http_code):
    global _request_block_until
    if http_code not in (403, 429) or session_uses_proxy():
        return
    pause = 8.0 if http_code == 403 else 4.0
    with _request_block_lock:
        _request_block_until = max(_request_block_until, time.time() + pause)


def note_network_failure():
    """Tạm dừng toàn phiên khi DNS/mạng fail hàng loạt (tránh đốt hết queue)."""
    global _request_block_until, _network_fail_streak
    with _request_block_lock:
        _network_fail_streak += 1
        if _network_fail_streak >= NETWORK_FAIL_STREAK_PAUSE:
            _request_block_until = max(
                _request_block_until,
                time.time() + NETWORK_FAIL_PAUSE_SECONDS,
            )
            _network_fail_streak = 0


def note_network_success():
    global _network_fail_streak
    with _request_block_lock:
        _network_fail_streak = 0


def request_shell_retry_delay():
    return REQUEST_SHELL_RETRY_DELAY_PROXY if session_uses_proxy() else REQUEST_SHELL_RETRY_DELAY


def request_candidate_limit():
    return REQUEST_CANDIDATE_LIMIT_PROXY if session_uses_proxy() else REQUEST_CANDIDATE_LIMIT_DIRECT

# Phân biệt hai loại "không có số liệu":
# - METRICS_UNREADABLE: HTML có dấu hiệu số liệu nhưng đọc lỗi (hiếm, đáng retry).
# - TIKTOK_NO_STATS: TikTok trả trang rỗng, không hề có số liệu (post ẩn view) -> đếm ẨN, không phải BỊ LỖI.
STATUS_METRICS_UNREADABLE = "Error: Không đọc được số liệu"
STATUS_TIKTOK_NO_STATS = "Ẩn số liệu: TikTok không trả lượt xem"
# Chuỗi cũ (trước khi tách ẩn/lỗi) — vẫn nhận diện khi đọc log/Excel cũ.
STATUS_TIKTOK_NO_STATS_LEGACY = "Lỗi: TikTok không trả số liệu"
STATUS_MEDIA_REDIRECT_MISMATCH = "Error: Redirect không khớp video"

MAX_WORKERS = 50
DEFAULT_WORKERS = 5
DEFAULT_RETRIES = 2
DEFAULT_SAVE_EVERY = 25
DEFAULT_REQUEST_TIMEOUT = 30
MAX_BROWSER_FALLBACK_WORKERS = 15
RESULT_SHEET_HEADERS = [
    "Stt",
    "Ngày",
    "Link",
    "Tên Kênh",
    "LƯỢT XEM",
    "TIM",
    "BÌNH LUẬN",
    "LƯỢT LƯU",
    "CHIA SẺ",
    "Đối tác",
    LAST_UPDATE_HEADER,
    TTBD_SCAN_STATUS_HEADER,
    TTBD_RESOLVED_URL_HEADER,
    TTBD_SOURCE_URL_HEADER,
]

SCRAPE_HISTORY_FILENAME = "scrape_history.json"
SCRAPE_HISTORY_LIMIT = 200


def _compute_workbook_totals(workbook, sheet_name=None):
    totals = {
        "totalLinks": 0,
        "totalViews": 0,
        "totalLikes": 0,
        "totalComments": 0,
        "totalSaves": 0,
        "totalShares": 0,
    }
    target_sheet = clean_text(sheet_name)
    if target_sheet and target_sheet in workbook.sheetnames:
        sheet_names = [target_sheet]
    else:
        sheet_names = workbook_data_sheet_names(workbook)
    for current_sheet in sheet_names:
        sheet = workbook[current_sheet]
        columns = detect_columns(sheet)
        url_column = columns.get("url")
        if not url_column:
            continue
        for row_index in range(2, (sheet.max_row or 1) + 1):
            raw_url = sheet.cell(row=row_index, column=url_column).value
            if not is_scrapable_tiktok_url(raw_url):
                continue
            totals["totalLinks"] += 1
            for metric_key, total_key in (
                ("views", "totalViews"),
                ("likes", "totalLikes"),
                ("comments", "totalComments"),
                ("saves", "totalSaves"),
                ("shares", "totalShares"),
            ):
                col = columns.get(metric_key)
                if not col:
                    continue
                totals[total_key] += metric_number(sheet.cell(row=row_index, column=col).value)
    return totals


def _compute_session_totals(workbook, rows_to_process):
    """Sum metrics only for rows included in the current scrape session."""
    totals = {
        "totalLinks": 0,
        "totalViews": 0,
        "totalLikes": 0,
        "totalComments": 0,
        "totalSaves": 0,
        "totalShares": 0,
    }
    column_cache = {}
    for item in rows_to_process:
        sheet_name = item.get("sheet_name")
        row_index = item.get("row")
        if not sheet_name or not row_index:
            continue
        if sheet_name not in workbook.sheetnames:
            continue
        if sheet_name not in column_cache:
            column_cache[sheet_name] = detect_columns(workbook[sheet_name])
        columns = column_cache[sheet_name]
        url_column = columns.get("url")
        if not url_column:
            continue
        raw_url = workbook[sheet_name].cell(row=row_index, column=url_column).value
        if not is_scrapable_tiktok_url(raw_url):
            continue
        totals["totalLinks"] += 1
        for metric_key, total_key in (
            ("views", "totalViews"),
            ("likes", "totalLikes"),
            ("comments", "totalComments"),
            ("saves", "totalSaves"),
            ("shares", "totalShares"),
        ):
            col = columns.get(metric_key)
            if not col:
                continue
            totals[total_key] += metric_number(
                workbook[sheet_name].cell(row=row_index, column=col).value
            )
    return totals


def scrape_history_path(base_dir):
    data_dir = os.path.join(base_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, SCRAPE_HISTORY_FILENAME)


def append_scrape_history(base_dir, entry):
    if not base_dir:
        return
    path = scrape_history_path(base_dir)
    history = []
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as file_obj:
                payload = json.load(file_obj)
            if isinstance(payload, dict):
                history = list(payload.get("history") or [])
        except (OSError, json.JSONDecodeError):
            history = []
    history.insert(0, entry)
    history = history[:SCRAPE_HISTORY_LIMIT]
    try:
        with open(path, "w", encoding="utf-8") as file_obj:
            json.dump({"history": history}, file_obj, ensure_ascii=False, indent=2)
    except OSError:
        pass


def read_scrape_history(base_dir, limit=50):
    path = scrape_history_path(base_dir)
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except (OSError, json.JSONDecodeError):
        return []
    history = payload.get("history") if isinstance(payload, dict) else None
    if not isinstance(history, list):
        return []
    return history[:limit]


def normalize_selected_partners(selected_partner=None, selected_partners=None):
    values = []
    if isinstance(selected_partners, (list, tuple, set)):
        values.extend(selected_partners)
    elif selected_partners:
        values.append(selected_partners)
    if isinstance(selected_partner, (list, tuple, set)):
        values.extend(selected_partner)
    elif selected_partner:
        values.append(selected_partner)

    partners = []
    seen = set()
    for value in values:
        name = clean_text(value)
        key = name.casefold()
        if name and key not in seen:
            partners.append(name)
            seen.add(key)
    return partners


def selected_partner_label(partners):
    if not partners:
        return ""
    if len(partners) == 1:
        return partners[0]
    return f"{len(partners)} đối tác"


def clamp_int(value, default, minimum, maximum):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))


def normalize_text(value):
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text)


def selected_data_sheet_names(workbook, sheet_name=None):
    target_sheet = clean_text(sheet_name)
    if target_sheet:
        return [target_sheet] if target_sheet in workbook.sheetnames else []
    return workbook_data_sheet_names(workbook)


def build_sheet_contexts(workbook, sheet_name=None):
    contexts = {}
    for sheet_name in selected_data_sheet_names(workbook, sheet_name):
        sheet = workbook[sheet_name]
        contexts[sheet_name] = {
            "worksheet": sheet,
            "columns": ensure_columns(sheet),
        }
    return contexts


def detect_columns(sheet):
    column_map = {key: None for key in METRIC_HEADERS}
    column_map["url"] = None
    column_map["channel"] = None
    column_map["date"] = None
    column_map["last_update"] = None
    column_map["scan_status"] = None
    column_map["resolved_url"] = None
    column_map["source_url"] = None

    for cell in sheet[1]:
        header = normalize_text(cell.value)
        if not header:
            continue
        if header == normalize_text(TTBD_SCAN_STATUS_HEADER):
            column_map["scan_status"] = cell.column
            continue
        if header == normalize_text(TTBD_RESOLVED_URL_HEADER):
            column_map["resolved_url"] = cell.column
            continue
        if header == normalize_text(TTBD_SOURCE_URL_HEADER):
            column_map["source_url"] = cell.column
            continue
        if "URL" in header or "LINK" in header:
            column_map["url"] = cell.column
        if "TÊN KÊNH" in header or "TEN KENH" in header:
            column_map["channel"] = cell.column
        if "NGÀY" in header or "NGAY" in header:
            column_map["date"] = cell.column
        if "CẬP NHẬT LẦN CUỐI" in header or "CAP NHAT LAN CUOI" in header:
            column_map["last_update"] = cell.column
        for key, expected_header in METRIC_HEADERS.items():
            if normalize_text(expected_header) in header:
                column_map[key] = cell.column

    max_column = sheet.max_column or 0
    if not column_map["url"]:
        for row in range(2, min(sheet.max_row or 1, 25) + 1):
            for col in range(1, max_column + 1):
                value = str(sheet.cell(row=row, column=col).value or "")
                if "tiktok.com" in value or "vt.tiktok.com" in value:
                    column_map["url"] = col
                    break
            if column_map["url"]:
                break

    for key, fallback in {"views": 5, "likes": 6, "comments": 7, "saves": 8, "shares": 9}.items():
        if not column_map[key] and max_column >= fallback and normalize_text(sheet.cell(row=1, column=fallback).value):
            column_map[key] = fallback

    return column_map


def ensure_columns(sheet):
    column_map = detect_columns(sheet)
    max_column = sheet.max_column or 0

    if not column_map["channel"]:
        channel_col = max_column + 1
        sheet.cell(row=1, column=channel_col).value = "Tên Kênh"
        column_map["channel"] = channel_col
        max_column = channel_col

    for key, header in METRIC_HEADERS.items():
        if not column_map[key]:
            max_column += 1
            sheet.cell(row=1, column=max_column).value = header
            column_map[key] = max_column

    if not column_map["last_update"]:
        max_column += 1
        sheet.cell(row=1, column=max_column).value = LAST_UPDATE_HEADER
        column_map["last_update"] = max_column

    for key, header in (
        ("scan_status", TTBD_SCAN_STATUS_HEADER),
        ("resolved_url", TTBD_RESOLVED_URL_HEADER),
        ("source_url", TTBD_SOURCE_URL_HEADER),
    ):
        if not column_map[key]:
            max_column += 1
            sheet.cell(row=1, column=max_column).value = header
            column_map[key] = max_column
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column_map[key])].hidden = True

    return column_map


def find_columns(sheet):
    return ensure_columns(sheet)


def collect_rows(workbook, selected_partner=None, selected_partners=None, sheet_name=None):
    selected_names = normalize_selected_partners(selected_partner, selected_partners)
    selected_keys = {partner.casefold() for partner in selected_names}
    rows = []

    for sheet_name in selected_data_sheet_names(workbook, sheet_name):
        worksheet = workbook[sheet_name]
        partner_columns = worksheet_partner_column_indexes(worksheet)
        columns = detect_columns(worksheet)
        url_column = columns["url"]
        if not url_column:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            raw_url = worksheet.cell(row=row_index, column=url_column).value
            if not is_scrapable_tiktok_url(raw_url):
                continue
            url = normalize_tiktok_url(raw_url)
            stored_source_url = normalize_tiktok_url(
                worksheet.cell(row=row_index, column=columns["source_url"]).value
            ) if columns.get("source_url") else ""
            stored_resolved_url = normalize_tiktok_url(
                worksheet.cell(row=row_index, column=columns["resolved_url"]).value
            ) if columns.get("resolved_url") else ""
            metadata_matches_source = (
                not stored_source_url
                or stored_source_url.casefold() == url.casefold()
            )
            expected_media_id = extract_media_id(stored_resolved_url) if metadata_matches_source else ""

            partners = []
            if partner_columns:
                partners = worksheet_row_partners(worksheet, row_index, partner_columns)

            if selected_keys:
                if not partners:
                    continue
                if not selected_keys.intersection({partner.casefold() for partner in partners}):
                    continue

            rows.append({
                "sequence": len(rows) + 1,
                "sheet_name": sheet_name,
                "row": row_index,
                "url": url,
                "partners": partners,
                "expected_media_id": expected_media_id,
            })

    return rows


def is_total_row(sheet, row_index, url_column):
    value = clean_text(sheet.cell(row=row_index, column=url_column).value)
    return normalize_text(value) == "TỔNG" or normalize_text(value) == "TONG"


def clear_existing_total_rows(workbook):
    for sheet_name in workbook_data_sheet_names(workbook):
        sheet = workbook[sheet_name]
        url_column = detect_columns(sheet)["url"]
        if not url_column:
            continue
        for row_index in range(sheet.max_row, 1, -1):
            if is_total_row(sheet, row_index, url_column):
                sheet.delete_rows(row_index, 1)


def append_sheet_total_rows(workbook):
    for sheet_name in workbook_data_sheet_names(workbook):
        sheet = workbook[sheet_name]
        columns = ensure_columns(sheet)
        url_column = columns.get("url")
        if not url_column:
            continue

        last_link_row = 1
        for row_index in range(2, sheet.max_row + 1):
            url = clean_text(sheet.cell(row=row_index, column=url_column).value)
            if "tiktok.com" in url or "vt.tiktok.com" in url:
                last_link_row = row_index
        if last_link_row < 2:
            continue

        total_row = last_link_row + 1
        sheet.cell(row=total_row, column=url_column).value = "TỔNG"
        for metric_key in METRIC_HEADERS:
            column_index = columns.get(metric_key)
            if column_index:
                col_letter = openpyxl.utils.get_column_letter(column_index)
                sheet.cell(row=total_row, column=column_index).value = f"=SUM({col_letter}2:{col_letter}{last_link_row})"


def build_result_sheet(workbook, rows_to_process, summary_update_time):
    sheet_name = result_sheet_display_name(format_excel_sheet_datetime())
    if sheet_name in workbook.sheetnames:
        base_name = sheet_name[:28]
        suffix = 2
        while f"{base_name}-{suffix}" in workbook.sheetnames:
            suffix += 1
        sheet_name = f"{base_name}-{suffix}"

    worksheet = workbook.create_sheet(title=sheet_name)
    for column_index, header in enumerate(RESULT_SHEET_HEADERS, start=1):
        worksheet.cell(row=1, column=column_index).value = header

    row_lookup = {}
    for item in rows_to_process:
        row_lookup[(item["sheet_name"], item["row"])] = item

    output_row = 2
    sequence = 1
    for source_sheet_name in workbook_data_sheet_names(workbook):
        source_sheet = workbook[source_sheet_name]
        columns = ensure_columns(source_sheet)
        url_column = columns.get("url")
        if not url_column:
            continue

        partner_columns = worksheet_partner_column_indexes(source_sheet)
        for source_row_index in range(2, source_sheet.max_row + 1):
            item = row_lookup.get((source_sheet_name, source_row_index))
            if not item:
                continue

            url = clean_text(source_sheet.cell(row=source_row_index, column=url_column).value)
            if not url or is_total_row(source_sheet, source_row_index, url_column):
                continue

            partner_names = worksheet_row_partners(source_sheet, source_row_index, partner_columns) if partner_columns else []
            worksheet.cell(row=output_row, column=1).value = sequence
            worksheet.cell(row=output_row, column=2).value = source_sheet.cell(row=source_row_index, column=columns.get("date") or 1).value if columns.get("date") else ""
            worksheet.cell(row=output_row, column=3).value = url
            worksheet.cell(row=output_row, column=4).value = clean_text(source_sheet.cell(row=source_row_index, column=columns.get("channel")).value) if columns.get("channel") else ""
            worksheet.cell(row=output_row, column=5).value = int(source_sheet.cell(row=source_row_index, column=columns.get("views")).value or 0) if columns.get("views") else 0
            worksheet.cell(row=output_row, column=6).value = int(source_sheet.cell(row=source_row_index, column=columns.get("likes")).value or 0) if columns.get("likes") else 0
            worksheet.cell(row=output_row, column=7).value = int(source_sheet.cell(row=source_row_index, column=columns.get("comments")).value or 0) if columns.get("comments") else 0
            worksheet.cell(row=output_row, column=8).value = int(source_sheet.cell(row=source_row_index, column=columns.get("saves")).value or 0) if columns.get("saves") else 0
            worksheet.cell(row=output_row, column=9).value = int(source_sheet.cell(row=source_row_index, column=columns.get("shares")).value or 0) if columns.get("shares") else 0
            worksheet.cell(row=output_row, column=10).value = "\n".join(partner_names)
            worksheet.cell(row=output_row, column=11).value = clean_text(source_sheet.cell(row=source_row_index, column=columns.get("last_update")).value) if columns.get("last_update") else summary_update_time
            worksheet.cell(row=output_row, column=12).value = clean_text(source_sheet.cell(row=source_row_index, column=columns.get("scan_status")).value) if columns.get("scan_status") else ""
            worksheet.cell(row=output_row, column=13).value = clean_text(source_sheet.cell(row=source_row_index, column=columns.get("resolved_url")).value) if columns.get("resolved_url") else ""
            worksheet.cell(row=output_row, column=14).value = clean_text(source_sheet.cell(row=source_row_index, column=columns.get("source_url")).value) if columns.get("source_url") else ""
            output_row += 1
            sequence += 1

    widths = [10, 16, 72, 24, 14, 12, 14, 14, 12, 28, 20]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    if output_row > 2:
        total_row = output_row
        worksheet.cell(row=total_row, column=3).value = "TỔNG"
        for column_index in range(5, 10):
            letter = openpyxl.utils.get_column_letter(column_index)
            worksheet.cell(row=total_row, column=column_index).value = f"=SUM({letter}2:{letter}{total_row - 1})"

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(12)].hidden = True
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(13)].hidden = True
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(14)].hidden = True
    return sheet_name


def extract_media_id(url):
    text = clean_text(url)
    for pattern in (r"/video/(\d+)", r"/photo/(\d+)"):
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def parse_count_value(raw):
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        number = float(raw)
        return str(int(number)) if number.is_integer() else str(int(number))

    text = clean_text(str(raw)).replace(",", "").replace(" ", "")
    if not text:
        return None
    if re.fullmatch(r"\d+", text):
        return text

    match = re.fullmatch(r"(\d+(?:\.\d+)?)([KMB])", text, flags=re.IGNORECASE)
    if match:
        number = float(match.group(1))
        suffix = match.group(2).upper()
        multiplier = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}[suffix]
        return str(int(number * multiplier))

    if re.fullmatch(r"\d+\.0+", text):
        return str(int(float(text)))
    return None


def stats_dict_to_metrics(stats):
    if not isinstance(stats, dict):
        return None
    return metrics_from_stats_sources(stats)


def metrics_from_stats_sources(*sources):
    """Read metrics from stats/statsV2 blobs and take the max per field."""
    field_values = {metric: [] for metric in COUNT_FIELD_MAP}

    def collect_from_dict(obj):
        if not isinstance(obj, dict):
            return
        blobs = []
        if isinstance(obj.get("stats"), dict):
            blobs.append(obj["stats"])
        if isinstance(obj.get("statsV2"), dict):
            blobs.append(obj["statsV2"])
        if not blobs:
            blobs.append(obj)
        for blob in blobs:
            if not isinstance(blob, dict):
                continue
            for metric, field in COUNT_FIELD_MAP.items():
                parsed = parse_count_value(blob.get(field))
                if parsed is not None:
                    field_values[metric].append(int(parsed))

    for source in sources:
        collect_from_dict(source)

    metrics = {}
    for metric in COUNT_FIELD_MAP:
        if not field_values[metric]:
            if metric == "Saves":
                metrics[metric] = "0"
                continue
            return None
        metrics[metric] = str(max(field_values[metric]))
    return metrics


def extract_raw_stats_fields(content, media_id=""):
    """Diagnostic helper: expose raw stats/statsV2 playCount values for probe tooling."""
    raw = {"stats": None, "statsV2": None, "source": ""}
    for blob in extract_embedded_state_blobs(content):
        scope = blob.get("__DEFAULT_SCOPE__")
        if isinstance(scope, dict):
            for key, value in scope.items():
                key_norm = normalize_text(key).casefold()
                if not any(marker in key_norm for marker in UNIVERSAL_DETAIL_KEY_MARKERS):
                    continue
                if not isinstance(value, dict):
                    continue
                item_struct = item_struct_from_scope_value(value)
                if not isinstance(item_struct, dict):
                    continue
                item_id = clean_text(item_struct.get("id") or item_struct.get("awemeId"))
                if media_id and item_id != media_id:
                    continue
                stats = item_struct.get("stats")
                stats_v2 = item_struct.get("statsV2")
                if isinstance(stats, dict):
                    raw["stats"] = {field: stats.get(field) for field in COUNT_FIELD_MAP.values()}
                if isinstance(stats_v2, dict):
                    raw["statsV2"] = {field: stats_v2.get(field) for field in COUNT_FIELD_MAP.values()}
                raw["source"] = key
                return raw

        item_module = blob.get("ItemModule")
        if isinstance(item_module, dict) and media_id:
            item = item_module.get(media_id) or item_module.get(str(media_id))
            if isinstance(item, dict):
                item_id = clean_text(item.get("id") or item.get("awemeId") or item.get("itemId"))
                if item_id and item_id != media_id:
                    continue
                stats = item.get("stats")
                stats_v2 = item.get("statsV2")
                if isinstance(stats, dict):
                    raw["stats"] = {field: stats.get(field) for field in COUNT_FIELD_MAP.values()}
                if isinstance(stats_v2, dict):
                    raw["statsV2"] = {field: stats_v2.get(field) for field in COUNT_FIELD_MAP.values()}
                raw["source"] = "ItemModule"
                return raw
    return raw


def extract_embedded_state_blobs(content):
    blobs = []
    for pattern in EMBEDDED_STATE_SCRIPT_PATTERNS:
        for match in re.finditer(pattern, content or "", flags=re.DOTALL | re.IGNORECASE):
            data = json_loads_safe(match.group(1))
            if data:
                blobs.append(data)
    return blobs


def universal_detail_scope_metrics(scope, media_id=""):
    if not isinstance(scope, dict):
        return None
    for key, value in scope.items():
        key_norm = normalize_text(key).casefold()
        if not any(marker in key_norm for marker in UNIVERSAL_DETAIL_KEY_MARKERS):
            continue
        if not isinstance(value, dict):
            continue
        item_struct = item_struct_from_scope_value(value)
        if not isinstance(item_struct, dict):
            continue
        item_id = clean_text(item_struct.get("id") or item_struct.get("awemeId"))
        if media_id and item_id != media_id:
            continue
        metrics = metrics_from_stats_sources(item_struct)
        if metrics:
            return metrics
    return None


def parse_counts_from_universal_data(data, media_id=""):
    if not isinstance(data, dict):
        return None

    scope = data.get("__DEFAULT_SCOPE__")
    metrics = universal_detail_scope_metrics(scope, media_id=media_id)
    if metrics:
        return metrics
    return None


def parse_counts_from_sigi_state(data, media_id=""):
    if not isinstance(data, dict):
        return None

    item_module = data.get("ItemModule")
    if isinstance(item_module, dict) and media_id:
        for module_key in (media_id, str(media_id)):
            item = item_module.get(module_key)
            if isinstance(item, dict):
                item_id = clean_text(item.get("id") or item.get("awemeId") or item.get("itemId"))
                if item_id and item_id != media_id:
                    continue
                metrics = metrics_from_stats_sources(item)
                if metrics:
                    return metrics

    if media_id:
        for obj in iter_nested_dicts(data):
            if not isinstance(obj, dict):
                continue
            item_id = clean_text(obj.get("id") or obj.get("awemeId") or obj.get("itemId"))
            if item_id != media_id:
                continue
            metrics = metrics_from_stats_sources(obj)
            if metrics:
                return metrics
    return None


def parse_counts_regex(content, media_id=""):
    data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    found = False
    search_text = content or ""

    if media_id:
        for match in re.finditer(re.escape(f'"{media_id}"'), search_text):
            window = search_text[match.start() : match.start() + 5000]
            window_data, window_found = parse_counts_regex(window, media_id="")
            if window_found:
                return window_data, True

    for key, pattern in COUNT_PATTERNS.items():
        matches = [int(value) for value in re.findall(pattern, search_text)]
        if matches:
            data[key] = str(max(matches))
            found = True
    return data, found


def embedded_state_is_ambiguous(content):
    for blob in extract_embedded_state_blobs(content):
        item_module = blob.get("ItemModule")
        if isinstance(item_module, dict) and len(item_module) > 1:
            return True
    return False


def parse_counts_from_embedded_json(content, media_id=""):
    for blob in extract_embedded_state_blobs(content):
        metrics = parse_counts_from_universal_data(blob, media_id=media_id)
        if metrics:
            return metrics, True
        metrics = parse_counts_from_sigi_state(blob, media_id=media_id)
        if metrics:
            return metrics, True
    return None, False


def validate_metrics(data):
    if not data:
        return False

    values = {}
    for metric in COUNT_FIELD_MAP:
        parsed = parse_count_value(data.get(metric))
        if parsed is None:
            return False
        values[metric] = int(parsed)

    views = values["Views"]
    if views < 0:
        return False

    engagement = values["Likes"] + values["Comments"] + values["Saves"] + values["Shares"]
    if views == 0 and engagement > 0:
        return False

    for metric in ("Likes", "Comments", "Saves", "Shares"):
        if values[metric] < 0:
            return False
        if views > 0 and values[metric] > views * 5:
            return False

    return True


def counts_match(left, right):
    if not left or not right:
        return False
    return all(str(left.get(metric, "0")) == str(right.get(metric, "0")) for metric in COUNT_FIELD_MAP)


def parse_counts(content, media_id=""):
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}

    metrics, found = parse_counts_from_embedded_json(content, media_id=media_id)
    if found and metrics and validate_metrics(metrics):
        return metrics, True

    if media_id:
        return empty, False

    if embedded_state_is_ambiguous(content):
        return empty, False

    data, found = parse_counts_regex(content, media_id="")
    if found and validate_metrics(data):
        return data, True
    return empty, False


def json_loads_safe(raw_value):
    if not isinstance(raw_value, str):
        return None
    candidates = [raw_value]
    unescaped = html.unescape(raw_value)
    if unescaped != raw_value:
        candidates.append(unescaped)
    for candidate in candidates:
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            continue
    return None


def iter_nested_dicts(value):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from iter_nested_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_nested_dicts(child)


def author_name_from_object(obj, profile_username="", *, require_username_match=True):
    if not isinstance(obj, dict):
        return ""

    unique_id = clean_text(obj.get("uniqueId") or obj.get("unique_id") or obj.get("uniqueID"))
    if require_username_match:
        if not profile_username:
            return ""
        if normalize_text(unique_id.lstrip("@")) != normalize_text(profile_username.lstrip("@")):
            return ""
        handle = profile_username
    else:
        if not unique_id:
            return ""
        handle = unique_id

    profile_url = f"https://www.tiktok.com/@{handle.lstrip('@')}"
    for key in ("nickname", "nickName", "authorName", "name", "displayName"):
        candidate = clean_text(obj.get(key))
        if (
            candidate
            and not is_generated_username_channel(candidate, profile_url)
            and not is_generic_tiktok_channel_name(candidate)
        ):
            validated = valid_channel_candidate(candidate, handle)
            if validated:
                return validated
    return ""


def iter_matching_item_structs(content, media_id=""):
    media_id = clean_text(media_id)
    if not media_id:
        return

    for blob in extract_embedded_state_blobs(content):
        scope = blob.get("__DEFAULT_SCOPE__")
        if isinstance(scope, dict):
            for value in scope.values():
                item_struct = item_struct_from_scope_value(value)
                if not isinstance(item_struct, dict):
                    continue
                item_id = clean_text(item_struct.get("id") or item_struct.get("awemeId"))
                if item_id == media_id:
                    yield item_struct

        item_module = blob.get("ItemModule")
        if isinstance(item_module, dict):
            item = item_module.get(media_id) or item_module.get(str(media_id))
            if isinstance(item, dict):
                item_id = clean_text(item.get("id") or item.get("awemeId") or item.get("itemId"))
                if not item_id or item_id == media_id:
                    yield item

        for obj in iter_nested_dicts(blob):
            if not isinstance(obj, dict):
                continue
            item_id = clean_text(obj.get("id") or obj.get("awemeId") or obj.get("itemId"))
            if item_id != media_id:
                continue
            if isinstance(obj.get("author"), dict) or "stats" in obj or "statsV2" in obj:
                yield obj


def channel_name_from_item_struct(item_struct):
    if not isinstance(item_struct, dict):
        return ""
    author = item_struct.get("author")
    if isinstance(author, dict):
        candidate = author_name_from_object(author, require_username_match=False)
        if candidate:
            return candidate
    return author_name_from_object(item_struct, require_username_match=False)


def parse_channel_name_from_page(content, profile_username="", media_id=""):
    if media_id:
        for item_struct in iter_matching_item_structs(content, media_id):
            candidate = channel_name_from_item_struct(item_struct)
            if candidate:
                return candidate
    return parse_channel_name(content, profile_username)


TIKTOK_NOT_FOUND_PHRASES = (
    "couldn't find this account",
    "couldnt find this account",
    "couldn't find this video",
    "video currently unavailable",
    "this video is currently unavailable",
    "page not available",
    "page unavailable",
    "video unavailable",
    "không tìm thấy tài khoản này",
    "không tìm thấy video này",
    "khong tim thay tai khoan nay",
    "khong tim thay video nay",
    "video này hiện không khả dụng",
    "trang không khả dụng",
    "unable to find",
)

TIKTOK_ERROR_HINT_WORDS = (
    "couldn't",
    "couldnt",
    "unavailable",
    "trending creators",
    "discover more",
    "try searching",
    "log in",
    "back to home",
    "go home",
)


def visible_page_text(content):
    text = content or ""
    text = re.sub(r"<script\b[^>]*>.*?</script>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<style\b[^>]*>.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    return text.lower().replace("\u2019", "'").replace("\u2018", "'")


def is_tiktok_error_page(content):
    visible_html = re.sub(r"<script\b[^>]*>.*?</script>", " ", content or "", flags=re.DOTALL | re.IGNORECASE)
    visible_html = re.sub(r"<style\b[^>]*>.*?</style>", " ", visible_html, flags=re.DOTALL | re.IGNORECASE)
    surfaces = []
    heading_pattern = re.compile(
        r"<(?P<tag>title|h[1-3])\b[^>]*>(?P<body>.*?)</(?P=tag)>",
        flags=re.DOTALL | re.IGNORECASE,
    )
    for match in heading_pattern.finditer(visible_html):
        surfaces.append(strip_html_text(match.group("body")).casefold())

    alert_pattern = re.compile(
        r"<(?P<tag>[a-z0-9]+)\b(?P<attrs>[^>]*(?:role=[\"']alert[\"']|(?:class|id|data-e2e)=[\"'][^\"']*(?:error|not-found|unavailable)[^\"']*[\"'])[^>]*)>"
        r"(?P<body>.*?)</(?P=tag)>",
        flags=re.DOTALL | re.IGNORECASE,
    )
    for match in alert_pattern.finditer(visible_html):
        surfaces.append(strip_html_text(match.group("body")).casefold())

    return any(
        phrase in surface
        for surface in surfaces
        for phrase in TIKTOK_NOT_FOUND_PHRASES
    )


def looks_like_error_text(value):
    text = str(value or "").lower().replace("\u2019", "'").replace("\u2018", "'")
    if not text:
        return False
    if any(phrase in text for phrase in TIKTOK_NOT_FOUND_PHRASES):
        return True
    if any(word in text for word in TIKTOK_ERROR_HINT_WORDS):
        return True
    return False


def valid_channel_candidate(value, profile_username):
    candidate = clean_text(value)
    if not candidate:
        return ""

    profile_url = f"https://www.tiktok.com/@{profile_username}"
    normalized_candidate = normalize_text(candidate.lstrip("@"))
    normalized_username = normalize_text(profile_username.lstrip("@"))
    blocked_labels = {
        "FOLLOW",
        "MESSAGE",
        "FOLLOWING",
        "FOLLOWERS",
        "LIKES",
    }
    if (
        candidate.startswith("@")
        or normalized_candidate == normalized_username
        or normalized_candidate in blocked_labels
        or "FOLLOWERS" in normalized_candidate
        or "FOLLOWING" in normalized_candidate
        or len(candidate) > 80
        or looks_like_error_text(candidate)
        or is_generated_username_channel(candidate, profile_url)
        or is_generic_tiktok_channel_name(candidate)
    ):
        return ""
    return candidate


def parse_channel_name(content, profile_username):
    if not profile_username:
        return ""

    script_patterns = [
        r'<script[^>]+id="__UNIVERSAL_DATA_FOR_REHYDRATION__"[^>]*>(.*?)</script>',
        r'<script[^>]+id="SIGI_STATE"[^>]*>(.*?)</script>',
    ]
    for pattern in script_patterns:
        match = re.search(pattern, content, flags=re.DOTALL)
        if not match:
            continue
        data = json_loads_safe(match.group(1))
        for obj in iter_nested_dicts(data):
            candidate = author_name_from_object(obj, profile_username)
            if candidate:
                return candidate

    # Photo posts often expose only the embedded "author": {...} object
    author_blocks_quoted = re.finditer(
        r'"author"\s*:\s*(\{[^{}]*"uniqueId"\s*:\s*"([^"]+)"[^{}]*\})',
        content,
    )
    for block_match in author_blocks_quoted:
        if normalize_text(block_match.group(2).lstrip("@")) != normalize_text(profile_username.lstrip("@")):
            continue
        data = json_loads_safe(block_match.group(1))
        candidate = author_name_from_object(data, profile_username)
        if candidate:
            return candidate

    author_blocks = re.finditer(r'\{[^{}]*"uniqueId"\s*:\s*"([^"]+)"[^{}]*\}', content)
    for block_match in author_blocks:
        if normalize_text(block_match.group(1).lstrip("@")) != normalize_text(profile_username.lstrip("@")):
            continue
        data = json_loads_safe(block_match.group(0))
        candidate = author_name_from_object(data, profile_username)
        if candidate:
            return candidate
    return ""


def strip_html_text(value):
    text = html.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    return clean_text(re.sub(r"\s+", " ", text))


def profile_title_candidate(raw_title, profile_username):
    title = strip_html_text(raw_title)
    if not title:
        return ""

    candidate = re.split(
        r"\s+\|\s+TikTok|\s+-\s+TikTok|\s+on\s+TikTok",
        title,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    candidate = re.sub(
        rf"\s*\(@?{re.escape(profile_username)}\)\s*$",
        "",
        candidate,
        flags=re.IGNORECASE,
    )
    candidate = re.sub(
        rf"\s*@{re.escape(profile_username)}\s*$",
        "",
        candidate,
        flags=re.IGNORECASE,
    )
    candidate = clean_text(candidate.strip(" -|"))
    return valid_channel_candidate(candidate, profile_username)


def parse_profile_channel_name(content, profile_username, media_id=""):
    candidate = parse_channel_name_from_page(content, profile_username, media_id=media_id)
    if candidate:
        return candidate

    for tag in re.findall(r"<meta\b[^>]*>", content or "", flags=re.IGNORECASE):
        if not re.search(r'(?:property|name)=["\'](?:og:)?title["\']', tag, flags=re.IGNORECASE):
            continue
        match = re.search(r'content=["\']([^"\']+)["\']', tag, flags=re.IGNORECASE)
        if not match:
            continue
        candidate = profile_title_candidate(match.group(1), profile_username)
        if candidate:
            return candidate

    match = re.search(r"<title[^>]*>(.*?)</title>", content or "", flags=re.DOTALL | re.IGNORECASE)
    if match:
        candidate = profile_title_candidate(match.group(1), profile_username)
        if candidate:
            return candidate
    return ""


def fetch_profile_channel_name_request(profile_username, timeout=DEFAULT_REQUEST_TIMEOUT):
    username = clean_text(profile_username).lstrip("@")
    if not username:
        return ""
    profile_url = f"https://www.tiktok.com/@{username}"
    try:
        _final_url, content = fetch_tiktok_html(profile_url, timeout=timeout)
        return parse_profile_channel_name(content, username)
    except Exception:
        return ""


def is_usable_channel_name(name, url=""):
    text = clean_text(name)
    if not text or text == "Lỗi" or is_numeric_channel_garbage(text):
        return False
    if is_generated_username_channel(text, url) or is_generic_tiktok_channel_name(text):
        return False
    return True


def author_unique_id_from_post(url, resolved_url="", timeout=DEFAULT_REQUEST_TIMEOUT):
    media_id = extract_media_id(resolved_url) or extract_media_id(url)
    if not media_id:
        return ""
    for candidate_url in (resolved_url, url):
        normalized = normalize_tiktok_url(clean_text(candidate_url))
        if not normalized:
            continue
        try:
            _final_url, content = fetch_tiktok_html(normalized, timeout=timeout)
        except Exception:
            continue
        for item_struct in iter_matching_item_structs(content, media_id):
            author = item_struct.get("author")
            if isinstance(author, dict):
                unique_id = clean_text(author.get("uniqueId") or author.get("unique_id"))
                if unique_id:
                    return unique_id
    return ""


def enrich_channel_name(
    url,
    channel_name,
    *,
    resolved_url="",
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
    timeout=DEFAULT_REQUEST_TIMEOUT,
):
    source_url = normalize_tiktok_url(url)
    lookup_url = normalize_tiktok_url(resolved_url or url)
    username = username_for_channel_lookup(source_url, lookup_url)
    overrides = channel_overrides or {}
    parsed = clean_text(channel_name)

    def read_cache():
        if not channel_cache or not username:
            return ""
        if cache_lock:
            with cache_lock:
                return clean_text(channel_cache.get(username.casefold(), ""))
        return clean_text(channel_cache.get(username.casefold(), ""))

    def write_cache(name):
        if channel_cache and username and name:
            if cache_lock:
                with cache_lock:
                    channel_cache[username.casefold()] = name
            else:
                channel_cache[username.casefold()] = name

    def profile_already_tried():
        if not profile_lookup_attempted or not username:
            return False
        key = username.casefold()
        if cache_lock:
            with cache_lock:
                return key in profile_lookup_attempted
        return key in profile_lookup_attempted

    def mark_profile_tried():
        if not profile_lookup_attempted or not username:
            return
        key = username.casefold()
        if cache_lock:
            with cache_lock:
                profile_lookup_attempted.add(key)
        else:
            profile_lookup_attempted.add(key)

    resolved = resolve_channel_name(source_url, parsed.lstrip("@") if parsed.startswith("@") else parsed, overrides)
    if not resolved and parsed.startswith("@") and is_usable_channel_name(parsed, lookup_url or source_url):
        resolved = parsed
    if is_usable_channel_name(resolved, lookup_url or source_url):
        write_cache(resolved)
        return resolved

    cached = read_cache()
    if is_usable_channel_name(cached, lookup_url or source_url):
        return cached

    if username and not profile_already_tried():
        mark_profile_tried()
        fetched = fetch_profile_channel_name_request(username, timeout=timeout)
        if not fetched and is_generated_username_channel(username) and not session_uses_proxy():
            alt_handle = author_unique_id_from_post(source_url, lookup_url, timeout=timeout)
            if alt_handle and alt_handle.casefold() != username.casefold():
                fetched = fetch_profile_channel_name_request(alt_handle, timeout=timeout)
        resolved = resolve_channel_name(source_url, fetched, overrides) or fetched
        if is_usable_channel_name(resolved, lookup_url or source_url):
            write_cache(resolved)
            return resolved

    if username:
        handle = f"@{username.lstrip('@')}"
        write_cache(handle)
        return handle

    return ""


def seed_channel_cache_from_workbook(rows_to_process, sheet_contexts, channel_cache):
    for item in rows_to_process:
        username = extract_profile_username(item.get("url", ""))
        if not username:
            continue
        context = sheet_contexts.get(item["sheet_name"])
        if not context:
            continue
        channel_col = context["columns"].get("channel")
        if not channel_col:
            continue
        raw = clean_text(context["worksheet"].cell(row=item["row"], column=channel_col).value)
        if raw and raw != "Lỗi" and is_usable_channel_name(raw, item["url"]):
            channel_cache[username.casefold()] = raw


def channel_name_for_sheet(
    url,
    channel_name,
    *,
    resolved_url="",
    status="Success",
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
):
    source_url = normalize_tiktok_url(url)
    resolved = enrich_channel_name(
        source_url,
        channel_name,
        resolved_url=resolved_url,
        channel_cache=channel_cache,
        channel_overrides=channel_overrides,
        profile_lookup_attempted=profile_lookup_attempted,
        cache_lock=cache_lock,
    )
    lookup_url = normalize_tiktok_url(resolved_url or url)
    if is_usable_channel_name(resolved, lookup_url or source_url):
        return resolved
    username = username_for_channel_lookup(source_url, lookup_url)
    if status == "Success" and username and not is_generated_username_channel(username):
        return f"@{username.lstrip('@')}"
    return "Lỗi"


def channel_name_quality(name, url=""):
    """Rank a channel cell so re-scrapes never downgrade a better existing value.

    2 = real nickname, 1 = @handle fallback, 0 = failed/garbage/empty.
    """
    text = clean_text(name)
    if not is_usable_channel_name(text, url):
        return 0
    return 1 if text.startswith("@") else 2


def format_metric_log_plain(data):
    views = int(metric_number(data.get("Views", 0)))
    likes = int(metric_number(data.get("Likes", 0)))
    comments = int(metric_number(data.get("Comments", 0)))
    saves = int(metric_number(data.get("Saves", 0)))
    shares = int(metric_number(data.get("Shares", 0)))
    return (
        f"Lượt xem {views} • Tim {likes} • Bình luận {comments} • "
        f"Lưu {saves} • Chia sẻ {shares}"
    )


def scrape_metric_details(data):
    return {
        "views": int(metric_number(data.get("Views", 0))),
        "likes": int(metric_number(data.get("Likes", 0))),
        "comments": int(metric_number(data.get("Comments", 0))),
        "saves": int(metric_number(data.get("Saves", 0))),
        "shares": int(metric_number(data.get("Shares", 0))),
    }


def format_metric_log_line(data):
    metrics = scrape_metric_details(data)
    return (
        f"view={metrics['views']} tim={metrics['likes']} cmt={metrics['comments']} "
        f"save={metrics['saves']} share={metrics['shares']}"
    )


def extract_profile_username(url):
    match = re.search(r"tiktok\.com/@([^/?]+)", url or "", re.IGNORECASE)
    return match.group(1).strip() if match else ""


def username_for_channel_lookup(source_url, resolved_url=""):
    for candidate in (resolved_url, source_url):
        normalized = normalize_tiktok_url(clean_text(candidate))
        if not normalized:
            continue
        username = extract_profile_username(normalized)
        if username:
            return username
    return ""


async def read_channel_name_from_dom(page, profile_username):
    if not profile_username:
        return ""

    selectors = [
        f'a[href*="@{profile_username}"]',
        f'a[href="/@{profile_username}"]',
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = await locator.count()
            if count == 0:
                continue
            for index in range(min(count, 8)):
                try:
                    text = clean_text(await locator.nth(index).inner_text(timeout=1500))
                except Exception:
                    continue
                if not text:
                    continue
                normalized = text.lstrip("@").strip()
                if normalize_text(normalized) == normalize_text(profile_username):
                    continue
                if text.startswith("@"):
                    continue
                candidate = valid_channel_candidate(text, profile_username)
                if candidate:
                    return candidate
        except Exception:
            continue
    return ""


def profile_text_candidates(text):
    for line in str(text or "").splitlines():
        line = clean_text(line)
        if line:
            yield line


async def read_profile_channel_name_from_dom(page, profile_username):
    selectors = [
        '[data-e2e="user-title"]',
        'h1[data-e2e="user-title"]',
        "main h1",
        "h1",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = await locator.count()
            for index in range(min(count, 5)):
                try:
                    text = await locator.nth(index).inner_text(timeout=1200)
                except Exception:
                    continue
                for line in profile_text_candidates(text):
                    candidate = valid_channel_candidate(line, profile_username)
                    if candidate:
                        return candidate
        except Exception:
            continue
    return ""


async def read_profile_channel_name(page, profile_username, channel_cache=None, timeout_ms=18000):
    if not profile_username:
        return ""

    cache_key = profile_username.casefold()
    if isinstance(channel_cache, dict):
        cached = clean_text(channel_cache.get(cache_key))
        if (
            cached
            and not is_generated_username_channel(cached, f"https://www.tiktok.com/@{profile_username}")
            and not is_generic_tiktok_channel_name(cached)
        ):
            return cached

    profile_url = f"https://www.tiktok.com/@{profile_username}"
    try:
        await page.goto(profile_url, wait_until="domcontentloaded", timeout=timeout_ms)
        for _ in range(20):
            content = await page.content()
            candidate = parse_profile_channel_name(content, profile_username)
            if not candidate:
                candidate = await read_profile_channel_name_from_dom(page, profile_username)
            if candidate:
                if isinstance(channel_cache, dict):
                    channel_cache[cache_key] = candidate
                return candidate
            await page.wait_for_timeout(500)
    except Exception:
        return ""
    return ""


def item_struct_from_scope_value(value):
    if not isinstance(value, dict):
        return None
    item_info = value.get("itemInfo")
    if isinstance(item_info, dict):
        item_struct = item_info.get("itemStruct")
        if isinstance(item_struct, dict):
            return item_struct
    item_struct = value.get("itemStruct")
    if isinstance(item_struct, dict):
        return item_struct
    extra_info = value.get("extra_info")
    if isinstance(extra_info, dict):
        nested = item_struct_from_scope_value(extra_info)
        if nested:
            return nested
        nested = extra_info.get("itemStruct")
        if isinstance(nested, dict):
            return nested
    return None


def request_html_has_metric_hints(content):
    return bool(REQUEST_METRIC_HINT_PATTERN.search(content or ""))


def is_hidden_stats_status(status):
    """True khi post TikTok ẩn số liệu (không phải lỗi quét mạng/HTTP)."""
    text = clean_text(status)
    if not text:
        return False
    if text in (STATUS_TIKTOK_NO_STATS, STATUS_TIKTOK_NO_STATS_LEGACY):
        return True
    lowered = text.casefold()
    return (
        "ẩn số liệu" in lowered
        or "tiktok không trả số liệu" in lowered
        or "tiktok không trả lượt xem" in lowered
    )


def should_clear_stale_metrics(status):
    text = clean_text(status)
    lowered = text.casefold()
    return (
        is_hidden_stats_status(status)
        or lowered == "error: trang tiktok không khả dụng"
        or bool(re.fullmatch(r"error:\s*http\s+(?:404|410)(?:\s+.*)?", lowered))
    )


def terminal_status_priority(status):
    if is_hidden_stats_status(status):
        return 40
    text = clean_text(status).casefold()
    if re.fullmatch(r"error:\s*http\s+410(?:\s+.*)?", text):
        return 30
    if text == "error: trang tiktok không khả dụng":
        return 20
    if re.fullmatch(r"error:\s*http\s+404(?:\s+.*)?", text):
        return 10
    return 0


def stronger_terminal_result(current, candidate):
    if current is None:
        return candidate
    if terminal_status_priority(candidate[2]) > terminal_status_priority(current[2]):
        return candidate
    return current


def item_struct_explicitly_hides_stats(item_struct):
    hidden_when_true = {
        "hidestats",
        "hideviewcount",
        "isstatshidden",
        "statshidden",
        "viewcounthidden",
    }
    hidden_when_false = {
        "isstatsvisible",
        "statisticsvisible",
        "statsvisible",
        "viewcountvisible",
    }
    visibility_fields = {"statisticsvisibility", "statsvisibility"}

    for obj in iter_nested_dicts(item_struct):
        for key, value in obj.items():
            normalized_key = re.sub(r"[^a-z0-9]", "", str(key).casefold())
            if normalized_key not in hidden_when_true | hidden_when_false | visibility_fields:
                continue
            normalized_value = clean_text(value).casefold()
            if normalized_key in hidden_when_true and (
                value is True or normalized_value in {"1", "true", "hidden"}
            ):
                return True
            if normalized_key in hidden_when_false and (
                value is False or normalized_value in {"0", "false", "hidden"}
            ):
                return True
            if normalized_key in visibility_fields and normalized_value in {"hidden", "private", "disabled"}:
                return True
    return False


def no_metrics_status(content, media_id=""):
    """Only classify hidden stats when the matching item says so explicitly."""
    if media_id:
        for item_struct in iter_matching_item_structs(content, media_id):
            if item_struct_explicitly_hides_stats(item_struct):
                return STATUS_TIKTOK_NO_STATS
    return STATUS_METRICS_UNREADABLE


def media_redirect_mismatch(source_url, final_url):
    source_media_id = extract_media_id(source_url)
    final_media_id = extract_media_id(final_url)
    return bool(source_media_id and final_media_id != source_media_id)


def build_request_url_candidates(url):
    url = normalize_tiktok_url(url)
    candidates = []
    seen = set()

    def add(candidate):
        normalized = normalize_tiktok_url(candidate) if candidate else ""
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        candidates.append(normalized)

    add(url)
    base = url.split("?")[0]
    media_id = extract_media_id(url)
    profile_username = extract_profile_username(url)
    if media_id and "/photo/" in url.lower():
        if "_r=1" not in url.lower():
            add(f"{base}?_r=1")
        if profile_username:
            add(f"https://www.tiktok.com/@{profile_username}/video/{media_id}")
    add(base)

    return candidates


def parse_fetched_request_page(source_url, final_url, content):
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    source_media_id = extract_media_id(source_url)
    final_media_id = extract_media_id(final_url)
    if media_redirect_mismatch(source_url, final_url):
        return empty, "", STATUS_MEDIA_REDIRECT_MISMATCH

    media_id = source_media_id or final_media_id
    profile_username = extract_profile_username(source_url) or extract_profile_username(final_url)
    if not media_id:
        channel_name = ""
        if profile_username and not is_generated_username_channel(profile_username):
            channel_name = f"@{profile_username.lstrip('@')}"
        return empty, channel_name, STATUS_METRICS_UNREADABLE

    metrics, found = parse_counts(content, media_id=media_id)
    channel_name = parse_profile_channel_name(content, profile_username, media_id=media_id)
    if not channel_name and profile_username and not is_generated_username_channel(profile_username):
        channel_name = f"@{profile_username.lstrip('@')}"
    if found and validate_metrics(metrics):
        return metrics, channel_name, "Success"
    if is_tiktok_error_page(content):
        return empty, "", "Error: Trang TikTok không khả dụng"
    return (metrics if found else empty), channel_name, no_metrics_status(content, media_id=media_id)


def hybrid_browser_worker_count(request_workers, total_links):
    if total_links <= 0:
        return 0
    if total_links <= 40:
        return min(max(request_workers // 2, 3), 8)
    return min(max(request_workers // 3, 5), MAX_BROWSER_FALLBACK_WORKERS)


def fetch_tiktok_html(url, timeout=DEFAULT_REQUEST_TIMEOUT):
    wait_if_request_blocked()
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": random.choice(USER_AGENTS),
            "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.tiktok.com/",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    )
    with _request_semaphore:
        with urlopen_request(request, timeout=timeout) as response:
            return response.geturl(), response.read().decode("utf-8", errors="replace")


def scrape_link_request(url, timeout=DEFAULT_REQUEST_TIMEOUT):
    data, channel_name, status, _hints, _resolved = _scrape_link_request_impl(url, timeout=timeout)
    return data, channel_name, status


def _scrape_link_request_impl(url, timeout=DEFAULT_REQUEST_TIMEOUT):
    url = normalize_tiktok_url(url)
    last_data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    last_channel = ""
    last_status = "Error: Không đọc được số liệu"
    last_resolved_url = ""
    saw_metric_hints = False
    best_terminal = None

    try:
        candidate_limit = request_candidate_limit()
        candidates = build_request_url_candidates(url)[:candidate_limit]
        best_success = None

        for candidate in candidates:
            try:
                final_url, content = fetch_tiktok_html(candidate, timeout=timeout)
            except urllib.error.HTTPError as error:
                last_status = f"Error: HTTP {error.code}"
                if should_clear_stale_metrics(last_status):
                    best_terminal = stronger_terminal_result(
                        best_terminal,
                        (last_data, last_channel, last_status, last_resolved_url),
                    )
                if error.code in (403, 429):
                    note_request_rate_limit(error.code)
                    release_thread_proxy()
                continue
            except Exception as error:
                last_status = f"Error: {str(error)}"
                if is_transient_network_status(last_status):
                    note_network_failure()
                continue

            note_network_success()
            redirect_mismatch = media_redirect_mismatch(url, final_url)
            if final_url and not redirect_mismatch:
                last_resolved_url = final_url
            if request_html_has_metric_hints(content):
                saw_metric_hints = True

            data, channel_name, status = parse_fetched_request_page(url, final_url, content)
            last_data, last_channel, last_status = data, channel_name, status
            if should_clear_stale_metrics(status):
                best_terminal = stronger_terminal_result(
                    best_terminal,
                    (data, channel_name, status, last_resolved_url),
                )
            if status == "Success":
                if best_success is None or (channel_name and not best_success[1]):
                    best_success = (data, channel_name, status, final_url)
                if channel_name:
                    return data, channel_name, status, saw_metric_hints, final_url

            if request_html_has_metric_hints(content):
                time.sleep(request_shell_retry_delay())
                try:
                    final_url, content = fetch_tiktok_html(candidate, timeout=timeout)
                except urllib.error.HTTPError as error:
                    last_status = f"Error: HTTP {error.code}"
                    if should_clear_stale_metrics(last_status):
                        best_terminal = stronger_terminal_result(
                            best_terminal,
                            (last_data, last_channel, last_status, last_resolved_url),
                        )
                    if error.code in (403, 429):
                        note_request_rate_limit(error.code)
                        release_thread_proxy()
                    continue
                except Exception as error:
                    last_status = f"Error: {str(error)}"
                    if is_transient_network_status(last_status):
                        note_network_failure()
                    continue
                redirect_mismatch = media_redirect_mismatch(url, final_url)
                if final_url and not redirect_mismatch:
                    last_resolved_url = final_url
                data, channel_name, status = parse_fetched_request_page(url, final_url, content)
                last_data, last_channel, last_status = data, channel_name, status
                if should_clear_stale_metrics(status):
                    best_terminal = stronger_terminal_result(
                        best_terminal,
                        (data, channel_name, status, last_resolved_url),
                    )
                if status == "Success":
                    if best_success is None or (channel_name and not best_success[1]):
                        best_success = (data, channel_name, status, final_url)
                    if channel_name:
                        return data, channel_name, status, saw_metric_hints, final_url

            redirect_candidate = normalize_tiktok_url(final_url.split("?")[0])
            if (
                redirect_candidate
                and not media_redirect_mismatch(url, final_url)
                and redirect_candidate not in candidates
                and len(candidates) < candidate_limit
            ):
                candidates.append(redirect_candidate)

        if best_success:
            data, channel_name, status, final_url = best_success
            return data, channel_name, status, saw_metric_hints, final_url
        if best_terminal:
            data, channel_name, status, final_url = best_terminal
            return data, channel_name, status, saw_metric_hints, final_url

        return last_data, last_channel, last_status, saw_metric_hints, last_resolved_url
    except urllib.error.HTTPError as error:
        return last_data, last_channel, f"Error: HTTP {error.code}", saw_metric_hints, last_resolved_url
    except Exception as error:
        return last_data, last_channel, f"Error: {str(error)}", saw_metric_hints, last_resolved_url


def scrape_link_with_retries_request(
    url,
    retries=DEFAULT_RETRIES,
    timeout=DEFAULT_REQUEST_TIMEOUT,
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
):
    last_data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    last_status = "Error: Chưa chạy"
    last_channel = ""
    last_resolved_url = ""

    attempts_used = 0
    for attempt in range(retries + 1):
        if attempt > 0:
            if is_request_rate_limited_status(last_status):
                if session_uses_proxy():
                    time.sleep(0.4 + attempt * 0.35 + random.uniform(0.1, 0.25))
                else:
                    time.sleep(2.5 + attempt * 2.0 + random.uniform(0.5, 2.0))
            else:
                time.sleep(1.0 + attempt * 0.75 + random.uniform(0.2, 0.8))
        data, channel_name, status, saw_metric_hints, resolved_url = _scrape_link_request_impl(
            url, timeout=timeout
        )
        attempts_used = attempt + 1
        if resolved_url:
            last_resolved_url = resolved_url
        channel_name = enrich_channel_name(
            url,
            channel_name,
            resolved_url=resolved_url if status == "Success" else "",
            channel_cache=channel_cache,
            channel_overrides=channel_overrides,
            profile_lookup_attempted=profile_lookup_attempted,
            cache_lock=cache_lock,
            timeout=timeout,
        )
        last_data, last_channel, last_status = data, channel_name, status
        if status == "Success":
            return data, channel_name, status, attempts_used, last_resolved_url
        if should_clear_stale_metrics(status):
            return data, channel_name, status, attempts_used, last_resolved_url
        if is_request_rate_limited_status(status):
            release_thread_proxy()
            continue
        if is_transient_network_status(status):
            continue
        if status == STATUS_METRICS_UNREADABLE:
            continue
        if not saw_metric_hints:
            break

    return last_data, last_channel, last_status, attempts_used or 1, last_resolved_url


async def block_heavy_resources(route):
    if route.request.resource_type in BLOCKED_RESOURCE_TYPES:
        await route.abort()
    else:
        await route.continue_()


async def make_browser_context(browser, proxy_configs=None):
    context_kwargs = {
        "user_agent": DEFAULT_USER_AGENT,
        "viewport": {"width": 390, "height": 844},
        "locale": BROWSER_LOCALE,
        "timezone_id": BROWSER_TIMEZONE,
        "geolocation": BROWSER_GEOLOCATION,
        "permissions": ["geolocation"],
    }
    configs = [item for item in (proxy_configs or []) if item and item.get("enabled")]
    if configs:
        proxy_config = random.choice(configs)
        context_kwargs["proxy"] = playwright_proxy_settings(proxy_config)
    context = await browser.new_context(**context_kwargs)
    await context.route("**/*", block_heavy_resources)
    return context


async def scrape_single_link(page, url, channel_cache=None, timeout_ms=45000):
    data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    channel_name = ""
    url = normalize_tiktok_url(url)
    profile_username = extract_profile_username(url)
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        try:
            await page.wait_for_selector(
                'script#__UNIVERSAL_DATA_FOR_REHYDRATION__, script#SIGI_STATE, script#api-data',
                timeout=12000,
            )
        except Exception:
            pass

        # Resolve final URL after redirect (vt.tiktok.com/... -> www.tiktok.com/@user/...)
        page_url = url
        try:
            page_url = page.url or url
        except Exception:
            page_url = url
        if not profile_username:
            profile_username = extract_profile_username(page_url)
        source_media_id = extract_media_id(url)
        final_media_id = extract_media_id(page_url)
        if media_redirect_mismatch(url, page_url):
            return data, channel_name, STATUS_MEDIA_REDIRECT_MISMATCH, url
        media_id = source_media_id or final_media_id
        if not media_id:
            if profile_username and not is_generated_username_channel(profile_username):
                channel_name = f"@{profile_username.lstrip('@')}"
            return data, channel_name, STATUS_METRICS_UNREADABLE, page_url

        previous_metrics = None
        found = False
        content = ""
        for _ in range(24):
            content = await page.content()

            candidate, candidate_found = parse_counts(content, media_id=media_id)
            if candidate_found and validate_metrics(candidate):
                if previous_metrics and counts_match(previous_metrics, candidate):
                    data = candidate
                    found = True
                    break
                previous_metrics = candidate
            elif is_tiktok_error_page(content):
                return data, channel_name, "Error: Trang TikTok không khả dụng", page_url
            parsed_channel = parse_channel_name_from_page(content, profile_username, media_id=media_id)
            if parsed_channel:
                channel_name = parsed_channel
            if profile_username and not channel_name:
                dom_channel = await read_channel_name_from_dom(page, profile_username)
                if dom_channel and not is_generated_username_channel(dom_channel, url) and not is_generic_tiktok_channel_name(dom_channel):
                    channel_name = dom_channel
            await page.wait_for_timeout(500)

        if profile_username:
            needs_profile_channel = (
                not channel_name
                or is_generated_username_channel(channel_name, url)
                or is_generic_tiktok_channel_name(channel_name)
            )
            if needs_profile_channel:
                profile_channel = await read_profile_channel_name(page, profile_username, channel_cache=channel_cache)
                if profile_channel:
                    channel_name = profile_channel

        # Final guard against accidentally captured error text
        if looks_like_error_text(channel_name):
            channel_name = ""

        # Fallback to @handle when we can detect the profile but no nickname is parsed.
        # Auto-generated user-ID handles (user1234567890) stay empty -> mark as "Lỗi".
        if not channel_name and profile_username and not is_generated_username_channel(profile_username):
            channel_name = f"@{profile_username}"

        if not found and previous_metrics and validate_metrics(previous_metrics):
            data = previous_metrics
            found = True

        if not found:
            return data, channel_name, no_metrics_status(content, media_id=media_id), page_url

        return data, channel_name, "Success", page_url
    except Exception as error:
        return data, channel_name, f"Error: {str(error)}", url


async def scrape_with_retries(page, url, retries, channel_cache=None):
    last_data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    last_status = "Error: Chưa chạy"
    last_channel = ""
    last_resolved_url = ""

    for attempt in range(retries + 1):
        if attempt > 0:
            await asyncio.sleep(random.uniform(1.5, 3.5))
        data, channel_name, status, resolved_url = await scrape_single_link(page, url, channel_cache=channel_cache)
        if resolved_url:
            last_resolved_url = resolved_url
        last_data, last_channel, last_status = data, channel_name, status
        if status == "Success":
            return data, channel_name, status, attempt + 1, last_resolved_url
        if should_clear_stale_metrics(status):
            return data, channel_name, status, attempt + 1, last_resolved_url

    return last_data, last_channel, last_status, retries + 1, last_resolved_url


def select_fallback_result(request_result, browser_result):
    if browser_result.get("status") == "Success":
        return browser_result
    if should_clear_stale_metrics(request_result.get("status")) and not should_clear_stale_metrics(
        browser_result.get("status")
    ):
        return request_result
    return browser_result


def guard_result_media_identity(result):
    if result.get("status") != "Success":
        return result

    raw_expected_ids = result.get("expected_media_ids") or []
    if isinstance(raw_expected_ids, str):
        raw_expected_ids = [raw_expected_ids]
    expected_ids = {clean_text(value) for value in raw_expected_ids if clean_text(value)}
    if not expected_ids:
        return result

    resolved_media_id = extract_media_id(result.get("resolved_url", ""))
    if len(expected_ids) == 1 and resolved_media_id in expected_ids:
        return result

    guarded = dict(result)
    guarded.update({
        "data": {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"},
        "channel_name": "",
        "status": STATUS_MEDIA_REDIRECT_MISMATCH,
        "resolved_url": "",
    })
    return guarded


async def worker_loop(
    worker_id,
    browser,
    scrape_queue,
    result_queue,
    retries,
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
    startup_semaphore=None,
    websocket_manager=None,
    worker_label=None,
    proxy_config=None,
    proxy_configs=None,
):
    RECYCLE_AFTER = 100
    display_worker = worker_label if worker_label is not None else worker_id
    browser_proxy_configs = proxy_configs if proxy_configs is not None else ([proxy_config] if proxy_config else [])

    async def make_context():
        return await make_browser_context(browser, proxy_configs=browser_proxy_configs)

    current_item = None

    if startup_semaphore is not None:
        async with startup_semaphore:
            await asyncio.sleep(min(worker_id - 1, 10) * 0.25)
            try:
                context = await make_context()
                page = await context.new_page()
            except Exception as error:
                if websocket_manager:
                    await websocket_manager.broadcast_log(f"Worker {display_worker} không khởi tạo được context: {str(error)}")
                return
    else:
        context = await make_context()
        page = await context.new_page()

    links_in_current_context = 0

    try:
        while True:
            item = await scrape_queue.get()
            current_item = item
            if item is None:
                scrape_queue.task_done()
                current_item = None
                break

            started_at = time.perf_counter()
            try:
                data, channel_name, status, attempts, resolved_url = await scrape_with_retries(page, item["url"], retries, channel_cache=channel_cache)
            except Exception as error:
                data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
                channel_name = ""
                status = f"Error: Worker {display_worker} crash ({str(error)})"
                attempts = 0
                resolved_url = ""
                try:
                    if page.is_closed():
                        page = await context.new_page()
                except Exception:
                    pass
            elapsed = time.perf_counter() - started_at
            browser_result = {
                "data": data,
                "channel_name": channel_name,
                "status": status,
                "attempts": attempts,
                "resolved_url": resolved_url,
                "elapsed": elapsed,
                "worker": display_worker,
            }
            selected_result = select_fallback_result(item.get("_request_result") or {}, browser_result)
            data = selected_result.get("data", data)
            channel_name = selected_result.get("channel_name", channel_name)
            status = selected_result.get("status", status)
            attempts = selected_result.get("attempts", attempts)
            resolved_url = selected_result.get("resolved_url", resolved_url)
            elapsed = selected_result.get("elapsed", elapsed)
            selected_worker = selected_result.get("worker", display_worker)
            channel_name = enrich_channel_name(
                item["url"],
                channel_name,
                resolved_url=resolved_url if status == "Success" else "",
                channel_cache=channel_cache,
                channel_overrides=channel_overrides,
                profile_lookup_attempted=profile_lookup_attempted,
                cache_lock=cache_lock,
            )
            result_item = {key: value for key, value in item.items() if key != "_request_result"}
            await result_queue.put({
                **result_item,
                "worker": selected_worker,
                "data": data,
                "channel_name": channel_name,
                "status": status,
                "attempts": attempts,
                "elapsed": elapsed,
                "resolved_url": resolved_url,
            })
            scrape_queue.task_done()
            current_item = None
            links_in_current_context += 1

            if links_in_current_context >= RECYCLE_AFTER:
                try:
                    await context.close()
                except Exception:
                    pass
                try:
                    context = await make_context()
                    page = await context.new_page()
                    links_in_current_context = 0
                except Exception as error:
                    if websocket_manager:
                        await websocket_manager.broadcast_log(f"Worker {display_worker} không khởi tạo lại được context: {str(error)}")
                    return

            await asyncio.sleep(random.uniform(0.3, 1.1))
    except Exception as error:
        if websocket_manager:
            await websocket_manager.broadcast_log(f"Worker {display_worker} dừng do lỗi: {str(error)}")
        if current_item is not None:
            await scrape_queue.put(current_item)
            scrape_queue.task_done()
    finally:
        try:
            await context.close()
        except Exception:
            pass


def _run_request_scrape(
    worker_index,
    url,
    retries=DEFAULT_RETRIES,
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
):
    # Gán proxy round-robin ngay trên thread thực thi hiện tại (thread pool của
    # asyncio có thể đổi thread giữa các item, nên phải gán lại mỗi lần, không
    # chỉ 1 lần khi worker khởi động).
    assign_worker_proxy(worker_index)
    return scrape_link_with_retries_request(
        url,
        retries=retries,
        channel_cache=channel_cache,
        channel_overrides=channel_overrides,
        profile_lookup_attempted=profile_lookup_attempted,
        cache_lock=cache_lock,
    )


async def request_worker_loop(
    worker_id,
    scrape_queue,
    browser_queue,
    result_queue,
    retries,
    websocket_manager=None,
    browser_fallback=True,
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
):
    loop = asyncio.get_running_loop()
    worker_label = f"R{worker_id}"
    worker_index = worker_id - 1
    current_item = None

    try:
        while True:
            item = await scrape_queue.get()
            current_item = item
            if item is None:
                scrape_queue.task_done()
                current_item = None
                break

            started_at = time.perf_counter()
            try:
                data, channel_name, status, attempts, resolved_url = await loop.run_in_executor(
                    None,
                    lambda url=item["url"]: _run_request_scrape(
                        worker_index,
                        url,
                        retries=retries,
                        channel_cache=channel_cache,
                        channel_overrides=channel_overrides,
                        profile_lookup_attempted=profile_lookup_attempted,
                        cache_lock=cache_lock,
                    ),
                )
            except Exception as error:
                data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
                channel_name = ""
                status = f"Error: Request worker {worker_label} crash ({str(error)})"
                attempts = 0
                resolved_url = ""

            elapsed = time.perf_counter() - started_at
            if status == "Success" or not browser_fallback:
                await result_queue.put({
                    **item,
                    "worker": worker_label,
                    "data": data,
                    "channel_name": channel_name,
                    "status": status,
                    "attempts": attempts,
                    "elapsed": elapsed,
                    "resolved_url": resolved_url,
                })
            else:
                await browser_queue.put({
                    **item,
                    "_request_result": {
                        "worker": worker_label,
                        "data": data,
                        "channel_name": channel_name,
                        "status": status,
                        "attempts": attempts,
                        "elapsed": elapsed,
                        "resolved_url": resolved_url,
                    },
                })
            scrape_queue.task_done()
            current_item = None
    except Exception as error:
        if websocket_manager:
            await websocket_manager.broadcast_log(f"Request worker {worker_label} dừng do lỗi: {str(error)}")
        if current_item is not None:
            await browser_queue.put(current_item)
            scrape_queue.task_done()


def write_result(
    sheet_contexts,
    item,
    data,
    channel_name,
    status,
    *,
    resolved_url="",
    channel_cache=None,
    channel_overrides=None,
    profile_lookup_attempted=None,
    cache_lock=None,
):
    context = sheet_contexts[item["sheet_name"]]
    sheet = context["worksheet"]
    columns = context["columns"]
    row_index = item["row"]
    update_time = format_display_datetime()

    write_metrics = status == "Success" or should_clear_stale_metrics(status)
    if write_metrics:
        for metric_key, data_key in METRIC_KEYS.items():
            column_index = columns.get(metric_key)
            if column_index:
                value = int(data.get(data_key) or 0) if status == "Success" else 0
                sheet.cell(row=row_index, column=column_index).value = value

    if columns.get("scan_status"):
        sheet.cell(row=row_index, column=columns["scan_status"]).value = clean_text(status)
    if status == "Success" and resolved_url and columns.get("resolved_url"):
        sheet.cell(row=row_index, column=columns["resolved_url"]).value = normalize_tiktok_url(resolved_url)
    if status == "Success" and columns.get("source_url"):
        sheet.cell(row=row_index, column=columns["source_url"]).value = normalize_tiktok_url(item["url"])

    if columns.get("channel"):
        channel_value = channel_name_for_sheet(
            item["url"],
            channel_name,
            resolved_url=resolved_url if status == "Success" else "",
            status=status,
            channel_cache=channel_cache,
            channel_overrides=channel_overrides,
            profile_lookup_attempted=profile_lookup_attempted,
            cache_lock=cache_lock,
        )
        existing = clean_text(sheet.cell(row=row_index, column=columns["channel"]).value)
        if status != "Success" and channel_name_quality(existing, item["url"]) > 0:
            channel_value = existing
        elif channel_name_quality(existing, item["url"]) > channel_name_quality(channel_value, item["url"]):
            channel_value = existing
        sheet.cell(row=row_index, column=columns["channel"]).value = channel_value

    # Chỉ cập nhật timestamp khi quét thành công — tránh hiểu nhầm đã quét OK.
    if status == "Success" and columns.get("last_update"):
        sheet.cell(row=row_index, column=columns["last_update"]).value = update_time


async def save_workbook(workbook, file_path, websocket_manager=None):
    temp_path = f"{file_path}.tmp"
    try:
        await asyncio.to_thread(workbook.save, temp_path)
        await asyncio.to_thread(os.replace, temp_path, file_path)
        return True, None
    except PermissionError:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        if websocket_manager:
            await websocket_manager.broadcast_log("CẢNH BÁO: File Excel đang mở, không thể ghi đè. Vui lòng đóng file rồi quét lại hoặc chờ lần lưu tiếp theo.")
        return False, "permission"
    except Exception as error:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        if websocket_manager:
            await websocket_manager.broadcast_log(f"CẢNH BÁO: Lỗi lưu file ({str(error)})")
        return False, "other"


def format_scrape_result_log(result, processed, total):
    url = clean_text(result.get("url", ""))
    worker = result.get("worker", "?")
    status = clean_text(result.get("status", ""))
    attempts = result.get("attempts", 0)
    elapsed = float(result.get("elapsed") or 0)
    data = result.get("data") or {}
    rows = result.get("rows") or []
    channel = clean_text(result.get("channel_name", "")) or "—"
    row_refs = ", ".join(
        f'{clean_text(row.get("sheet_name", ""))}#{row.get("row", "")}'
        for row in rows[:6]
    )
    if len(rows) > 6:
        row_refs = f"{row_refs} +{len(rows) - 6}"

    base = {
        "processed": processed,
        "total": total,
        "worker": worker,
        "elapsed": round(elapsed, 1),
        "rows": row_refs or "—",
        "url": url,
    }

    if status == "Success":
        metrics = scrape_metric_details(data)
        message = (
            f"[{processed}/{total}] OK • {channel} • {format_metric_log_plain(data)} • "
            f"{elapsed:.1f}s • luồng {worker} • dòng {row_refs or '—'}"
        )
        details = {
            "kind": "scrape_ok",
            **base,
            "channel": channel,
            "metrics": metrics,
        }
        return message, "OK", details

    if is_hidden_stats_status(status):
        message = (
            f"[{processed}/{total}] Ẩn số liệu • {elapsed:.1f}s • luồng {worker} • "
            f"TikTok không trả lượt xem • dòng {row_refs or '—'}"
        )
        details = {
            "kind": "scrape_hidden",
            **base,
            "attempts": attempts,
            "status": status,
        }
        return message, "WARN", details

    message = (
        f"[{processed}/{total}] Lỗi • {elapsed:.1f}s • luồng {worker} • "
        f"thử {attempts} lần • {status} • dòng {row_refs or '—'}"
    )
    details = {
        "kind": "scrape_error",
        **base,
        "attempts": attempts,
        "status": status,
    }
    return message, "ERROR", details


def progress_payload(total, processed, success_count, error_count, worker_count, started_at, done=False, mode="full", partner="", phase="scanning", uses_browser=False, hidden_count=0):
    elapsed = max(time.perf_counter() - started_at, 0.001)
    rate = processed / elapsed * 60 if processed else 0
    remaining = max(total - processed, 0)
    eta_seconds = int((remaining / rate) * 60) if rate > 0 else None
    return {
        "total": total,
        "processed": processed,
        "success": success_count,
        "error": error_count,
        "hidden": hidden_count,
        "workers": worker_count,
        "rate": round(rate, 1),
        "etaSeconds": eta_seconds,
        "done": done,
        "mode": mode,
        "partner": partner,
        "phase": phase,
        "usesBrowser": uses_browser,
    }


async def run_scraper(file_path, websocket_manager=None, worker_count=DEFAULT_WORKERS, retries=DEFAULT_RETRIES, save_every=DEFAULT_SAVE_EVERY, selected_partner=None, selected_partners=None, create_result_sheet=False, base_dir=None, file_label="", sheet_name=None, use_request=True, browser_fallback=False, use_proxy=False, proxy_text=""):
    if not os.path.exists(file_path):
        if websocket_manager:
            await websocket_manager.broadcast_log(f"Lỗi: Không tìm thấy file {file_path}")
        return

    retries = clamp_int(retries, DEFAULT_RETRIES, 0, 5)
    save_every = clamp_int(save_every, DEFAULT_SAVE_EVERY, 5, 100)
    selected_names = normalize_selected_partners(selected_partner, selected_partners)
    partner_label = selected_partner_label(selected_names)
    scrape_base_dir = base_dir or os.path.dirname(os.path.abspath(file_path))
    proxy_configs = resolve_proxy_configs(scrape_base_dir, proxy_text=proxy_text) if use_proxy else []
    has_proxy = bool(proxy_configs)
    worker_count = clamp_worker_count(worker_count, proxy_count=len(proxy_configs))
    heavy_proxy_load = has_proxy and (worker_count / len(proxy_configs)) > MAX_WORKERS_PER_PROXY

    workbook = openpyxl.load_workbook(file_path)
    clear_existing_total_rows(workbook)
    sheet_contexts = build_sheet_contexts(workbook, sheet_name=sheet_name)
    rows_to_process = collect_rows(workbook, selected_partners=selected_names, sheet_name=sheet_name)

    # Dedup URL: nhiều dòng cùng URL chỉ scrape 1 lần, ghi kết quả về tất cả các dòng
    unique_buckets = {}
    bucket_order = []
    for item in rows_to_process:
        key = item["url"].strip().casefold()
        bucket = unique_buckets.get(key)
        if bucket is None:
            bucket = {
                "sequence": len(bucket_order) + 1,
                "url": item["url"],
                "rows": [],
                "expected_media_ids": [],
            }
            unique_buckets[key] = bucket
            bucket_order.append(bucket)
        bucket["rows"].append({
            "sheet_name": item["sheet_name"],
            "row": item["row"],
            "partners": item["partners"],
        })
        expected_media_id = clean_text(item.get("expected_media_id", ""))
        if expected_media_id and expected_media_id not in bucket["expected_media_ids"]:
            bucket["expected_media_ids"].append(expected_media_id)

    duplicate_count = len(rows_to_process) - len(bucket_order)
    channel_cache = {}
    channel_overrides = load_channel_overrides(file_path)
    profile_lookup_attempted = set()
    cache_lock = threading.Lock()
    seed_channel_cache_from_workbook(rows_to_process, sheet_contexts, channel_cache)
    total_rows = len(rows_to_process)
    total = len(bucket_order)
    started_at = time.perf_counter()
    mode = "partner" if selected_names else "full"
    if use_request:
        configure_request_concurrency(worker_count, proxy_count=len(proxy_configs))

    # Adaptive save_every: file lớn save thưa hơn để giảm I/O
    if total > 500:
        save_every = max(50, min(100, total // 30))

    if total == 0:
        if websocket_manager:
            await websocket_manager.broadcast_status(progress_payload(0, 0, 0, 0, worker_count, started_at, done=True, mode=mode, partner=partner_label, phase="done"))
            message = "Không tìm thấy link TikTok nào phù hợp để quét."
            if selected_names:
                message = f"Không tìm thấy link nào cho {partner_label}."
            await websocket_manager.broadcast_log(message)
        return

    worker_count = min(worker_count, total)
    active_sheet = clean_text(sheet_name) or (rows_to_process[0].get("sheet_name") if rows_to_process else "")
    if websocket_manager:
        sheet_part = f'Sheet "{active_sheet}" • ' if active_sheet else ""
        await websocket_manager.broadcast_log(
            f"{sheet_part}{total_rows} dòng • {total} URL sau dedup sẽ quét."
        )
        if duplicate_count > 0:
            await websocket_manager.broadcast_log(
                f"Tiết kiệm {duplicate_count} lượt nhờ gộp URL trùng."
            )
        if selected_names:
            await websocket_manager.broadcast_log(
                f"Bắt đầu cập nhật {partner_label}: {total} URL, {worker_count} luồng, retry {retries} lần."
            )
        elif use_request:
            if browser_fallback:
                fallback_count = hybrid_browser_worker_count(worker_count, total)
                await websocket_manager.broadcast_log(
                    f"Bắt đầu quét hybrid {total} URL: {worker_count} luồng Request + {fallback_count} luồng trình duyệt fallback, retry {retries} lần, lưu mỗi {save_every} kết quả."
                )
            else:
                proxy_note = ""
                per_proxy = (worker_count / len(proxy_configs)) if proxy_configs else 0
                if proxy_configs:
                    if len(proxy_configs) == 1:
                        proxy_note = f" • proxy {proxy_label(proxy_configs[0])}"
                    else:
                        proxy_note = (
                            f" • {len(proxy_configs)} proxy, chia đều ~{per_proxy:.1f} luồng/proxy"
                        )
                elif use_proxy:
                    proxy_note = " • proxy: chưa cấu hình"
                if not has_proxy and worker_count > DIRECT_MAX_WORKERS:
                    await websocket_manager.broadcast_log(
                        f"Lưu ý: Không dùng proxy nhưng chạy {worker_count} luồng — tất cả dùng chung 1 IP máy, "
                        f"dễ bị TikTok chặn (khuyến nghị ≤{DIRECT_MAX_WORKERS} luồng khi không có proxy)."
                    )
                elif heavy_proxy_load:
                    await websocket_manager.broadcast_log(
                        f"Lưu ý: {len(proxy_configs)} proxy nhưng {worker_count} luồng — mỗi proxy đang gánh trung bình "
                        f"~{per_proxy:.1f} luồng (khuyến nghị ≤{MAX_WORKERS_PER_PROXY}/proxy). Vẫn chạy đủ {worker_count} luồng; "
                        f"nếu thấy nhiều lỗi HTTP 403, nên thêm proxy hoặc giảm luồng."
                    )
                elif total > 500 and worker_count < 25:
                    await websocket_manager.broadcast_log(
                        f"Gợi ý: Sheet lớn + proxy — thử 25–30 luồng để quét nhanh hơn (hiện {worker_count} luồng)."
                    )
                await websocket_manager.broadcast_log(
                    f"Bắt đầu quét {total} URL bằng Request (HTTP), {worker_count} luồng, retry {retries} lần, lưu mỗi {save_every} kết quả{proxy_note}."
                )
        else:
            await websocket_manager.broadcast_log(
                f"Bắt đầu quét {total} URL bằng {worker_count} luồng trình duyệt, retry {retries} lần, lưu mỗi {save_every} kết quả."
            )
        await websocket_manager.broadcast_status(
            progress_payload(
                total, 0, 0, 0, worker_count, started_at,
                mode=mode, partner=partner_label, phase="starting",
                uses_browser=(not use_request) or browser_fallback,
            )
        )

    work_queue = asyncio.Queue()
    browser_queue = asyncio.Queue()
    result_queue = asyncio.Queue()
    for bucket in bucket_order:
        await work_queue.put(bucket)

    request_worker_count = worker_count if use_request else 0
    browser_worker_count = (
        hybrid_browser_worker_count(worker_count, total)
        if use_request and browser_fallback
        else (worker_count if not use_request else 0)
    )

    if use_request:
        for _ in range(request_worker_count):
            await work_queue.put(None)
    else:
        for _ in range(browser_worker_count):
            await work_queue.put(None)

    processed = 0
    success_count = 0
    error_count = 0
    hidden_count = 0
    pending_save_count = 0
    save_skip_until_processed = 0
    last_status_broadcast = 0.0
    active_worker_count = request_worker_count + browser_worker_count
    completed_sequences = set()

    async with async_playwright() as playwright:
        set_session_proxies(proxy_configs if use_proxy else [])
        if websocket_manager:
            if use_request and browser_fallback:
                await websocket_manager.broadcast_log(
                    f"Đang khởi tạo {request_worker_count} luồng Request và {browser_worker_count} luồng trình duyệt fallback..."
                )
            elif use_request:
                await websocket_manager.broadcast_log(
                    f"Đang khởi tạo {request_worker_count} luồng Request..."
                )
            else:
                await websocket_manager.broadcast_log(f"Đang khởi tạo trình duyệt và {browser_worker_count} luồng...")
        browser = None
        if browser_worker_count > 0:
            browser = await playwright.chromium.launch(
                headless=True,
                args=[
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-features=IsolateOrigins,site-per-process",
                    "--no-sandbox",
                ],
            )
        startup_semaphore = asyncio.Semaphore(min(browser_worker_count, 6)) if browser_worker_count else None
        workers = []

        if use_request:
            workers.extend([
                asyncio.create_task(
                    request_worker_loop(
                        index + 1,
                        work_queue,
                        browser_queue,
                        result_queue,
                        retries,
                        websocket_manager=websocket_manager,
                        browser_fallback=browser_fallback,
                        channel_cache=channel_cache,
                        channel_overrides=channel_overrides,
                        profile_lookup_attempted=profile_lookup_attempted,
                        cache_lock=cache_lock,
                    )
                )
                for index in range(request_worker_count)
            ])

        if browser_worker_count > 0:
            workers.extend([
                asyncio.create_task(
                    worker_loop(
                        index + 1,
                        browser,
                        browser_queue if use_request else work_queue,
                        result_queue,
                        retries,
                        channel_cache=channel_cache,
                        channel_overrides=channel_overrides,
                        profile_lookup_attempted=profile_lookup_attempted,
                        cache_lock=cache_lock,
                        startup_semaphore=startup_semaphore,
                        websocket_manager=websocket_manager,
                        worker_label=f"B{index + 1}" if use_request else None,
                        proxy_configs=proxy_configs,
                    )
                )
                for index in range(browser_worker_count)
            ])
        if websocket_manager:
            await websocket_manager.broadcast_status(
                progress_payload(total, 0, 0, 0, active_worker_count, started_at, mode=mode, partner=partner_label, phase="scanning")
            )

        async def finalize_hybrid_workers():
            if not use_request:
                return
            request_tasks = workers[:request_worker_count]
            browser_tasks = workers[request_worker_count:]
            await asyncio.gather(*request_tasks, return_exceptions=True)
            for _ in range(browser_worker_count):
                await browser_queue.put(None)
            if browser_tasks:
                await asyncio.gather(*browser_tasks, return_exceptions=True)

        finalize_task = asyncio.create_task(finalize_hybrid_workers()) if use_request and browser_fallback and browser_worker_count > 0 else None

        try:
            stalled_seconds = 0
            while processed < total:
                try:
                    result = await asyncio.wait_for(result_queue.get(), timeout=30.0)
                    stalled_seconds = 0
                except asyncio.TimeoutError:
                    stalled_seconds += 30
                    alive_workers = sum(1 for task in workers if not task.done())
                    if alive_workers == 0:
                        remaining_buckets = [
                            bucket for bucket in bucket_order
                            if bucket["sequence"] not in completed_sequences
                        ]
                        if websocket_manager:
                            await websocket_manager.broadcast_log(
                                f"Tất cả worker đã dừng, hủy {len(remaining_buckets)} link còn lại."
                            )
                        empty_data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
                        for bucket in remaining_buckets:
                            error_count += 1
                            processed += 1
                            completed_sequences.add(bucket["sequence"])
                            for target in bucket.get("rows") or []:
                                write_result(
                                    sheet_contexts,
                                    {
                                        "sheet_name": target["sheet_name"],
                                        "row": target["row"],
                                        "url": bucket["url"],
                                    },
                                    empty_data,
                                    "",
                                    "Error: Worker đã dừng",
                                )
                            pending_save_count += max(len(bucket.get("rows") or []), 1)
                        break
                    if websocket_manager and stalled_seconds % 60 == 0:
                        await websocket_manager.broadcast_log(
                            f"Đang chờ kết quả... {alive_workers}/{active_worker_count} luồng còn sống ({processed}/{total})."
                        )
                    continue

                result = guard_result_media_identity(result)
                processed += 1
                data = result["data"]
                status = result["status"]
                bucket_rows = result.get("rows") or []
                completed_sequences.add(result.get("sequence"))
                resolved_url = clean_text(result.get("resolved_url", ""))
                result["channel_name"] = enrich_channel_name(
                    result["url"],
                    result.get("channel_name", ""),
                    resolved_url=resolved_url if status == "Success" else "",
                    channel_cache=channel_cache,
                    channel_overrides=channel_overrides,
                    profile_lookup_attempted=profile_lookup_attempted,
                    cache_lock=cache_lock,
                )

                if status == "Success":
                    success_count += 1
                elif is_hidden_stats_status(status):
                    hidden_count += 1
                else:
                    error_count += 1

                if websocket_manager:
                    log_message, log_level, log_details = format_scrape_result_log(result, processed, total)
                    await websocket_manager.broadcast_log(log_message, level=log_level, details=log_details)

                # Write the same scrape result to every spreadsheet row that shares this URL
                for target in bucket_rows:
                    write_result(
                        sheet_contexts,
                        {
                            "sheet_name": target["sheet_name"],
                            "row": target["row"],
                            "url": result["url"],
                        },
                        data,
                        result["channel_name"],
                        status,
                        resolved_url=resolved_url,
                        channel_cache=channel_cache,
                        channel_overrides=channel_overrides,
                        profile_lookup_attempted=profile_lookup_attempted,
                        cache_lock=cache_lock,
                    )
                pending_save_count += max(len(bucket_rows), 1)

                if websocket_manager:
                    primary_target = bucket_rows[0] if bucket_rows else {"sheet_name": ""}
                    # Chỉ tô cam khi dòng chính (primary) đúng 1 đối tác — khớp Excel/preview.
                    single_partner = len(primary_target.get("partners") or []) == 1
                    video_link = should_highlight_video_link(
                        result["url"],
                        resolved_url=resolved_url,
                        likes=data.get("Likes", 0),
                        shares=data.get("Shares", 0),
                        metrics_readable=status == "Success",
                    )
                    await websocket_manager.broadcast_data({
                        "id": result["sequence"],
                        "url": result["url"],
                        "views": int(metric_number(data.get("Views", 0))),
                        "likes": int(metric_number(data.get("Likes", 0))),
                        "comments": int(metric_number(data.get("Comments", 0))),
                        "saves": int(metric_number(data.get("Saves", 0))),
                        "shares": int(metric_number(data.get("Shares", 0))),
                        "status": status,
                        "worker": result["worker"],
                        "channelName": channel_name_for_sheet(
                            result["url"],
                            result["channel_name"],
                            resolved_url=resolved_url if status == "Success" else "",
                            status=status,
                            channel_cache=channel_cache,
                            channel_overrides=channel_overrides,
                            profile_lookup_attempted=profile_lookup_attempted,
                            cache_lock=cache_lock,
                        ),
                        "sheetName": primary_target.get("sheet_name", ""),
                        "singlePartner": single_partner,
                        "videoLink": video_link,
                    })
                    now = time.perf_counter()
                    if processed == total or now - last_status_broadcast > 0.25:
                        await websocket_manager.broadcast_status(
                            progress_payload(
                                total,
                                processed,
                                success_count,
                                error_count,
                                active_worker_count,
                                started_at,
                                mode=mode,
                                partner=partner_label,
                                phase="scanning",
                                hidden_count=hidden_count,
                            )
                        )
                        last_status_broadcast = now

                if pending_save_count >= save_every and save_skip_until_processed <= processed:
                    saved, save_reason = await save_workbook(workbook, file_path, websocket_manager)
                    if saved:
                        pending_save_count = 0
                        if websocket_manager:
                            await websocket_manager.broadcast_log(f"Đã lưu tạm workbook tại {processed}/{total} links.")
                    else:
                        # Reset counter and skip saving for the next `save_every` items to avoid log spam
                        pending_save_count = 0
                        save_skip_until_processed = processed + save_every

                result_queue.task_done()

            if finalize_task is not None:
                await finalize_task
            else:
                try:
                    await asyncio.wait_for(asyncio.gather(*workers, return_exceptions=True), timeout=20.0)
                except asyncio.TimeoutError:
                    pass
        finally:
            for task in workers:
                if not task.done():
                    task.cancel()
            try:
                await asyncio.wait_for(asyncio.gather(*workers, return_exceptions=True), timeout=10.0)
            except asyncio.TimeoutError:
                pass
            # Close any remaining contexts gracefully before shutting browser down
            if browser is not None:
                try:
                    for ctx in list(browser.contexts):
                        try:
                            await ctx.close()
                        except Exception:
                            pass
                except Exception:
                    pass
                try:
                    await browser.close()
                except Exception:
                    pass
            # Brief grace period so the Playwright Node side can drain pending
            # events instead of hitting EPIPE on the parent shutdown
            try:
                await asyncio.sleep(0.3)
            except Exception:
                pass
            set_session_proxies([])

    scan_sheet_for_summary = clean_text(sheet_name) or (
        clean_text(rows_to_process[0].get("sheet_name", "")) if rows_to_process else ""
    )
    try:
        append_sheet_total_rows(workbook)
        summary_update_time = format_display_datetime()
        created_sheet_name = ""
        if create_result_sheet:
            created_sheet_name = build_result_sheet(workbook, rows_to_process, summary_update_time)
            if websocket_manager and created_sheet_name:
                await websocket_manager.broadcast_log(f"Đã tạo sheet kết quả mới: {created_sheet_name}.")
        summary_count = rebuild_summary_sheet(
            workbook,
            summary_update_time=summary_update_time,
            selected_partners=selected_names,
            data_sheet_name=scan_sheet_for_summary,
        )
        if websocket_manager and scan_sheet_for_summary:
            summary_title = summary_sheet_title_for_data_sheet(scan_sheet_for_summary)
            await websocket_manager.broadcast_log(
                f"Đã cập nhật {summary_title} ({summary_count} đối tác)."
            )
    except Exception as error:
        if websocket_manager:
            await websocket_manager.broadcast_log(f"CẢNH BÁO: Không cập nhật được sheet Tổng kết ({str(error)})")

    try:
        if scan_sheet_for_summary:
            highlighted_count = highlight_single_partner_link_rows(workbook, scan_sheet_for_summary)
            if websocket_manager and highlighted_count:
                await websocket_manager.broadcast_log(
                    f"Đã bôi cam {highlighted_count} dòng link 1 đối tác chưa đủ điều kiện xanh."
                )
            video_highlighted = highlight_video_link_rows(workbook, scan_sheet_for_summary)
            if websocket_manager and video_highlighted:
                await websocket_manager.broadcast_log(
                    f"Đã bôi xanh {video_highlighted} dòng video có hoạt động."
                )
    except Exception as error:
        if websocket_manager:
            await websocket_manager.broadcast_log(
                f"CẢNH BÁO: Không bôi cam được dòng link 1 đối tác ({str(error)})"
            )

    final_saved, final_save_reason = await save_workbook(workbook, file_path, websocket_manager)
    if not final_saved:
        if final_save_reason == "permission":
            raise RuntimeError("Lưu file Excel thất bại vì file đang mở.")
        raise RuntimeError("Lưu file Excel thất bại.")
    duration_seconds = max(int(time.perf_counter() - started_at), 0)
    scan_sheet_name = clean_text(sheet_name) or ""
    if not scan_sheet_name and rows_to_process:
        scan_sheet_name = clean_text(rows_to_process[0].get("sheet_name", ""))
    session_totals = _compute_session_totals(workbook, rows_to_process)
    sheet_totals = _compute_workbook_totals(workbook, sheet_name=scan_sheet_name or None)
    history_entry = {
        "timestamp": format_display_datetime(),
        "fileLabel": file_label or os.path.basename(file_path),
        "scanSheet": scan_sheet_name,
        "scrapedUrls": total,
        "scrapedRows": total_rows,
        "success": success_count,
        "error": error_count,
        "hidden": hidden_count,
        "workers": active_worker_count if use_request else worker_count,
        "durationSeconds": duration_seconds,
        "sheetTotalLinks": sheet_totals["totalLinks"],
        **session_totals,
    }
    append_scrape_history(base_dir or os.path.dirname(os.path.abspath(file_path)), history_entry)
    if websocket_manager:
        await websocket_manager.broadcast_status(
            progress_payload(
                total,
                processed,
                success_count,
                error_count,
                active_worker_count if use_request else worker_count,
                started_at,
                done=True,
                mode=mode,
                partner=partner_label,
                phase="done",
                hidden_count=hidden_count,
            )
        )
        duration_seconds = max(int(time.perf_counter() - started_at), 0)
        summary_counts = (
            f"thành công {success_count}, ẩn số liệu {hidden_count}, lỗi {error_count}"
        )
        if selected_names:
            await websocket_manager.broadcast_log(
                f"HOÀN THÀNH: Đã cập nhật {partner_label} với {processed} URL ({total_rows} dòng), "
                f"{summary_counts}, thời lượng {duration_seconds}s, file={os.path.basename(file_path)}.",
                level="OK",
            )
        else:
            await websocket_manager.broadcast_log(
                f"HOÀN THÀNH: Đã quét {processed}/{total} URL ({total_rows} dòng), "
                f"{summary_counts}, thời lượng {duration_seconds}s, file={os.path.basename(file_path)}.",
                level="OK",
            )
