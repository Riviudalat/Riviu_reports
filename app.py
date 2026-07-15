from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, UploadFile, File, Query
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import asyncio
import io
import json
import os
import re
import zipfile
from datetime import datetime
import urllib.error
from urllib.parse import quote

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.units import pixels_to_EMU

from scraper import run_scraper, read_scrape_history
from google_sheets_sync import (
    authorize_google,
    download_google_sheet_authenticated,
    oauth_status,
    push_rows_to_new_sheet,
    save_oauth_client,
    try_load_credentials,
)
from proxy_utils import (
    load_proxy_list_text,
    parse_proxy_text,
    resolve_proxy_configs,
    save_proxy_list_text,
    test_proxy_text,
)
from workbook_utils import (
    GOOGLE_SHEET_LABEL,
    LEGACY_GOOGLE_SHEET_FILE_ID,
    REPORT_COLUMNS,
    SINGLE_LINK_FILL_COLOR,
    VIDEO_LINK_FILL_COLOR,
    build_workbook_rows,
    clean_text,
    download_google_sheet,
    ensure_data_dir,
    find_data_sheet_names,
    google_sheet_source_for_file,
    is_failed_channel_name,
    is_tiktok_video_link,
    list_workbook_partners,
    list_workbook_partners_with_link_counts,
    normalize_header,
    parse_google_spreadsheet_id,
    read_summary_dashboard,
    read_sheet_preview,
    register_google_sheet_source,
    safe_join,
    metric_number,
    format_metric,
    fetch_google_spreadsheet_title,
    format_display_datetime,
    format_filename_datetime,
    google_sheet_file_id_from_title,
    google_sheet_sync_label,
    workbook_file_entries,
    workbook_sheet_names,
)


app = FastAPI()
templates = Jinja2Templates(directory="templates")

_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
if os.path.isdir(_STATIC_DIR):
    app.mount("/static", StaticFiles(directory=_STATIC_DIR), name="static")


class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast_log(self, message: str, *, level: str = "", details=None):
        payload = {"type": "log", "message": message}
        if level:
            payload["level"] = level
        if details:
            payload["details"] = details
        for conn in self.active_connections:
            try:
                await conn.send_json(payload)
            except Exception:
                pass

    async def broadcast_status(self, status: dict):
        for conn in self.active_connections:
            try:
                await conn.send_json({"type": "status", "data": status})
            except Exception:
                pass

    async def broadcast_data(self, row: dict):
        for conn in self.active_connections:
            try:
                await conn.send_json({"type": "data", "row": row})
            except Exception:
                pass


manager = ConnectionManager()
EXCEL_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = ensure_data_dir(EXCEL_DIR)
SCRAPE_TASK = None
CURRENT_SELECTED_FILE = ""
CURRENT_SELECTED_SHEET = ""
CURRENT_SCAN_SHEET = ""
GOOGLE_SHEET_SOURCE_URL = ""
LOGO_PATH = os.path.join(EXCEL_DIR, "logo.png")


def file_entries():
    return workbook_file_entries(EXCEL_DIR)


def resolve_file_path(file_id):
    if not file_id:
        return ""
    return safe_join(EXCEL_DIR, file_id.replace("\\", "/"))


def ensure_selected_file():
    global CURRENT_SELECTED_FILE, CURRENT_SELECTED_SHEET, CURRENT_SCAN_SHEET
    entries = file_entries()
    available_ids = {entry["id"] for entry in entries}

    if CURRENT_SELECTED_FILE and CURRENT_SELECTED_FILE in available_ids:
        return CURRENT_SELECTED_FILE

    if not entries:
        CURRENT_SELECTED_FILE = ""
        CURRENT_SELECTED_SHEET = ""
        CURRENT_SCAN_SHEET = ""
        return ""

    google_ids = [
        entry["id"]
        for entry in entries
        if entry.get("source") == "google" and entry["id"] != LEGACY_GOOGLE_SHEET_FILE_ID
    ]
    google_ids = sorted(
        google_ids,
        key=lambda file_id: os.path.getmtime(resolve_file_path(file_id)) if os.path.exists(resolve_file_path(file_id)) else 0,
        reverse=True,
    )
    preferred_ids = google_ids + [LEGACY_GOOGLE_SHEET_FILE_ID, "Report_v1.xlsx", "Report_v1_with_partners.xlsx"]
    CURRENT_SELECTED_FILE = next((file_id for file_id in preferred_ids if file_id in available_ids), entries[0]["id"])
    return CURRENT_SELECTED_FILE


def current_excel_path():
    selected = ensure_selected_file()
    return resolve_file_path(selected) if selected else ""


def current_google_sheet_source():
    current_id = ensure_selected_file()
    return google_sheet_source_for_file(EXCEL_DIR, current_id)


def current_display_label():
    current_id = ensure_selected_file()
    for entry in file_entries():
        if entry["id"] == current_id:
            return entry["label"]
    return current_id


def default_sheet_for_file(file_id):
    if not file_id:
        return ""
    try:
        sheets = find_data_sheet_names(resolve_file_path(file_id))
        return sheets[0] if sheets else ""
    except Exception:
        return ""


def sheets_for_current_file():
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return []
    try:
        return find_data_sheet_names(target_path)
    except Exception:
        return []


def resolve_scan_sheet(requested_sheet=""):
    global CURRENT_SCAN_SHEET
    sheets = sheets_for_current_file()
    if not sheets:
        return "", []
    sheet_name = clean_text(requested_sheet) or CURRENT_SCAN_SHEET or CURRENT_SELECTED_SHEET or sheets[0]
    if sheet_name not in sheets:
        raise ValueError("Sheet không tồn tại")
    CURRENT_SCAN_SHEET = sheet_name
    return sheet_name, sheets


def safe_report_name(name):
    filename = re.sub(r'[\\/:*?"<>|]+', "-", clean_text(name))
    filename = re.sub(r"\s+", " ", filename).strip(" .")
    return filename[:90] if filename else "doi_tac"


def spreadsheet_text(value):
    text = clean_text(value)
    return f"'{text}" if text else ""


def spreadsheet_date_text(value):
    if pd.isna(value):
        return ""
    if hasattr(value, "to_pydatetime"):
        return spreadsheet_text(value.to_pydatetime().strftime("%d/%m/%Y"))
    if isinstance(value, datetime):
        return spreadsheet_text(value.strftime("%d/%m/%Y"))
    parsed = pd.to_datetime(value, errors="coerce")
    if not pd.isna(parsed):
        return spreadsheet_text(parsed.to_pydatetime().strftime("%d/%m/%Y"))
    text = clean_text(value)
    if text.endswith(" 00:00:00"):
        text = text[:10]
    return spreadsheet_text(text)


def spreadsheet_formula_text(value):
    return clean_text(value).replace('"', '""')


def spreadsheet_hyperlink(value):
    return clean_text(value)


def excel_column_name(index):
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def column_width_to_pixels(width):
    if not width:
        width = 8.43
    return int(width * 7 + 5)


def row_height_to_pixels(height):
    if not height:
        height = 15
    return int(height * 96 / 72)


def add_centered_image_to_cell(ws, cell_address, image_path, max_height_px=36):
    if not image_path or not os.path.exists(image_path):
        return False
    try:
        img = ExcelImage(image_path)
    except Exception:
        return False

    col_letter, row_number = coordinate_from_string(cell_address)
    col_idx = ord(col_letter.upper()) - ord("A")
    row_idx = row_number - 1

    if img.height > max_height_px:
        scale = max_height_px / img.height
        img.height = max_height_px
        img.width = int(img.width * scale)

    cell_width_px = column_width_to_pixels(ws.column_dimensions[col_letter].width)
    max_width_px = max(cell_width_px - 4, 24)
    if img.width > max_width_px:
        scale = max_width_px / img.width
        img.width = max_width_px
        img.height = int(img.height * scale)

    cell_height_px = row_height_to_pixels(ws.row_dimensions[row_number].height)
    col_off = pixels_to_EMU(max((cell_width_px - img.width) / 2, 0))
    row_off = pixels_to_EMU(max((cell_height_px - img.height) / 2, 0))
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col_idx, colOff=col_off, row=row_idx, rowOff=row_off),
        ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)),
    )
    ws.add_image(img)
    return True


def find_report_column(frame, header):
    target = normalize_header(header)
    for column in frame.columns:
        if normalize_header(column) == target:
            return column
    if header == "LINK AIR":
        for column in frame.columns:
            name = normalize_header(column)
            if "link" in name or "url" in name:
                return column
    if header == "TÊN KÊNH":
        for column in frame.columns:
            name = normalize_header(column)
            if "tên kênh" in name or "ten kenh" in name:
                return column
    if header == "NGÀY AIR":
        for column in frame.columns:
            name = normalize_header(column)
            if "ngày" in name or "ngay" in name:
                return column
    return None


def build_partner_report(partner, rows, *, apply_min_views=True, min_views=100):
    threshold = max(int(min_views or 0), 0) if apply_min_views else 0
    if apply_min_views:
        valid_rows = [
            row
            for row in rows
            if not is_failed_channel_name(row.get("TÊN KÊNH", ""))
            and metric_number(row.get("LƯỢT XEM", 0)) >= threshold
        ]
    else:
        valid_rows = [
            row
            for row in rows
            if not is_failed_channel_name(row.get("TÊN KÊNH", ""))
        ]
    frame = pd.DataFrame(valid_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"
    updated_at = format_display_datetime()
    table_header_row = 6
    data_start_row = table_header_row + 1
    last_column = get_column_letter(len(REPORT_COLUMNS))

    title_fill = PatternFill("solid", fgColor="FF6B00")
    header_fill = PatternFill("solid", fgColor="FF6B00")
    total_fill = PatternFill("solid", fgColor="FFF7ED")
    thin = Side(style="thin", color="D8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 22
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    add_centered_image_to_cell(ws, "A1", LOGO_PATH)

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(REPORT_COLUMNS))
    title_cell = ws["B1"]
    title_cell.value = f"BÁO CÁO ĐỐI TÁC: {partner}"
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REPORT_COLUMNS))
    ws["A2"] = f"Tổng link: {len(frame)} • Ngày cập nhật: {updated_at}"
    ws["A2"].font = Font(color="9A3412", italic=True, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    for col_index, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=table_header_row, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    single_link_fill = PatternFill("solid", fgColor=SINGLE_LINK_FILL_COLOR)
    video_link_fill = PatternFill("solid", fgColor=VIDEO_LINK_FILL_COLOR)
    for row_index, (_, source_row) in enumerate(frame.iterrows(), start=data_start_row):
        for col_index, header in enumerate(REPORT_COLUMNS, start=1):
            value = source_row.get(header, "")
            if header == "LINK AIR":
                value = clean_text(value)
            elif header == "TÊN KÊNH":
                value = clean_text(value)
            elif header == "NGÀY AIR":
                if pd.isna(value):
                    value = ""
                elif hasattr(value, "to_pydatetime"):
                    value = value.to_pydatetime()
                else:
                    parsed_date = pd.to_datetime(value, errors="coerce")
                    value = parsed_date.to_pydatetime() if not pd.isna(parsed_date) else clean_text(value)
            else:
                value = format_metric(value)

            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(header in {"LINK AIR", "TÊN KÊNH"}))
            if header == "LINK AIR" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
            elif header == "NGÀY AIR":
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif header not in {"LINK AIR", "TÊN KÊNH"}:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")

        row_partners = source_row.get("partners")
        if not isinstance(row_partners, (list, tuple)):
            row_partners = []
        link_for_type = clean_text(source_row.get("LINK AIR", ""))
        row_fill = None
        if is_tiktok_video_link(link_for_type):
            row_fill = video_link_fill
        elif len(row_partners) == 1:
            row_fill = single_link_fill
        if row_fill is not None:
            for col_index in range(1, len(REPORT_COLUMNS) + 1):
                ws.cell(row=row_index, column=col_index).fill = row_fill

    total_row = None
    if len(frame) > 0:
        total_row = len(frame) + data_start_row
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        for col_index in range(1, len(REPORT_COLUMNS) + 1):
            cell = ws.cell(row=total_row, column=col_index)
            cell.fill = total_fill
            cell.border = border
        total_label = ws.cell(row=total_row, column=1, value="TỔNG")
        total_label.font = Font(bold=True, color="9A3412")
        total_label.alignment = Alignment(horizontal="center", vertical="center")
        for header in REPORT_COLUMNS[3:]:
            if header not in frame.columns:
                continue
            col_index = REPORT_COLUMNS.index(header) + 1
            total_value = int(frame[header].map(metric_number).sum())
            cell = ws.cell(row=total_row, column=col_index, value=total_value)
            cell.fill = total_fill
            cell.font = Font(bold=True, color="9A3412")
            cell.border = border
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = f"A{data_start_row}"
    filter_last_row = (total_row - 1) if total_row else max(data_start_row, table_header_row)
    ws.auto_filter.ref = f"A{table_header_row}:{last_column}{filter_last_row}"
    widths = [14, 24, 72, 14, 12, 14, 14, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def content_disposition(filename):
    return f"attachment; filename*=UTF-8''{quote(filename)}"


def validate_proxy_start(use_proxy: bool, proxy_text: str, base_dir: str) -> str | None:
    if not use_proxy:
        return None
    if resolve_proxy_configs(base_dir, proxy_text):
        return None
    return "Bật proxy nhưng chưa có proxy hợp lệ. Mở Cấu hình và dán proxy."


def static_asset_version(base_dir: str) -> str:
    paths = [
        os.path.join(base_dir, "static", "app.js"),
        os.path.join(base_dir, "static", "styles.css"),
    ]
    mtimes = [int(os.path.getmtime(path)) for path in paths if os.path.exists(path)]
    return str(max(mtimes)) if mtimes else "1"


def sync_download_google_sheet(base_dir, source_url, spreadsheet_id, destination_path):
    creds = try_load_credentials(base_dir)
    if creds and creds.valid:
        try:
            download_google_sheet_authenticated(base_dir, spreadsheet_id, destination_path)
            return
        except Exception:
            pass
    try:
        download_google_sheet(source_url, destination_path)
    except Exception as public_error:
        if creds and creds.valid:
            raise public_error
        message = str(public_error)
        if isinstance(public_error, urllib.error.HTTPError) and public_error.code in (401, 403):
            raise ValueError("Sheet riêng tư — cần đăng nhập Google hợp lệ.") from public_error
        if "401" in message or "403" in message or "private" in message.lower():
            raise ValueError("Sheet riêng tư — cần đăng nhập Google hợp lệ.") from public_error
        raise


async def run_scraper_safely(target_path, worker_count, partner=None, partners=None, create_result_sheet=False, push_to_google=False, sheet_name="", use_request=True, browser_fallback=False, use_proxy=False, proxy_text=""):
    try:
        scan_sheet, _ = resolve_scan_sheet(sheet_name)
        await run_scraper(
            target_path,
            manager,
            worker_count=worker_count,
            selected_partner=partner,
            selected_partners=partners,
            create_result_sheet=create_result_sheet,
            base_dir=EXCEL_DIR,
            file_label=current_display_label(),
            sheet_name=scan_sheet,
            use_request=use_request,
            browser_fallback=browser_fallback,
            use_proxy=use_proxy,
            proxy_text=proxy_text,
        )
        if push_to_google:
            source = current_google_sheet_source()
            spreadsheet_id = source.get("spreadsheetId", "")
            if spreadsheet_id:
                rows = build_workbook_rows(target_path, sheet_name=scan_sheet)
                values = build_google_push_rows(rows)
                sheet_title = push_rows_to_new_sheet(EXCEL_DIR, spreadsheet_id, values)
                await manager.broadcast_log(f"Đã đẩy kết quả lên Google Sheet: {sheet_title}.")
            else:
                await manager.broadcast_log("BỎ QUA đẩy Google: file hiện tại chưa gắn với Google Sheet gốc.")
    except asyncio.CancelledError:
        await manager.broadcast_log("Đã hủy phiên quét.")
        await manager.broadcast_status({"total": 0, "processed": 0, "success": 0, "error": 0, "done": True, "cancelled": True})
        raise
    except Exception as error:
        await manager.broadcast_log(f"Lỗi quét: {str(error)}")
        await manager.broadcast_status({"total": 0, "processed": 0, "success": 0, "error": 1, "done": True})


def build_google_push_rows(rows):
    normalized_rows = []
    max_partner_columns = 1
    for row in rows:
        partner_names = row.get("partners", [])
        if not isinstance(partner_names, list):
            partner_names = []
        partner_names = [clean_text(name) for name in partner_names if clean_text(name)]
        max_partner_columns = max(max_partner_columns, len(partner_names))
        normalized_rows.append((row, partner_names))

    headers = ["Stt", "Ngày", "Link", "Tên Kênh", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"]
    headers.extend(["Đối tác" if index == 0 else f"Đối tác {index + 1}" for index in range(max_partner_columns)])
    headers.append("Cập nhật lần cuối")

    values = [headers]
    for index, (row, partner_names) in enumerate(normalized_rows, start=1):
        padded_partners = partner_names + [""] * (max_partner_columns - len(partner_names))
        channel_name = clean_text(row.get("TÊN KÊNH", ""))
        if is_failed_channel_name(channel_name):
            channel_name = "Lỗi"
        values.append([
            index,
            spreadsheet_date_text(row.get("NGÀY AIR", "")),
            spreadsheet_hyperlink(row.get("LINK AIR", "")),
            channel_name,
            metric_number(row.get("LƯỢT XEM", 0)),
            metric_number(row.get("TIM", 0)),
            metric_number(row.get("BÌNH LUẬN", 0)),
            metric_number(row.get("LƯỢT LƯU", 0)),
            metric_number(row.get("CHIA SẺ", 0)),
            *padded_partners,
            format_display_datetime(),
        ])
    if len(values) > 1:
        total_row_number = len(values) + 1
        data_last_row = total_row_number - 1
        total_row = ["", "", "TỔNG", ""]
        for column_index in range(5, 10):
            column_name = excel_column_name(column_index)
            total_row.append(f"=SUM({column_name}2:{column_name}{data_last_row})")
        total_row.extend([""] * (max_partner_columns + 1))
        values.append(total_row)
    return values



def build_export_payload(target_path, selected_partners, apply_min_views, min_views, requested_sheet_name=""):
    data_sheets = find_data_sheet_names(target_path)
    report_sheet = clean_text(requested_sheet_name) or (data_sheets[0] if data_sheets else "")
    if report_sheet and report_sheet not in data_sheets:
        raise ValueError(f"Sheet {report_sheet} không tồn tại trong file.")
    available_partners = set(list_workbook_partners(target_path, sheet_name=report_sheet))
    partners = [clean_text(name) for name in selected_partners if clean_text(name) in available_partners]
    if not partners:
        raise ValueError("Không tìm thấy đối tác đã chọn trong file")
    export_timestamp = format_filename_datetime()
    sheet_label = safe_report_name(report_sheet) if report_sheet else ""

    def report_base_name(partner_name):
        parts = [safe_report_name(partner_name)]
        if sheet_label:
            parts.append(sheet_label)
        parts.append(export_timestamp)
        return " ".join(parts)

    if len(partners) == 1:
        partner = partners[0]
        report_rows = build_workbook_rows(target_path, selected_partner=partner, sheet_name=report_sheet)
        report_bytes = build_partner_report(
            partner,
            report_rows,
            apply_min_views=apply_min_views,
            min_views=min_views,
        )
        filename = f"{report_base_name(partner)}.xlsx"
        return {
            "content": report_bytes,
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": filename,
        }

    archive = io.BytesIO()
    used_names = set()
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for partner in partners:
            report_rows = build_workbook_rows(target_path, selected_partner=partner, sheet_name=report_sheet)
            report_bytes = build_partner_report(
                partner,
                report_rows,
                apply_min_views=apply_min_views,
                min_views=min_views,
            )
            base_name = report_base_name(partner)
            filename = f"{base_name}.xlsx"
            counter = 2
            while filename in used_names:
                filename = f"{base_name}_{counter}.xlsx"
                counter += 1
            used_names.add(filename)
            zip_file.writestr(filename, report_bytes)

    archive.seek(0)
    zip_parts = ["bao_cao_doi_tac"]
    if sheet_label:
        zip_parts.append(sheet_label)
    zip_parts.append(export_timestamp)
    return {
        "content": archive.getvalue(),
        "media_type": "application/zip",
        "filename": f"{' '.join(zip_parts)}.zip",
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"asset_version": static_asset_version(EXCEL_DIR)},
    )


@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return Response(status_code=204)


@app.get("/logo.png", include_in_schema=False)
async def logo_png():
    if os.path.exists(LOGO_PATH):
        return FileResponse(LOGO_PATH, media_type="image/png")
    return Response(status_code=404)


@app.get("/riviu-logo.png", include_in_schema=False)
async def riviu_logo():
    return await logo_png()


@app.get("/list-files")
async def list_files():
    ensure_selected_file()
    oauth = oauth_status(EXCEL_DIR)
    google_source = current_google_sheet_source()
    target_sheet_url = GOOGLE_SHEET_SOURCE_URL or google_source.get("url", "")
    target_spreadsheet_id = ""
    if target_sheet_url:
        try:
            target_spreadsheet_id = parse_google_spreadsheet_id(target_sheet_url)
        except Exception:
            target_spreadsheet_id = ""
    return {
        "files": file_entries(),
        "current": CURRENT_SELECTED_FILE,
        "currentLabel": current_display_label(),
        "currentSheet": CURRENT_SELECTED_SHEET,
        "sheets": sheets_for_current_file(),
        "scanSheet": CURRENT_SCAN_SHEET or CURRENT_SELECTED_SHEET,
        "googleSheetUrl": target_sheet_url,
        "googlePushReady": bool(target_spreadsheet_id) and oauth.get("valid"),
        "googleOAuthConfigured": oauth.get("configured"),
        "googleOAuthAuthorized": oauth.get("valid"),
    }


@app.post("/select-file")
async def select_file(data: dict):
    global CURRENT_SELECTED_FILE, CURRENT_SELECTED_SHEET, CURRENT_SCAN_SHEET
    file_id = data.get("filename")
    target_path = resolve_file_path(file_id) if file_id else ""
    if target_path and os.path.exists(target_path):
        CURRENT_SELECTED_FILE = file_id.replace("\\", "/")
        CURRENT_SELECTED_SHEET = default_sheet_for_file(CURRENT_SELECTED_FILE)
        CURRENT_SCAN_SHEET = CURRENT_SELECTED_SHEET
        return {"success": True, "selected": CURRENT_SELECTED_FILE, "sheet": CURRENT_SELECTED_SHEET, "scanSheet": CURRENT_SCAN_SHEET}
    return {"success": False, "error": "File không tồn tại"}


@app.post("/sync-google-sheet")
async def sync_google_sheet(data: dict):
    global CURRENT_SELECTED_FILE, CURRENT_SELECTED_SHEET, CURRENT_SCAN_SHEET, GOOGLE_SHEET_SOURCE_URL
    source_url = (data.get("url") or "").strip()
    if not source_url:
        return JSONResponse(content={"error": "Vui lòng nhập URL Google Sheet"}, status_code=400)

    try:
        spreadsheet_id = parse_google_spreadsheet_id(source_url)
        spreadsheet_title = fetch_google_spreadsheet_title(source_url)
        timestamp_filename = format_filename_datetime()
        file_label = google_sheet_sync_label(spreadsheet_title)
        file_id = google_sheet_file_id_from_title(spreadsheet_title, timestamp_filename)
        target_path = resolve_file_path(file_id)
        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        sync_download_google_sheet(EXCEL_DIR, source_url, spreadsheet_id, target_path)
        preview = read_sheet_preview(target_path)
        GOOGLE_SHEET_SOURCE_URL = source_url
        CURRENT_SELECTED_FILE = file_id
        CURRENT_SELECTED_SHEET = preview.get("currentSheet", "")
        CURRENT_SCAN_SHEET = CURRENT_SELECTED_SHEET
        register_google_sheet_source(EXCEL_DIR, file_id, source_url, title=file_label)
        await manager.broadcast_log(
            f"Đã nạp Google Sheet thành file mới: {file_label} → {os.path.basename(file_id)} • sheet={CURRENT_SELECTED_SHEET or ''} • url={source_url}",
            level="OK",
        )
        return {
            "success": True,
            "file": file_id,
            "label": file_label,
            "sheets": preview.get("sheets", []),
            "currentSheet": CURRENT_SELECTED_SHEET,
            "scanSheet": CURRENT_SCAN_SHEET,
            "spreadsheetId": spreadsheet_id,
        }
    except Exception as error:
        return JSONResponse(content={"error": f"Lỗi đồng bộ Google Sheet: {str(error)}"}, status_code=500)


@app.get("/google-oauth-status")
async def google_oauth_status():
    source = current_google_sheet_source()
    target_url = GOOGLE_SHEET_SOURCE_URL or source.get("url", "")
    target_id = ""
    if target_url:
        try:
            target_id = parse_google_spreadsheet_id(target_url)
        except Exception:
            target_id = ""
    status = oauth_status(EXCEL_DIR)
    return {
        **status,
        "connectedSheetUrl": target_url,
        "connectedSpreadsheetId": target_id,
    }


@app.get("/proxy-list")
async def get_proxy_list():
    text = load_proxy_list_text(EXCEL_DIR)
    configs = parse_proxy_text(text)
    return {
        "text": text,
        "count": len(configs),
    }


@app.post("/proxy-list")
async def save_proxy_list(data: dict):
    text = str(data.get("text") or "")
    save_proxy_list_text(EXCEL_DIR, text)
    configs = parse_proxy_text(text)
    return {
        "ok": True,
        "count": len(configs),
        "message": f"Đã lưu {len(configs)} proxy" if configs else "Đã lưu danh sách (chưa đọc được proxy hợp lệ)",
    }


@app.get("/api/version")
async def api_version():
    from proxy_utils import PROXY_TEST_BUILD
    from scraper import MAX_WORKERS

    return {
        "proxyTestBuild": PROXY_TEST_BUILD,
        "app": "riviu-reports",
        "maxWorkers": MAX_WORKERS,
    }


@app.post("/proxy-test")
async def test_proxy_list(data: dict | None = None):
    payload = data or {}
    text = str(payload.get("text") or load_proxy_list_text(EXCEL_DIR))
    if payload.get("save"):
        save_proxy_list_text(EXCEL_DIR, text)
    return test_proxy_text(text)


@app.post("/google-oauth-client")
async def upload_google_oauth_client(file: UploadFile = File(...)):
    try:
        if not file.filename.lower().endswith(".json"):
            return JSONResponse(content={"error": "Chỉ nhận file JSON OAuth client."}, status_code=400)
        payload = json.loads((await file.read()).decode("utf-8"))
        save_oauth_client(EXCEL_DIR, payload)
        return {"success": True}
    except Exception as error:
        return JSONResponse(content={"error": f"Không lưu được OAuth client: {str(error)}"}, status_code=500)


@app.post("/google-oauth-login")
async def google_oauth_login():
    try:
        authorize_google(EXCEL_DIR)
        return {"success": True}
    except Exception as error:
        return JSONResponse(content={"error": f"Đăng nhập Google thất bại: {str(error)}"}, status_code=500)


@app.post("/push-google-sheet")
async def push_google_sheet(data: dict | None = None):
    global GOOGLE_SHEET_SOURCE_URL
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return JSONResponse(content={"error": "File không tồn tại"}, status_code=404)

    request_url = clean_text((data or {}).get("url", ""))
    source_url = request_url or GOOGLE_SHEET_SOURCE_URL or current_google_sheet_source().get("url", "")
    spreadsheet_id = ""
    try:
        spreadsheet_id = parse_google_spreadsheet_id(source_url)
    except Exception:
        spreadsheet_id = ""
    if not spreadsheet_id:
        return JSONResponse(content={"error": "Vui lòng nhập đúng URL Google Sheet đích trước khi tạo sheet."}, status_code=400)

    try:
        GOOGLE_SHEET_SOURCE_URL = source_url
        register_google_sheet_source(EXCEL_DIR, ensure_selected_file(), source_url)
        upload_sheet = clean_text((data or {}).get("sourceSheet", "")) or CURRENT_SCAN_SHEET or CURRENT_SELECTED_SHEET
        available_sheets = find_data_sheet_names(target_path)
        if upload_sheet and upload_sheet not in available_sheets:
            return JSONResponse(content={"error": f"Sheet {upload_sheet} không tồn tại trong file."}, status_code=400)
        rows = build_workbook_rows(target_path, sheet_name=upload_sheet)
        if not rows:
            return JSONResponse(content={"error": f"Sheet \"{upload_sheet or 'đang chọn'}\" không có link TikTok để tạo sheet."}, status_code=400)
        values = build_google_push_rows(rows)
        sheet_title = push_rows_to_new_sheet(EXCEL_DIR, spreadsheet_id, values)
        return {"success": True, "sheetTitle": sheet_title, "sourceSheet": upload_sheet}
    except Exception as error:
        return JSONResponse(content={"error": f"Đẩy dữ liệu lên Google Sheet thất bại: {str(error)}"}, status_code=500)


@app.get("/preview-excel")
async def preview_excel(sheet_name: str = Query(default="")):
    global CURRENT_SELECTED_SHEET
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return {"sheets": [], "currentSheet": "", "columns": [], "data": [], "message": f"Không tìm thấy file {CURRENT_SELECTED_FILE}."}
    try:
        requested_sheet = sheet_name or CURRENT_SELECTED_SHEET or default_sheet_for_file(CURRENT_SELECTED_FILE)

        def _load_preview():
            return read_sheet_preview(target_path, sheet_name=requested_sheet)

        preview = await asyncio.to_thread(_load_preview)
        CURRENT_SELECTED_SHEET = preview.get("currentSheet", CURRENT_SELECTED_SHEET)
        preview["file"] = CURRENT_SELECTED_FILE
        preview["fileLabel"] = current_display_label()
        return preview
    except Exception as error:
        return {"sheets": [], "currentSheet": "", "columns": [], "data": [], "message": f"File đang bận hoặc lỗi: {str(error)}"}


@app.get("/summary-dashboard")
async def summary_dashboard(sheet_name: str = Query(default="")):
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return JSONResponse(content={"error": "File không tồn tại"}, status_code=404)
    try:
        requested_sheet = clean_text(sheet_name) or CURRENT_SCAN_SHEET or CURRENT_SELECTED_SHEET or ""
        summary = await asyncio.to_thread(read_summary_dashboard, target_path, requested_sheet or None)
        summary["file"] = CURRENT_SELECTED_FILE
        summary["fileLabel"] = current_display_label()
        return summary
    except Exception as error:
        return JSONResponse(content={"error": f"Không đọc được sheet Tổng kết: {str(error)}"}, status_code=500)


@app.get("/download-excel")
async def download_excel():
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return JSONResponse(content={"error": "File không tồn tại"}, status_code=404)
    download_name = os.path.basename(CURRENT_SELECTED_FILE)
    return FileResponse(
        path=target_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/scrape-history")
async def scrape_history(limit: int = Query(default=50)):
    safe_limit = max(1, min(int(limit or 50), 200))
    return {"history": read_scrape_history(EXCEL_DIR, limit=safe_limit)}


@app.get("/report-partners")
async def report_partners(
    sheet_name: str = Query(default=""),
    apply_min_views: bool = Query(default=True),
    min_views: int = Query(default=100),
):
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return JSONResponse(content={"error": "File không tồn tại"}, status_code=404)

    try:
        data_sheets = await asyncio.to_thread(find_data_sheet_names, target_path)
        requested_sheet = clean_text(sheet_name) or (data_sheets[0] if data_sheets else "")
        if requested_sheet and requested_sheet not in data_sheets:
            return JSONResponse(content={"error": f"Sheet {requested_sheet} không tồn tại trong file."}, status_code=400)
        safe_min_views = max(int(min_views or 0), 0)

        def _load_partners():
            return list_workbook_partners_with_link_counts(
                target_path,
                sheet_name=requested_sheet,
                apply_min_views=apply_min_views,
                min_views=safe_min_views,
            )

        partners = await asyncio.to_thread(_load_partners)
        all_sheets = await asyncio.to_thread(workbook_sheet_names, target_path)
        return {
            "partners": partners,
            "total": len(partners),
            "file": CURRENT_SELECTED_FILE,
            "fileLabel": current_display_label(),
            "sheets": data_sheets,
            "currentSheet": requested_sheet,
            "dataSheet": requested_sheet,
            "allSheets": all_sheets,
        }
    except Exception as error:
        return JSONResponse(content={"error": f"Không đọc được danh sách đối tác: {str(error)}"}, status_code=500)


@app.post("/export-report")
async def export_report(data: dict):
    target_path = current_excel_path()
    if not os.path.exists(target_path):
        return JSONResponse(content={"error": "File không tồn tại"}, status_code=404)

    selected_partners = data.get("partners") or []
    if not isinstance(selected_partners, list) or not selected_partners:
        return JSONResponse(content={"error": "Vui lòng chọn ít nhất một đối tác"}, status_code=400)

    apply_min_views = bool(data.get("applyMinViews", True))
    try:
        min_views = int(data.get("minViews", 100))
    except (TypeError, ValueError):
        min_views = 100
    min_views = max(min_views, 0)

    try:
        payload = await asyncio.to_thread(
            build_export_payload,
            target_path,
            selected_partners,
            apply_min_views,
            min_views,
            data.get("sheetName", ""),
        )
        return Response(
            content=payload["content"],
            media_type=payload["media_type"],
            headers={"Content-Disposition": content_disposition(payload["filename"])},
        )
    except ValueError as error:
        status_code = 404 if "Không tìm thấy" in str(error) else 400
        return JSONResponse(content={"error": str(error)}, status_code=status_code)
    except Exception as error:
        return JSONResponse(content={"error": f"Lỗi xuất báo cáo: {str(error)}"}, status_code=500)


@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    global CURRENT_SELECTED_FILE, CURRENT_SELECTED_SHEET, CURRENT_SCAN_SHEET
    try:
        original_name = os.path.basename(file.filename or "")
        if not (original_name.endswith(".xlsx") or original_name.endswith(".xls")):
            return {"success": False, "error": "Chỉ chấp nhận file Excel (.xlsx, .xls)"}

        save_path = safe_join(EXCEL_DIR, original_name)
        if not save_path:
            return {"success": False, "error": "Tên file không hợp lệ"}

        with open(save_path, "wb") as file_obj:
            content = await file.read()
            file_obj.write(content)

        file_id = original_name.replace("\\", "/")
        CURRENT_SELECTED_FILE = file_id
        CURRENT_SELECTED_SHEET = default_sheet_for_file(file_id)
        CURRENT_SCAN_SHEET = CURRENT_SELECTED_SHEET
        return {"success": True, "filename": file_id, "sheet": CURRENT_SELECTED_SHEET, "scanSheet": CURRENT_SCAN_SHEET}
    except Exception as error:
        return {"success": False, "error": str(error)}


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    global SCRAPE_TASK
    await manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            try:
                payload = json.loads(data)
            except json.JSONDecodeError:
                payload = {"action": data}

            action = payload.get("action")
            if action == "start":
                if SCRAPE_TASK and not SCRAPE_TASK.done():
                    await manager.broadcast_log("Đang có phiên quét chạy. Vui lòng chờ hoàn tất trước khi chạy phiên mới.")
                    continue

                target_path = current_excel_path()
                worker_count = payload.get("workers", 20)
                create_result_sheet = bool(payload.get("create_result_sheet", False))
                push_to_google = bool(payload.get("push_to_google", False))
                partner = clean_text(payload.get("partner", ""))
                sheet_name = clean_text(payload.get("sheet_name", ""))
                try:
                    sheet_name, _ = resolve_scan_sheet(sheet_name)
                except ValueError as error:
                    await manager.broadcast_log(str(error))
                    await manager.broadcast_status({"total": 0, "processed": 0, "success": 0, "error": 1, "done": True})
                    continue
                partners_payload = payload.get("partners", [])
                partners = []
                if isinstance(partners_payload, list):
                    partners = [clean_text(name) for name in partners_payload if clean_text(name)]
                scrape_mode = clean_text(payload.get("scrape_mode", "")).lower()
                if scrape_mode == "browser":
                    use_request = False
                    browser_fallback = False
                elif scrape_mode == "hybrid":
                    use_request = True
                    browser_fallback = True
                else:
                    use_request = True
                    browser_fallback = False
                use_proxy = bool(payload.get("use_proxy", False))
                proxy_text = str(payload.get("proxy_text") or "")
                proxy_error = validate_proxy_start(use_proxy, proxy_text, EXCEL_DIR)
                if proxy_error:
                    await manager.broadcast_log(proxy_error)
                    await manager.broadcast_status({"total": 0, "processed": 0, "success": 0, "error": 1, "done": True})
                    continue
                from scraper import clamp_worker_count
                from proxy_utils import resolve_proxy_configs
                proxy_count = len(resolve_proxy_configs(EXCEL_DIR, proxy_text=proxy_text)) if use_proxy else 0
                worker_count = clamp_worker_count(worker_count, proxy_count=proxy_count)
                SCRAPE_TASK = asyncio.create_task(
                    run_scraper_safely(
                        target_path,
                        worker_count,
                        partner=partner or None,
                        partners=partners,
                        create_result_sheet=create_result_sheet,
                        push_to_google=push_to_google,
                        sheet_name=sheet_name,
                        use_request=use_request,
                        browser_fallback=browser_fallback,
                        use_proxy=use_proxy,
                        proxy_text=proxy_text,
                    )
                )
            elif action == "cancel":
                if SCRAPE_TASK and not SCRAPE_TASK.done():
                    SCRAPE_TASK.cancel()
                    await manager.broadcast_log("Đang hủy phiên quét hiện tại...")
                else:
                    await manager.broadcast_log("Không có phiên quét nào đang chạy.")
    except WebSocketDisconnect:
        manager.disconnect(websocket)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=1231)
