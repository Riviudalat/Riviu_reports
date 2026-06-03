import openpyxl

from scraper import _compute_session_totals, collect_rows


def _build_sample_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tháng 6"
    ws.append(["Ngày", "LINK AIR", "TÊN KÊNH", "LƯỢT XEM", "TIM", "Đối tác"])
    ws.append(["", "https://www.tiktok.com/@a/video/1", "A", 100, 10, "Partner A"])
    ws.append(["", "tiktok.com/@b/video/2", "B", 200, 20, "Partner B"])
    ws.append(["", "TỔNG", "", 300, 30, ""])
    wb.save(path)
    wb.close()


def test_collect_rows_skips_total_row(tmp_path):
    file_path = tmp_path / "sample.xlsx"
    _build_sample_workbook(file_path)
    wb = openpyxl.load_workbook(file_path)
    rows = collect_rows(wb, sheet_name="Tháng 6")
    wb.close()
    assert len(rows) == 2
    assert all("TỔNG" not in row["url"] for row in rows)


def test_compute_session_totals_sums_only_session_rows(tmp_path):
    file_path = tmp_path / "sample.xlsx"
    _build_sample_workbook(file_path)
    wb = openpyxl.load_workbook(file_path)
    rows = collect_rows(wb, sheet_name="Tháng 6")
    totals = _compute_session_totals(wb, rows)
    wb.close()
    assert totals["totalLinks"] == 2
    assert totals["totalViews"] == 300
    assert totals["totalLikes"] == 30
