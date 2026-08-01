import pytest

from scraper import (
    channel_name_for_sheet,
    channel_name_quality,
    counts_match,
    enrich_channel_name,
    extract_media_id,
    fetch_profile_channel_name_request,
    format_metric_log_line,
    format_scrape_result_log,
    is_tiktok_error_page,
    parse_count_value,
    parse_counts,
    validate_metrics,
)


def test_parse_counts_extracts_metrics_from_reflow_detail_json():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.reflow.video.detail": {
          "statusCode": 0,
          "itemInfo": {
            "itemStruct": {
              "id": "555",
              "statsV2": {
                "playCount": "4321",
                "diggCount": "99",
                "commentCount": "4",
                "collectCount": "2",
                "shareCount": "1"
              }
            }
          }
        }
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="555")
    assert found is True
    assert data["Views"] == "4321"
    assert data["Likes"] == "99"


def test_build_request_url_candidates_includes_video_rewrite():
    from scraper import build_request_url_candidates

    candidates = build_request_url_candidates(
        "https://www.tiktok.com/@demo/photo/764002"
    )
    assert "https://www.tiktok.com/@demo/photo/764002" in candidates
    assert "https://www.tiktok.com/@demo/video/764002" in candidates
    assert "https://www.tiktok.com/@demo/photo/764002?_r=1" in candidates
    assert candidates.index("https://www.tiktok.com/@demo/photo/764002?_r=1") < candidates.index(
        "https://www.tiktok.com/@demo/video/764002"
    )


def test_build_duplicate_link_payload_lists_all_sheet_rows():
    import scraper

    assert hasattr(scraper, "build_duplicate_link_payload")
    payload = scraper.build_duplicate_link_payload([
        {
            "url": "https://www.tiktok.com/@demo/video/123",
            "rows": [
                {"sheet_name": "Data", "row": 2},
                {"sheet_name": "Data", "row": 7},
                {"sheet_name": "Data", "row": 15},
            ],
        }
    ])

    assert payload == {
        "duplicateUrlCount": 1,
        "duplicateRowCount": 2,
        "items": [
            {
                "id": 1,
                "url": "https://www.tiktok.com/@demo/video/123",
                "locations": [
                    {"sheetName": "Data", "row": 2},
                    {"sheetName": "Data", "row": 7},
                    {"sheetName": "Data", "row": 15},
                ],
            }
        ],
    }


def test_build_duplicate_link_payload_ignores_unique_urls_and_preserves_sheets():
    import scraper

    assert hasattr(scraper, "build_duplicate_link_payload")
    payload = scraper.build_duplicate_link_payload([
        {
            "url": "https://www.tiktok.com/@demo/video/unique",
            "rows": [{"sheet_name": "Thang 6", "row": 4}],
        },
        {
            "url": "https://www.tiktok.com/@demo/video/duplicate",
            "rows": [
                {"sheet_name": "Thang 6", "row": 8},
                {"sheet_name": "Thang 7", "row": 11},
            ],
        },
    ])

    assert payload["duplicateUrlCount"] == 1
    assert payload["duplicateRowCount"] == 1
    assert payload["items"][0]["locations"] == [
        {"sheetName": "Thang 6", "row": 8},
        {"sheetName": "Thang 7", "row": 11},
    ]


def test_proxy_request_budget_keeps_photo_r1_candidate(monkeypatch):
    import scraper

    media_id = "7657817433285135636"
    source_url = (
        f"https://www.tiktok.com/@demo/photo/{media_id}"
        "?is_from_webapp=1&sender_device=pc"
    )
    r1_url = f"https://www.tiktok.com/@demo/photo/{media_id}?_r=1"
    success_content = f"""
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {{
      "__DEFAULT_SCOPE__": {{
        "webapp.photo-detail": {{
          "itemInfo": {{
            "itemStruct": {{
              "id": "{media_id}",
              "statsV2": {{
                "playCount": "525",
                "diggCount": "14",
                "commentCount": "0",
                "collectCount": "5",
                "shareCount": "0"
              }}
            }}
          }}
        }}
      }}
    }}
    </script>
    """
    fetched_urls = []

    def fake_fetch(url, timeout=30):
        fetched_urls.append(url)
        return url, success_content if url == r1_url else "<html></html>"

    monkeypatch.setattr(scraper, "session_uses_proxy", lambda: True)
    monkeypatch.setattr(scraper, "fetch_tiktok_html", fake_fetch)

    data, _channel, status, _hints, _resolved = scraper._scrape_link_request_impl(source_url)

    assert status == "Success"
    assert data == {
        "Views": "525",
        "Likes": "14",
        "Comments": "0",
        "Saves": "5",
        "Shares": "0",
    }
    assert r1_url in fetched_urls


def test_scrape_link_request_parses_photo_statsv2():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.photo-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "764002",
              "stats": {
                "playCount": 0,
                "diggCount": 45,
                "commentCount": 2,
                "collectCount": 0,
                "shareCount": 1
              },
              "statsV2": {
                "playCount": "1234",
                "diggCount": "45",
                "commentCount": "2",
                "collectCount": "0",
                "shareCount": "1"
              }
            }
          }
        }
      }
    }
    </script>
    """
    from scraper import scrape_link_request

    def fake_fetch(url, timeout=30):
        return "https://www.tiktok.com/@demo/photo/764002", content

    import scraper as scraper_module

    original = scraper_module.fetch_tiktok_html
    scraper_module.fetch_tiktok_html = fake_fetch
    try:
        data, channel, status = scrape_link_request("https://www.tiktok.com/@demo/photo/764002")
    finally:
        scraper_module.fetch_tiktok_html = original

    assert status == "Success"
    assert data["Views"] == "1234"


def test_parse_counts_extracts_matching_photo_metrics_from_api_data():
    from scraper import parse_counts

    content = """
    <script id="api-data" type="application/json">
    {
      "videoDetail": {
        "itemInfo": {
          "itemStruct": {
            "id": "7657817433285135636",
            "stats": {
              "playCount": 525,
              "diggCount": 14,
              "commentCount": 0,
              "collectCount": 5,
              "shareCount": 0
            }
          }
        }
      }
    }
    </script>
    """

    data, found = parse_counts(content, media_id="7657817433285135636")

    assert found is True
    assert data == {
        "Views": "525",
        "Likes": "14",
        "Comments": "0",
        "Saves": "5",
        "Shares": "0",
    }


def test_parse_counts_handles_api_data_script_variants():
    import json

    media_id = "7657817433285135636"
    payload = {
        "videoDetail": {
            "itemInfo": {
                "itemStruct": {
                    "id": media_id,
                    "desc": "A &quot;quoted&quot; caption",
                    "stats": {
                        "playCount": 525,
                        "diggCount": 14,
                        "commentCount": 0,
                        "collectCount": 5,
                        "shareCount": 0,
                    },
                }
            }
        }
    }
    encoded = json.dumps(payload)
    variants = [
        f"<script ID='API-DATA' type='application/json'>{encoded}</script>",
        f'<script id="api-data">{{not-json}}</script><script id="api-data">{encoded}</script>',
    ]

    for content in variants:
        data, found = parse_counts(content, media_id=media_id)
        assert found is True
        assert data == {
            "Views": "525",
            "Likes": "14",
            "Comments": "0",
            "Saves": "5",
            "Shares": "0",
        }


def test_universal_detail_parser_matches_normalized_scope_keys():
    from scraper import parse_counts_from_universal_data

    data = {
        "__DEFAULT_SCOPE__": {
            "webapp.photo-detail": {
                "itemInfo": {
                    "itemStruct": {
                        "id": "123",
                        "stats": {
                            "playCount": 10,
                            "diggCount": 2,
                            "commentCount": 0,
                            "collectCount": 1,
                            "shareCount": 0,
                        },
                    }
                }
            }
        }
    }

    assert parse_counts_from_universal_data(data, media_id="123") == {
        "Views": "10",
        "Likes": "2",
        "Comments": "0",
        "Saves": "1",
        "Shares": "0",
    }


def test_hybrid_browser_worker_count_caps_fallback_workers():
    from scraper import hybrid_browser_worker_count, MAX_BROWSER_FALLBACK_WORKERS

    assert hybrid_browser_worker_count(50, 1000) == MAX_BROWSER_FALLBACK_WORKERS
    assert hybrid_browser_worker_count(6, 20) >= 3


def test_parse_counts_uses_statsv2_when_stats_playcount_is_zero():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.photo-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "764002",
              "stats": {
                "playCount": 0,
                "diggCount": 45,
                "commentCount": 2,
                "collectCount": 0,
                "shareCount": 1
              },
              "statsV2": {
                "playCount": "1234",
                "diggCount": "45",
                "commentCount": "2",
                "collectCount": "0",
                "shareCount": "1"
              }
            }
          }
        }
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="764002")
    assert found is True
    assert data["Views"] == "1234"
    assert data["Likes"] == "45"


def test_validate_metrics_rejects_zero_views_with_engagement():
    assert validate_metrics(
        {"Views": "0", "Likes": "45", "Comments": "0", "Saves": "0", "Shares": "1"}
    ) is False
    assert validate_metrics(
        {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    ) is True


def test_parse_counts_extracts_metrics_from_photo_detail_json():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.photo-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "764001",
              "stats": {
                "playCount": 321,
                "diggCount": 12,
                "commentCount": 3,
                "collectCount": 1,
                "shareCount": 2
              }
            }
          }
        }
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="764001")
    assert found is True
    assert data["Views"] == "321"


def test_parse_counts_extracts_metrics_from_embedded_json():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "999",
              "stats": {
                "playCount": 12345,
                "diggCount": 678,
                "commentCount": 90,
                "collectCount": 12,
                "shareCount": 34
              }
            }
          }
        }
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="999")
    assert found is True
    assert data["Views"] == "12345"
    assert data["Likes"] == "678"
    assert data["Comments"] == "90"
    assert data["Saves"] == "12"
    assert data["Shares"] == "34"


def test_parse_counts_prefers_matching_media_id_over_other_videos():
    content = """
    <script id="SIGI_STATE" type="application/json">
    {
      "ItemModule": {
        "111": {
          "id": "111",
          "stats": {
            "playCount": 10,
            "diggCount": 1,
            "commentCount": 0,
            "collectCount": 0,
            "shareCount": 0
          }
        },
        "222": {
          "id": "222",
          "stats": {
            "playCount": 50000,
            "diggCount": 4000,
            "commentCount": 200,
            "collectCount": 50,
            "shareCount": 30
          }
        }
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="222")
    assert found is True
    assert data["Views"] == "50000"
    assert data["Likes"] == "4000"


def test_parse_counts_rejects_single_embedded_item_with_different_media_id():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "999",
              "stats": {
                "playCount": 5000,
                "diggCount": 100,
                "commentCount": 3,
                "collectCount": 2,
                "shareCount": 10
              }
            }
          }
        }
      }
    }
    </script>
    """

    data, found = parse_counts(content, media_id="123")

    assert found is False
    assert data == {
        "Views": "0",
        "Likes": "0",
        "Comments": "0",
        "Saves": "0",
        "Shares": "0",
    }


def test_parse_counts_rejects_sigi_item_key_with_mismatched_inner_id():
    content = """
    <script id="SIGI_STATE" type="application/json">
    {
      "ItemModule": {
        "123": {
          "id": "999",
          "stats": {
            "playCount": 500,
            "diggCount": 25,
            "commentCount": 1,
            "collectCount": 2,
            "shareCount": 3
          }
        }
      }
    }
    </script>
    """

    data, found = parse_counts(content, media_id="123")

    assert found is False
    assert data["Views"] == "0"


def test_parse_counts_treats_missing_collect_count_as_zero_for_exact_item():
    content = """
    <script id="SIGI_STATE" type="application/json">
    {
      "ItemModule": {
        "123": {
          "id": "123",
          "stats": {
            "playCount": 500,
            "diggCount": 25,
            "commentCount": 1,
            "shareCount": 3
          }
        }
      }
    }
    </script>
    """

    data, found = parse_counts(content, media_id="123")

    assert found is True
    assert data == {
        "Views": "500",
        "Likes": "25",
        "Comments": "1",
        "Saves": "0",
        "Shares": "3",
    }


def test_parse_counts_rejects_idless_item_when_media_id_is_expected():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "stats": {
                "playCount": 5000,
                "diggCount": 100,
                "commentCount": 3,
                "collectCount": 2,
                "shareCount": 10
              }
            }
          }
        }
      }
    }
    </script>
    """

    data, found = parse_counts(content, media_id="123")

    assert found is False
    assert data["Views"] == "0"


def test_parse_counts_returns_defaults_when_missing():
    data, found = parse_counts("<html></html>")
    assert found is False
    assert data["Views"] == "0"


def test_parse_counts_rejects_ambiguous_multiple_items_without_media_id():
    content = """
    <script id="SIGI_STATE" type="application/json">
    {
      "ItemModule": {
        "111": {"id":"111","stats":{"playCount":10,"diggCount":1,"commentCount":0,"collectCount":0,"shareCount":0}},
        "222": {"id":"222","stats":{"playCount":20,"diggCount":2,"commentCount":0,"collectCount":0,"shareCount":0}}
      }
    }
    </script>
    """
    data, found = parse_counts(content, media_id="")
    assert found is False
    assert data["Views"] == "0"


def test_parse_count_value_supports_suffixes():
    assert parse_count_value("1.2M") == "1200000"
    assert parse_count_value("12K") == "12000"
    assert parse_count_value(4567) == "4567"


def test_validate_metrics_rejects_implausible_values():
    assert validate_metrics(
        {"Views": "100", "Likes": "900", "Comments": "0", "Saves": "0", "Shares": "0"}
    ) is False
    assert validate_metrics(
        {"Views": "1000", "Likes": "50", "Comments": "4", "Saves": "2", "Shares": "1"}
    ) is True


def test_counts_match_requires_all_metrics_equal():
    left = {"Views": "10", "Likes": "1", "Comments": "0", "Saves": "0", "Shares": "0"}
    right = {"Views": "10", "Likes": "1", "Comments": "0", "Saves": "0", "Shares": "0"}
    assert counts_match(left, right) is True
    assert counts_match(left, {**right, "Views": "11"}) is False


def test_format_scrape_result_log_error_and_success():
    success_msg, success_level, success_details = format_scrape_result_log(
        {
            "url": "https://www.tiktok.com/@a/video/1",
            "worker": 2,
            "status": "Success",
            "attempts": 1,
            "elapsed": 1.25,
            "channel_name": "Channel A",
            "data": {"Views": "10", "Likes": "1", "Comments": "0", "Saves": "0", "Shares": "0"},
            "rows": [{"sheet_name": "Tháng 6", "row": 5}],
        },
        3,
        10,
    )
    assert success_level == "OK"
    assert success_details["kind"] == "scrape_ok"
    assert success_details["metrics"]["views"] == 10
    assert "Lượt xem 10" in success_msg
    assert "Tim 1" in success_msg
    assert "Tháng 6#5" in success_msg


def test_format_metric_log_line_uses_integers():
    assert format_metric_log_line(
        {"Views": "39", "Likes": "5", "Comments": "0", "Saves": "0", "Shares": "0"}
    ) == "view=39 tim=5 cmt=0 save=0 share=0"


def test_parse_channel_name_from_page_uses_item_struct_when_url_handle_mismatch():
    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.photo-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "7647351675514146069",
              "author": {
                "uniqueId": "vc.balo.i.trn",
                "nickname": "Vác Balo Đi Trốn"
              },
              "statsV2": {
                "playCount": "281",
                "diggCount": "13",
                "commentCount": "0",
                "collectCount": "3",
                "shareCount": "3"
              }
            }
          }
        }
      }
    }
    </script>
    """
    from scraper import parse_channel_name_from_page, parse_fetched_request_page

    name = parse_channel_name_from_page(
        content,
        "user16205752394791",
        media_id="7647351675514146069",
    )
    assert name == "Vác Balo Đi Trốn"

    metrics, channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@user16205752394791/photo/7647351675514146069",
        "https://www.tiktok.com/@user16205752394791/photo/7647351675514146069",
        content,
    )
    assert status == "Success"
    assert channel == "Vác Balo Đi Trốn"
    assert metrics["Views"] == "281"


def test_is_usable_channel_name_rejects_numeric_garbage():
    from scraper import is_usable_channel_name

    assert is_usable_channel_name("1.0") is False
    assert is_usable_channel_name("1") is False
    assert is_usable_channel_name("Lỗi") is False
    assert is_usable_channel_name("Đà Lạt Hệ Chill") is True


def test_channel_name_quality_ranks_real_name_above_handle():
    url = "https://www.tiktok.com/@demo/photo/1"
    assert channel_name_quality("Đà Lạt Hệ Chill", url) == 2
    assert channel_name_quality("@demo", url) == 1
    assert channel_name_quality("Lỗi", url) == 0
    assert channel_name_quality("1.0", url) == 0
    assert channel_name_quality("", url) == 0
    assert channel_name_quality("@user1234567890", url) == 0


def test_channel_name_for_sheet_does_not_write_generated_handle():
    url = "https://www.tiktok.com/@user16205752394791/photo/1"
    result = channel_name_for_sheet(
        url,
        "",
        resolved_url=url,
        status="Success",
        profile_lookup_attempted={"user16205752394791"},
    )
    assert result == "Lỗi"


def test_channel_name_for_sheet_keeps_real_handle_fallback():
    url = "https://www.tiktok.com/@.lt.h.chill/photo/1"
    result = channel_name_for_sheet(
        url,
        "",
        resolved_url=url,
        status="Success",
        profile_lookup_attempted={".lt.h.chill"},
    )
    assert result == "@.lt.h.chill"


def test_fetch_profile_channel_name_request_allows_user_id_accounts():
    profile_html = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {"userInfo": {"user": {"uniqueId": "user2657244715931", "nickname": "Lê Việt Khoa"}}}
    </script>
    """
    import scraper as scraper_module

    original = scraper_module.fetch_tiktok_html
    scraper_module.fetch_tiktok_html = lambda url, timeout=30: (url, profile_html)
    try:
        assert fetch_profile_channel_name_request("user2657244715931") == "Lê Việt Khoa"
    finally:
        scraper_module.fetch_tiktok_html = original


def test_enrich_channel_name_uses_cache_for_generated_user_accounts():
    url = "https://www.tiktok.com/@user2663512211600/photo/1"
    cache = {"user2663512211600": "Quỳnh Tiên"}
    name = enrich_channel_name(url, "", channel_cache=cache)
    assert name == "Quỳnh Tiên"


def test_enrich_channel_name_uses_resolved_url_for_short_links():
    short_url = "https://vt.tiktok.com/ZSxcMvARr/"
    resolved = "https://www.tiktok.com/@uyn.nh8183/photo/7652254869969014034"
    profile_html = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {"author": {"uniqueId": "uyn.nh8183", "nickname": "Uyển Như"}}
    </script>
    """
    import scraper as scraper_module

    original = scraper_module.fetch_tiktok_html
    scraper_module.fetch_tiktok_html = lambda url, timeout=30: (url, profile_html)
    try:
        name = enrich_channel_name(short_url, "", resolved_url=resolved, profile_lookup_attempted=set())
        assert name == "Uyển Như"
    finally:
        scraper_module.fetch_tiktok_html = original


def test_format_scrape_result_log_error():
    error_msg, error_level, error_details = format_scrape_result_log(
        {
            "url": "https://www.tiktok.com/@b/video/2",
            "worker": 1,
            "status": "Error: Không đọc được số liệu",
            "attempts": 3,
            "elapsed": 4.5,
            "rows": [{"sheet_name": "Tháng 6", "row": 9}],
        },
        4,
        10,
    )
    assert error_level == "ERROR"
    assert error_details["kind"] == "scrape_error"
    assert "Lỗi" in error_msg
    assert "Không đọc được số liệu" in error_msg


def test_extract_media_id_from_video_and_photo_urls():
    assert extract_media_id("https://www.tiktok.com/@a/video/1234567890") == "1234567890"
    assert extract_media_id("https://www.tiktok.com/@a/photo/9876543210") == "9876543210"


def test_is_tiktok_error_page_ignores_i18n_strings_in_script_tags():
    content = """
    <html><body>
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {"couldn't find this account":"couldn't find this account","page not available":"page not available"}
    </script>
    <div>8629 views</div>
    </body></html>
    """
    assert is_tiktok_error_page(content) is False


def test_is_tiktok_error_page_detects_visible_error_text():
    content = """
    <html><body>
    <h1>Couldn't find this account</h1>
    </body></html>
    """
    assert is_tiktok_error_page(content) is True


def test_parse_fetched_request_page_rejects_redirect_to_different_media_id():
    from scraper import STATUS_MEDIA_REDIRECT_MISMATCH, parse_fetched_request_page, should_clear_stale_metrics

    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "999",
              "stats": {
                "playCount": 777,
                "diggCount": 50,
                "commentCount": 2,
                "collectCount": 3,
                "shareCount": 4
              }
            }
          }
        }
      }
    }
    </script>
    """

    metrics, channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@source/video/123",
        "https://www.tiktok.com/@other/video/999",
        content,
    )

    assert status == STATUS_MEDIA_REDIRECT_MISMATCH
    assert should_clear_stale_metrics(status) is False
    assert channel == ""
    assert metrics["Views"] == "0"


def test_parse_fetched_request_page_rejects_redirect_without_media_id():
    from scraper import STATUS_MEDIA_REDIRECT_MISMATCH, parse_fetched_request_page

    content = """
    <script id="api-data" type="application/json">
    {"videoDetail":{"itemInfo":{"itemStruct":{
      "id":"123",
      "stats":{"playCount":525,"diggCount":14,"commentCount":0,"collectCount":5,"shareCount":0}
    }}}}
    </script>
    """

    metrics, channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@source/video/123",
        "https://www.tiktok.com/@source",
        content,
    )

    assert status == STATUS_MEDIA_REDIRECT_MISMATCH
    assert channel == ""
    assert metrics["Views"] == "0"


def test_parse_fetched_request_page_blank_response_is_unreadable():
    from scraper import STATUS_METRICS_UNREADABLE, parse_fetched_request_page

    metrics, channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@source/video/123",
        "https://www.tiktok.com/@source/video/123",
        "",
    )

    assert status == STATUS_METRICS_UNREADABLE
    assert channel == "@source"
    assert metrics["Views"] == "0"


def test_parse_fetched_request_page_prefers_exact_metrics_over_caption_error_phrase():
    from scraper import parse_fetched_request_page

    content = """
    <html><body>
    <p>My review: video unavailable yesterday</p>
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "123",
              "stats": {
                "playCount": 100,
                "diggCount": 5,
                "commentCount": 1,
                "collectCount": 0,
                "shareCount": 1
              }
            }
          }
        }
      }
    }
    </script>
    </body></html>
    """

    metrics, _channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@a/video/123",
        "https://www.tiktok.com/@a/video/123",
        content,
    )

    assert status == "Success"
    assert metrics["Views"] == "100"


def test_partial_matching_item_with_caption_phrase_stays_unreadable():
    from scraper import STATUS_METRICS_UNREADABLE, parse_fetched_request_page

    content = """
    <html><body>
    <p>Không tìm thấy bài hát này trong thư viện.</p>
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "123",
              "stats": {
                "playCount": 100,
                "diggCount": 5,
                "collectCount": 0,
                "shareCount": 1
              }
            }
          }
        }
      }
    }
    </script>
    </body></html>
    """

    metrics, _channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@a/video/123",
        "https://www.tiktok.com/@a/video/123",
        content,
    )

    assert status == STATUS_METRICS_UNREADABLE
    assert metrics["Views"] == "0"


def test_parse_fetched_request_page_requires_media_id_for_success():
    from scraper import STATUS_METRICS_UNREADABLE, parse_fetched_request_page

    content = """
    <script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">
    {
      "__DEFAULT_SCOPE__": {
        "webapp.video-detail": {
          "itemInfo": {
            "itemStruct": {
              "id": "999",
              "stats": {
                "playCount": 777,
                "diggCount": 50,
                "commentCount": 2,
                "collectCount": 3,
                "shareCount": 4
              }
            }
          }
        }
      }
    }
    </script>
    """

    metrics, _channel, status = parse_fetched_request_page(
        "https://www.tiktok.com/@source",
        "https://www.tiktok.com/@source",
        content,
    )

    assert status == STATUS_METRICS_UNREADABLE
    assert metrics["Views"] == "0"


def test_configure_request_concurrency_never_reduces_worker_count():
    import scraper
    from scraper import configure_request_concurrency

    configure_request_concurrency(20, proxy_count=0)
    assert scraper._request_semaphore._value == 20
    configure_request_concurrency(20, proxy_count=10)
    assert scraper._request_semaphore._value == 20


def test_clamp_worker_count_no_proxy_keeps_requested():
    from scraper import clamp_worker_count

    # Không proxy vẫn chạy đúng số luồng đã chọn, không tự giảm.
    assert clamp_worker_count(50, proxy_count=0) == 50


def test_clamp_worker_count_plenty_of_proxies_keeps_requested():
    from scraper import clamp_worker_count

    assert clamp_worker_count(30, proxy_count=10) == 30


def test_clamp_worker_count_few_proxies_keeps_requested_and_distributes_evenly():
    from scraper import clamp_worker_count

    # 50 luồng chỉ có 3 proxy -> vẫn chạy đủ 50 luồng, chia đều cho 3 proxy
    # (round-robin ở assign_worker_proxy), không tự giảm số luồng.
    assert clamp_worker_count(50, proxy_count=3) == 50


def test_clamp_worker_count_bounds_to_valid_range():
    from scraper import MAX_WORKERS, clamp_worker_count

    assert clamp_worker_count(0, proxy_count=0) == 1
    assert clamp_worker_count(999, proxy_count=0) == MAX_WORKERS


def test_is_request_rate_limited_status():
    from scraper import is_request_rate_limited_status

    assert is_request_rate_limited_status("Error: HTTP 403") is True
    assert is_request_rate_limited_status("Error: HTTP 429") is True
    assert is_request_rate_limited_status("Error: HTTP 404") is False


def test_is_transient_network_status():
    from scraper import is_transient_network_status

    assert is_transient_network_status("Error: [Errno 11001] getaddrinfo failed") is True
    assert is_transient_network_status("Error: [WinError 10054] connection was forcibly closed") is True
    assert is_transient_network_status("Error: timed out") is True
    assert is_transient_network_status(
        "Error: <urlopen error [SSL: UNEXPECTED_EOF_WHILE_READING] EOF occurred in violation of protocol (_ssl.c:1081)>"
    ) is True
    assert is_transient_network_status("Error: HTTP 404") is False
    assert is_transient_network_status("Ẩn số liệu: TikTok không trả lượt xem") is False
    assert is_transient_network_status("Lỗi: TikTok không trả số liệu") is False


def test_is_hidden_stats_status_accepts_new_and_legacy():
    from scraper import (
        STATUS_TIKTOK_NO_STATS,
        STATUS_TIKTOK_NO_STATS_LEGACY,
        is_hidden_stats_status,
    )

    assert STATUS_TIKTOK_NO_STATS.startswith("Ẩn số liệu")
    assert not STATUS_TIKTOK_NO_STATS.startswith("Lỗi:")
    assert is_hidden_stats_status(STATUS_TIKTOK_NO_STATS) is True
    assert is_hidden_stats_status(STATUS_TIKTOK_NO_STATS_LEGACY) is True
    assert is_hidden_stats_status("Error: HTTP 403") is False
    assert is_hidden_stats_status("Success") is False


def test_format_scrape_result_log_hidden_is_warn():
    from scraper import STATUS_TIKTOK_NO_STATS, format_scrape_result_log

    message, level, details = format_scrape_result_log(
        {
            "url": "https://www.tiktok.com/@a/video/1",
            "worker": "R1",
            "status": STATUS_TIKTOK_NO_STATS,
            "attempts": 1,
            "elapsed": 1.2,
            "data": {},
            "rows": [{"sheet_name": "Tháng 7", "row": 9}],
            "channel_name": "",
        },
        15,
        100,
    )
    assert level == "WARN"
    assert "Ẩn số liệu" in message
    assert details["kind"] == "scrape_hidden"


def test_scrape_link_retries_transient_network_errors(monkeypatch):
    import scraper

    calls = {"n": 0}

    def fake_impl(url, timeout=30):
        calls["n"] += 1
        empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
        return empty, "", "Error: [Errno 11001] getaddrinfo failed", False, ""

    monkeypatch.setattr(scraper, "_scrape_link_request_impl", fake_impl)
    monkeypatch.setattr(scraper.time, "sleep", lambda *_args, **_kwargs: None)

    _data, _channel, status, attempts, _resolved = scraper.scrape_link_with_retries_request(
        "https://www.tiktok.com/@demo/video/1",
        retries=2,
    )
    assert attempts == 3
    assert calls["n"] == 3
    assert "11001" in status


def test_scrape_link_retries_metrics_unreadable_without_hints(monkeypatch):
    import scraper

    calls = {"n": 0}

    def fake_impl(url, timeout=30):
        calls["n"] += 1
        empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
        return empty, "", scraper.STATUS_METRICS_UNREADABLE, False, ""

    monkeypatch.setattr(scraper, "_scrape_link_request_impl", fake_impl)
    monkeypatch.setattr(scraper.time, "sleep", lambda *_args, **_kwargs: None)

    _data, _channel, status, attempts, _resolved = scraper.scrape_link_with_retries_request(
        "https://www.tiktok.com/@demo/video/1",
        retries=2,
    )

    assert status == scraper.STATUS_METRICS_UNREADABLE
    assert attempts == 3
    assert calls["n"] == 3


def test_author_only_matching_shell_is_unreadable_not_hidden():
    from scraper import STATUS_METRICS_UNREADABLE, no_metrics_status

    content = """
    <script id="SIGI_STATE" type="application/json">
    {"ItemModule":{"123":{"id":"123","author":{"uniqueId":"demo"}}}}
    </script>
    """

    assert no_metrics_status(content, media_id="123") == STATUS_METRICS_UNREADABLE


def test_explicit_matching_stats_hidden_marker_is_classified_as_hidden():
    from scraper import STATUS_TIKTOK_NO_STATS, no_metrics_status

    content = """
    <script id="SIGI_STATE" type="application/json">
    {"ItemModule":{"123":{"id":"123","statsHidden":true,"author":{"uniqueId":"demo"}}}}
    </script>
    """

    assert no_metrics_status(content, media_id="123") == STATUS_TIKTOK_NO_STATS


def test_should_clear_stale_metrics_only_matches_terminal_statuses():
    from scraper import STATUS_TIKTOK_NO_STATS, should_clear_stale_metrics

    assert should_clear_stale_metrics(STATUS_TIKTOK_NO_STATS) is True
    assert should_clear_stale_metrics("Error: Trang TikTok không khả dụng") is True
    assert should_clear_stale_metrics("Error: HTTP 404") is True
    assert should_clear_stale_metrics("Error: HTTP 410 Gone") is True
    assert should_clear_stale_metrics("Error: Dịch vụ tạm thời không khả dụng") is False


def test_request_redirect_to_different_media_does_not_expose_target_url(monkeypatch):
    import scraper

    content = """
    <script id="SIGI_STATE" type="application/json">
    {"ItemModule":{"999":{"id":"999","stats":{"playCount":10,"diggCount":1,"commentCount":0,"collectCount":0,"shareCount":0}}}}
    </script>
    """

    monkeypatch.setattr(
        scraper,
        "fetch_tiktok_html",
        lambda *_args, **_kwargs: ("https://www.tiktok.com/@target/video/999", content),
    )
    monkeypatch.setattr(scraper, "request_candidate_limit", lambda: 1)

    _data, _channel, status, _hints, resolved_url = scraper._scrape_link_request_impl(
        "https://www.tiktok.com/@source/video/123"
    )

    assert status == scraper.STATUS_MEDIA_REDIRECT_MISMATCH
    assert resolved_url == ""


def test_request_candidate_terminal_status_is_not_overwritten_by_later_transient(monkeypatch):
    import urllib.error
    import scraper

    calls = {"n": 0}

    def fake_fetch(candidate, **_kwargs):
        calls["n"] += 1
        if calls["n"] == 1:
            raise urllib.error.HTTPError(candidate, 404, "Not Found", None, None)
        raise TimeoutError("timed out")

    monkeypatch.setattr(scraper, "fetch_tiktok_html", fake_fetch)
    monkeypatch.setattr(scraper, "request_candidate_limit", lambda: 2)
    monkeypatch.setattr(scraper, "note_network_failure", lambda: None)

    _data, _channel, status, _hints, _resolved_url = scraper._scrape_link_request_impl(
        "https://www.tiktok.com/@source/video/123?tracking=1"
    )

    assert calls["n"] == 2
    assert status == "Error: HTTP 404"


def test_request_shell_retry_preserves_terminal_http_error(monkeypatch):
    import urllib.error
    import scraper

    calls = {"count": 0}

    def fake_fetch(candidate, **_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return candidate, '<script>{"playCount":525}</script>'
        raise urllib.error.HTTPError(candidate, 404, "Not Found", None, None)

    monkeypatch.setattr(scraper, "fetch_tiktok_html", fake_fetch)

    _data, _channel, status, _hints, _resolved_url = scraper._scrape_link_request_impl(
        "https://www.tiktok.com/@source/video/123"
    )

    assert calls["count"] == 2
    assert status == "Error: HTTP 404"
    assert scraper.should_clear_stale_metrics(status) is True


def test_request_prefers_exact_hidden_status_over_earlier_http_404(monkeypatch):
    import urllib.error
    import scraper

    calls = {"count": 0}
    hidden_content = """
    <script id="SIGI_STATE" type="application/json">
    {"ItemModule":{"123":{"id":"123","statsHidden":true}}}
    </script>
    """

    def fake_fetch(candidate, **_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            raise urllib.error.HTTPError(candidate, 404, "Not Found", None, None)
        return candidate, hidden_content

    monkeypatch.setattr(scraper, "fetch_tiktok_html", fake_fetch)

    _data, _channel, status, _hints, _resolved_url = scraper._scrape_link_request_impl(
        "https://www.tiktok.com/@source/video/123?tracking=1"
    )

    assert status == scraper.STATUS_TIKTOK_NO_STATS


def test_proxy_candidate_budget_reaches_photo_video_alias(monkeypatch):
    import scraper

    media_id = "7657817433285135636"
    source_url = f"https://www.tiktok.com/@demo/photo/{media_id}?sender_device=pc"
    base_url = f"https://www.tiktok.com/@demo/photo/{media_id}"
    video_url = f"https://www.tiktok.com/@demo/video/{media_id}"
    success_content = f"""
    <script id="api-data" type="application/json">
    {{"videoDetail":{{"itemInfo":{{"itemStruct":{{
      "id":"{media_id}",
      "stats":{{"playCount":525,"diggCount":14,"commentCount":0,"collectCount":5,"shareCount":0}}
    }}}}}}}}
    </script>
    """
    fetched_urls = []

    def fake_fetch(url, **_kwargs):
        fetched_urls.append(url)
        if url == video_url:
            return url, success_content
        return (base_url if url == source_url else url), "<html></html>"

    monkeypatch.setattr(scraper, "session_uses_proxy", lambda: True)
    monkeypatch.setattr(scraper, "fetch_tiktok_html", fake_fetch)

    data, _channel, status, _hints, _resolved_url = scraper._scrape_link_request_impl(source_url)

    assert status == "Success"
    assert data["Views"] == "525"
    assert video_url in fetched_urls
    assert len(fetched_urls) <= scraper.request_candidate_limit()


def test_guard_result_media_identity_rejects_changed_short_url_target():
    from scraper import STATUS_MEDIA_REDIRECT_MISMATCH, guard_result_media_identity

    result = {
        "status": "Success",
        "resolved_url": "https://www.tiktok.com/@new/video/999",
        "expected_media_ids": ["123"],
        "channel_name": "Wrong channel",
        "data": {"Views": "900", "Likes": "90", "Comments": "9", "Saves": "9", "Shares": "9"},
    }

    guarded = guard_result_media_identity(result)

    assert guarded["status"] == STATUS_MEDIA_REDIRECT_MISMATCH
    assert guarded["resolved_url"] == ""
    assert guarded["channel_name"] == ""
    assert guarded["data"] == {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}


def test_collect_rows_binds_resolved_media_only_to_the_same_source_url():
    import openpyxl
    from scraper import collect_rows

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Link", "__TTBD_RESOLVED_URL", "__TTBD_SOURCE_URL"])
    sheet.append([
        "https://vt.tiktok.com/ZSsame/",
        "https://www.tiktok.com/@same/video/123",
        "https://vt.tiktok.com/ZSsame/",
    ])
    sheet.append([
        "https://vt.tiktok.com/ZSnew/",
        "https://www.tiktok.com/@old/video/456",
        "https://vt.tiktok.com/ZSold/",
    ])

    rows = collect_rows(workbook)

    assert rows[0]["expected_media_id"] == "123"
    assert rows[1]["expected_media_id"] == ""
    workbook.close()


def test_scrape_with_retries_stops_after_terminal_browser_result(monkeypatch):
    import asyncio
    import scraper

    calls = {"n": 0}

    async def fake_scrape_single_link(*_args, **_kwargs):
        calls["n"] += 1
        empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
        return empty, "", "Error: Trang TikTok không khả dụng", ""

    monkeypatch.setattr(scraper, "scrape_single_link", fake_scrape_single_link)

    result = asyncio.run(
        scraper.scrape_with_retries(object(), "https://www.tiktok.com/@demo/video/1", retries=2)
    )

    assert calls["n"] == 1
    assert result[2] == "Error: Trang TikTok không khả dụng"
    assert result[3] == 1


def test_select_fallback_result_preserves_terminal_request_evidence():
    from scraper import select_fallback_result

    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    request_result = {
        "data": empty,
        "channel_name": "",
        "status": "Error: HTTP 404",
        "attempts": 1,
        "resolved_url": "",
    }
    browser_result = {
        "data": empty,
        "channel_name": "",
        "status": "Error: Không đọc được số liệu",
        "attempts": 3,
        "resolved_url": "",
    }

    assert select_fallback_result(request_result, browser_result) is request_result
    browser_result["status"] = "Success"
    assert select_fallback_result(request_result, browser_result) is browser_result


def test_scrape_link_breaks_early_on_no_stats(monkeypatch):
    import scraper

    calls = {"n": 0}

    def fake_impl(url, timeout=30):
        calls["n"] += 1
        empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
        return empty, "", scraper.STATUS_TIKTOK_NO_STATS, False, ""

    monkeypatch.setattr(scraper, "_scrape_link_request_impl", fake_impl)
    monkeypatch.setattr(scraper.time, "sleep", lambda *_args, **_kwargs: None)

    _data, _channel, status, attempts, _resolved = scraper.scrape_link_with_retries_request(
        "https://www.tiktok.com/@demo/video/1",
        retries=2,
    )
    assert attempts == 1
    assert calls["n"] == 1
    assert status == scraper.STATUS_TIKTOK_NO_STATS


def test_request_retries_stop_after_terminal_status(monkeypatch):
    import scraper

    calls = {"n": 0}

    def fake_impl(url, timeout=30):
        calls["n"] += 1
        empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
        return empty, "", "Error: HTTP 404", True, ""

    monkeypatch.setattr(scraper, "_scrape_link_request_impl", fake_impl)
    monkeypatch.setattr(scraper.time, "sleep", lambda *_args, **_kwargs: None)

    _data, _channel, status, attempts, _resolved = scraper.scrape_link_with_retries_request(
        "https://www.tiktok.com/@demo/video/1",
        retries=2,
    )

    assert status == "Error: HTTP 404"
    assert attempts == 1
    assert calls["n"] == 1


def test_note_network_failure_pauses_after_streak(monkeypatch):
    import scraper

    scraper.configure_request_concurrency(5)
    sleeps = []
    monkeypatch.setattr(scraper.time, "sleep", lambda seconds: sleeps.append(seconds))
    monkeypatch.setattr(scraper.time, "time", lambda: 1000.0)

    for _ in range(scraper.NETWORK_FAIL_STREAK_PAUSE):
        scraper.note_network_failure()

    scraper.wait_if_request_blocked()
    assert sleeps
    assert sleeps[0] >= scraper.NETWORK_FAIL_PAUSE_SECONDS - 0.01
    scraper.configure_request_concurrency(5)


def test_progress_payload_includes_hidden_count():
    from scraper import progress_payload

    payload = progress_payload(10, 5, 3, 1, 2, 1000.0, hidden_count=1)
    assert payload["success"] == 3
    assert payload["error"] == 1
    assert payload["hidden"] == 1


def test_write_result_updates_timestamp_only_on_success():
    import openpyxl
    from scraper import LAST_UPDATE_HEADER, write_result

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tháng 7"
    ws.append(["Link", "Tên Kênh", "LƯỢT XEM", "TIM", LAST_UPDATE_HEADER])
    ws.append(["https://www.tiktok.com/@a/video/1", "", 0, 0, "01/01/2026-00:00"])
    contexts = {
        "Tháng 7": {
            "worksheet": ws,
            "columns": {
                "views": 3,
                "likes": 4,
                "channel": 2,
                "last_update": 5,
            },
        }
    }
    item = {"sheet_name": "Tháng 7", "row": 2, "url": "https://www.tiktok.com/@a/video/1"}
    before = ws.cell(row=2, column=5).value

    write_result(
        contexts,
        item,
        {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"},
        "",
        "Error: Worker đã dừng",
    )
    assert ws.cell(row=2, column=5).value == before

    write_result(
        contexts,
        item,
        {"Views": "10", "Likes": "1", "Comments": "0", "Saves": "0", "Shares": "0"},
        "@a",
        "Success",
    )
    assert ws.cell(row=2, column=5).value != before
    assert ws.cell(row=2, column=3).value == 10
    wb.close()


def test_write_result_clears_metrics_for_definitive_no_data_statuses():
    import openpyxl
    from scraper import LAST_UPDATE_HEADER, STATUS_TIKTOK_NO_STATS, write_result

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tháng 7"
    ws.append([
        "Link",
        "Tên Kênh",
        "LƯỢT XEM",
        "TIM",
        "BÌNH LUẬN",
        "LƯỢT LƯU",
        "CHIA SẺ",
        LAST_UPDATE_HEADER,
    ])
    ws.append([
        "https://www.tiktok.com/@a/video/1",
        "Kênh A",
        100,
        10,
        2,
        3,
        4,
        "01/01/2026-00:00",
    ])
    contexts = {
        "Tháng 7": {
            "worksheet": ws,
            "columns": {
                "views": 3,
                "likes": 4,
                "comments": 5,
                "saves": 6,
                "shares": 7,
                "channel": 2,
                "last_update": 8,
            },
        }
    }
    item = {"sheet_name": "Tháng 7", "row": 2, "url": "https://www.tiktok.com/@a/video/1"}
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}

    for status in (STATUS_TIKTOK_NO_STATS, "Error: Trang TikTok không khả dụng", "Error: HTTP 404"):
        for column, value in zip(range(3, 8), (100, 10, 2, 3, 4)):
            ws.cell(row=2, column=column).value = value

        write_result(contexts, item, empty, "", status)

        assert [ws.cell(row=2, column=column).value for column in range(3, 8)] == [0, 0, 0, 0, 0]

    wb.close()


def test_write_result_preserves_metrics_for_transient_failure():
    import openpyxl
    from scraper import LAST_UPDATE_HEADER, STATUS_METRICS_UNREADABLE, write_result

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tháng 7"
    ws.append(["Link", "Tên Kênh", "LƯỢT XEM", "TIM", "CHIA SẺ", LAST_UPDATE_HEADER])
    ws.append(["https://www.tiktok.com/@a/video/1", "Kênh A", 100, 10, 4, "01/01/2026-00:00"])
    contexts = {
        "Tháng 7": {
            "worksheet": ws,
            "columns": {
                "views": 3,
                "likes": 4,
                "shares": 5,
                "channel": 2,
                "last_update": 6,
            },
        }
    }
    item = {"sheet_name": "Tháng 7", "row": 2, "url": "https://www.tiktok.com/@a/video/1"}
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}

    for status in ("Error: timed out", "Error: HTTP 429", STATUS_METRICS_UNREADABLE):
        write_result(contexts, item, empty, "", status)
        assert [ws.cell(row=2, column=column).value for column in (3, 4, 5)] == [100, 10, 4]

    wb.close()


def test_write_result_persists_scan_status_and_resolved_url():
    import openpyxl
    from scraper import STATUS_METRICS_UNREADABLE, write_result

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([
        "Link",
        "Tên Kênh",
        "LƯỢT XEM",
        "TIM",
        "CHIA SẺ",
        "__TTBD_SCAN_STATUS",
        "__TTBD_RESOLVED_URL",
        "__TTBD_SOURCE_URL",
    ])
    ws.append(["https://vt.tiktok.com/ZSdemo/", "Kênh A", 100, 10, 4, "", "", ""])
    contexts = {
        "Data": {
            "worksheet": ws,
            "columns": {
                "views": 3,
                "likes": 4,
                "shares": 5,
                "channel": 2,
                "scan_status": 6,
                "resolved_url": 7,
                "source_url": 8,
            },
        }
    }
    item = {"sheet_name": "Data", "row": 2, "url": "https://vt.tiktok.com/ZSdemo/"}
    data = {"Views": "120", "Likes": "12", "Comments": "0", "Saves": "0", "Shares": "5"}
    resolved = "https://www.tiktok.com/@demo/video/123"

    write_result(contexts, item, data, "@demo", "Success", resolved_url=resolved)
    assert ws.cell(row=2, column=6).value == "Success"
    assert ws.cell(row=2, column=7).value == resolved
    assert ws.cell(row=2, column=8).value == item["url"]

    write_result(contexts, item, data, "", STATUS_METRICS_UNREADABLE, resolved_url=resolved)
    assert ws.cell(row=2, column=6).value == STATUS_METRICS_UNREADABLE
    assert ws.cell(row=2, column=7).value == resolved
    assert ws.cell(row=2, column=8).value == item["url"]
    assert [ws.cell(row=2, column=column).value for column in (3, 4, 5)] == [120, 12, 5]
    wb.close()


def test_write_result_preserves_existing_channel_on_failed_scan(monkeypatch):
    import openpyxl
    import scraper

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Link", "Tên Kênh", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    ws.append(["https://www.tiktok.com/@source/video/123", "Source Channel", 100, 10, 4])
    contexts = {
        "Data": {
            "worksheet": ws,
            "columns": {"channel": 2, "views": 3, "likes": 4, "shares": 5},
        }
    }
    item = {"sheet_name": "Data", "row": 2, "url": "https://www.tiktok.com/@source/video/123"}
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    monkeypatch.setattr(scraper, "channel_name_for_sheet", lambda *_args, **_kwargs: "Target Channel")

    scraper.write_result(
        contexts,
        item,
        empty,
        "Target Channel",
        "Error: Trang TikTok không khả dụng",
        resolved_url="https://www.tiktok.com/@target/video/999",
    )

    assert ws.cell(row=2, column=2).value == "Source Channel"
    wb.close()


def test_write_result_does_not_use_failed_redirect_for_channel_lookup(monkeypatch):
    import openpyxl
    import scraper

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Link", "Tên Kênh"])
    ws.append(["https://www.tiktok.com/@source/video/123", ""])
    contexts = {"Data": {"worksheet": ws, "columns": {"channel": 2}}}
    item = {"sheet_name": "Data", "row": 2, "url": "https://www.tiktok.com/@source/video/123"}
    empty = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    captured = {}

    def fake_channel_name_for_sheet(*_args, **kwargs):
        captured["resolved_url"] = kwargs.get("resolved_url")
        return "Lỗi"

    monkeypatch.setattr(scraper, "channel_name_for_sheet", fake_channel_name_for_sheet)

    scraper.write_result(
        contexts,
        item,
        empty,
        "",
        "Error: Trang TikTok không khả dụng",
        resolved_url="https://www.tiktok.com/@target/video/999",
    )

    assert captured["resolved_url"] == ""
    wb.close()


def test_run_scraper_does_not_report_completion_when_final_save_fails(tmp_path, monkeypatch):
    import asyncio
    import openpyxl
    import scraper

    file_path = tmp_path / "report.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Link", "Tên Kênh", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@demo/video/123", "", 0, 0, 0, 0, 0])
    workbook.save(file_path)
    workbook.close()

    class FakePlaywrightContext:
        async def __aenter__(self):
            return object()

        async def __aexit__(self, _exc_type, _exc, _traceback):
            return False

    def fake_scrape(*_args, **_kwargs):
        return (
            {"Views": "100", "Likes": "10", "Comments": "1", "Saves": "2", "Shares": "3"},
            "Demo",
            "Success",
            1,
            "https://www.tiktok.com/@demo/video/123",
        )

    async def failed_save(*_args, **_kwargs):
        return False, "permission"

    history_entries = []
    monkeypatch.setattr(scraper, "async_playwright", lambda: FakePlaywrightContext())
    monkeypatch.setattr(scraper, "_run_request_scrape", fake_scrape)
    monkeypatch.setattr(scraper, "save_workbook", failed_save)
    monkeypatch.setattr(scraper, "append_scrape_history", lambda _base, entry: history_entries.append(entry))

    with pytest.raises(RuntimeError, match="Excel"):
        asyncio.run(
            scraper.run_scraper(
                str(file_path),
                worker_count=1,
                retries=0,
                save_every=5,
                sheet_name="Data",
                use_request=True,
                browser_fallback=False,
            )
        )

    assert history_entries == []
