import pandas as pd

from workbook_utils import (
    dataframe_partner_columns,
    fetch_google_spreadsheet_title,
    find_link_column_name,
    google_sheet_file_id_from_title,
    format_display_datetime,
    format_excel_sheet_datetime,
    format_filename_datetime,
    google_sheet_filename_to_label,
    google_sheet_sync_label,
    highlight_single_partner_link_rows,
    highlight_video_link_rows,
    is_internal_workbook_filename,
    parse_filename_datetime_stamp,
    result_sheet_display_name,
    is_exportable_report_row,
    display_channel_name_from_file,
    is_failed_channel_name,
    is_generated_username_channel,
    is_scrapable_tiktok_url,
    is_result_sheet_name,
    is_result_sheet_timestamp_name,
    is_summary_sheet_name,
    metric_number,
    normalize_tiktok_url,
    build_workbook_rows,
    build_summary_totals_row,
    read_sheet_preview,
    rebuild_summary_sheet,
    safe_join,
    safe_workbook_filename,
    SINGLE_LINK_FILL_COLOR,
    VIDEO_LINK_FILL_COLOR,
    TIKTOK_MEDIA_PHOTO,
    TIKTOK_MEDIA_VIDEO,
    detect_tiktok_media_type,
    is_tiktok_photo_link,
    is_tiktok_video_link,
    month_label_for_sheet_name,
    should_highlight_video_link,
    split_partner_value,
    summary_sheet_title_for_data_sheet,
    to_number,
    workbook_file_entries,
    worksheet_find_link_column_index,
)


def test_metric_number_handles_thousand_separators():
    assert metric_number("1.234.567") == 1234567
    assert metric_number("1,234,567") == 1234567
    assert metric_number("150") == 150
    assert metric_number("") == 0
    assert metric_number(None) == 0


def test_link_column_fallback_ignores_internal_metadata_headers():
    import openpyxl

    frame = pd.DataFrame({"Nội dung": ["row"], "__TTBD_SOURCE_URL": ["https://vt.tiktok.com/ZSold/"]})
    assert find_link_column_name(frame) is None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Nội dung", "__TTBD_RESOLVED_URL", "__TTBD_SOURCE_URL"])
    assert worksheet_find_link_column_index(sheet) is None
    workbook.close()


def test_to_number_delegates_to_metric_number():
    assert to_number("2.500") == 2500


def test_split_partner_value_multiline():
    raw = "Partner A\nPartner B"
    assert split_partner_value(raw) == ["Partner A", "Partner B"]


def test_dataframe_partner_columns_detects_partner_block():
    frame = pd.DataFrame(
        {
            "Đối tác": ["A"],
            "Đối tác 2": ["B"],
            "LINK AIR": ["https://www.tiktok.com/@x/video/1"],
        }
    )
    columns = dataframe_partner_columns(frame)
    assert columns == ["Đối tác", "Đối tác 2"]


def test_is_failed_channel_name():
    assert is_failed_channel_name("") is True
    assert is_failed_channel_name("Lỗi") is True
    assert is_failed_channel_name("Error: timeout") is True
    assert is_failed_channel_name("Nice Cafe") is False
    assert is_failed_channel_name("Screen time breaks") is True
    assert is_failed_channel_name("1.0") is True
    assert is_failed_channel_name("1") is True
    assert is_failed_channel_name(1.0) is True


def test_display_channel_name_maps_tiktok_ui_garbage_to_loi():
    assert display_channel_name_from_file("https://www.tiktok.com/@a/video/1", "Screen time breaks") == "Lỗi"


def test_display_channel_name_maps_numeric_garbage_to_loi():
    link = "https://www.tiktok.com/@a/video/1"
    assert display_channel_name_from_file(link, "1.0") == "Lỗi"
    assert display_channel_name_from_file(link, 1.0) == "Lỗi"
    assert display_channel_name_from_file(link, "") == ""
    assert display_channel_name_from_file(link, "Nice Cafe") == "Nice Cafe"


def test_is_generated_username_channel():
    assert is_generated_username_channel("user1234567890") is True
    assert is_generated_username_channel("Nice Cafe") is False


def test_is_exportable_report_row_respects_min_views():
    row = {"TÊN KÊNH": "Cafe", "LƯỢT XEM": 50}
    assert is_exportable_report_row(row, apply_min_views=True, min_views=100) is False
    assert is_exportable_report_row(row, apply_min_views=False, min_views=100) is True


def test_google_sheet_file_id_uses_spreadsheet_title(monkeypatch):
    monkeypatch.setattr(
        "workbook_utils.fetch_google_spreadsheet_title",
        lambda _url: "Report Seeding Tiktok 2026",
    )
    file_id = google_sheet_file_id_from_title("Report Seeding Tiktok 2026", "05-06-2026-10-42")
    assert file_id == "data/Report Seeding Tiktok 2026 05-06-2026-10-42.xlsx"


def test_google_sheet_sync_label_includes_display_timestamp():
    label = google_sheet_sync_label("Report Seeding Tiktok 2026", "03/06/2026-14:30")
    assert label == "Report Seeding Tiktok 2026 03/06/2026-14:30"


def test_datetime_formats_are_consistent():
    moment = __import__("datetime").datetime(2026, 6, 3, 14, 30)
    assert format_display_datetime(moment) == "03/06/2026-14:30"
    assert format_filename_datetime(moment) == "03-06-2026-14-30"
    assert format_excel_sheet_datetime(moment) == "03-06-2026-14:30"
    assert parse_filename_datetime_stamp("03-06-2026-14-30") == "03/06/2026-14:30"


def test_result_sheet_display_name_uses_excel_safe_timestamp():
    name = result_sheet_display_name("17-06-2026-14:44")
    assert name == "17-06-2026-14:44"
    assert len(name) <= 31


def test_is_result_sheet_name_recognizes_timestamp_and_legacy_prefix():
    assert is_result_sheet_timestamp_name("17-06-2026-14:44") is True
    assert is_result_sheet_timestamp_name("17-06-2026-14:44-2") is True
    assert is_result_sheet_timestamp_name("03-06-2026-14-30") is True
    assert is_result_sheet_timestamp_name("03-06-2026-14-30-2") is True
    assert is_result_sheet_timestamp_name("Tháng 6") is False
    assert is_result_sheet_name("03-06-2026-14-30") is True
    assert is_result_sheet_name("Report Seeding Tiktok 03-06-2026-14-30") is True
    assert is_result_sheet_name("Tháng 6") is False


def test_google_sheet_filename_to_label():
    assert (
        google_sheet_filename_to_label("Report Seeding Tiktok 2026 05-06-2026-10-42.xlsx")
        == "Report Seeding Tiktok 2026 05/06/2026-10:42"
    )
    assert (
        google_sheet_filename_to_label("Report Seeding Tiktok 2026-05-06-2026-10-42.xlsx")
        == "Report Seeding Tiktok 2026 05/06/2026-10:42"
    )


def test_fetch_google_spreadsheet_title_from_html(monkeypatch):
    class FakeResponse:
        def read(self):
            return b"<html><title>Report Seeding Tiktok 2026 - Google Sheets</title></html>"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

    monkeypatch.setattr("urllib.request.urlopen", lambda *args, **kwargs: FakeResponse())
    title = fetch_google_spreadsheet_title("https://docs.google.com/spreadsheets/d/abc123/edit")
    assert title == "Report Seeding Tiktok 2026"


def test_safe_workbook_filename():
    assert safe_workbook_filename('File:name?') == "File-name-"


def test_summary_sheet_title_for_data_sheet():
    assert summary_sheet_title_for_data_sheet("Tháng 6") == "Tổng kết tháng 6"
    assert summary_sheet_title_for_data_sheet("Tháng 5") == "Tổng kết tháng 5"
    assert is_summary_sheet_name("Tổng kết tháng 6") is True
    assert is_summary_sheet_name("Tháng 6") is False


def test_month_label_for_sheet_name():
    assert month_label_for_sheet_name("Tháng 6") == "T6"
    assert month_label_for_sheet_name("Tháng 07") == "T7"
    assert month_label_for_sheet_name("thang12") == "T12"
    assert month_label_for_sheet_name("Tổng kết tháng 6") == "T6"
    assert month_label_for_sheet_name("Tháng 13") == ""
    assert month_label_for_sheet_name("") == ""
    assert month_label_for_sheet_name("Danh sách kênh") == ""


def test_rebuild_summary_sheet_only_one_data_sheet(tmp_path):
    import openpyxl

    file_path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    may = wb.active
    may.title = "Tháng 5"
    may.append(["LINK AIR", "Đối tác", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"])
    may.append(["https://www.tiktok.com/@a/video/1", "Partner May", 10, 1, 0, 0, 0])
    june = wb.create_sheet("Tháng 6")
    june.append(["LINK AIR", "Đối tác", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"])
    june.append(["https://www.tiktok.com/@b/video/2", "Partner June", 100, 2, 0, 0, 0])
    wb.save(file_path)
    wb.close()

    wb = openpyxl.load_workbook(file_path)
    count = rebuild_summary_sheet(wb, data_sheet_name="Tháng 6")
    assert count == 1
    assert "Tổng kết tháng 6" in wb.sheetnames
    summary = wb["Tổng kết tháng 6"]
    partner_names = [summary.cell(row=r, column=2).value for r in range(2, summary.max_row + 1)]
    assert partner_names == ["Partner June"]
    assert wb.sheetnames.index("Tổng kết tháng 6") == wb.sheetnames.index("Tháng 6") + 1
    wb.close()


def test_detect_tiktok_media_type_from_video_and_photo_urls():
    video_url = (
        "https://www.tiktok.com/@ngkhangg.008/video/7635258581851360520"
        "?is_from_webapp=1&web_id=7636006415740044807"
    )
    photo_url = (
        "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453"
        "?image_index=2&is_from_webapp=1&web_id=7636006415740044807"
    )
    assert detect_tiktok_media_type(video_url) == TIKTOK_MEDIA_VIDEO
    assert detect_tiktok_media_type(photo_url) == TIKTOK_MEDIA_PHOTO
    assert is_tiktok_video_link(video_url) is True
    assert is_tiktok_video_link(photo_url) is False
    assert is_tiktok_photo_link(photo_url) is True
    assert is_tiktok_photo_link(video_url) is False
    assert detect_tiktok_media_type("https://vt.tiktok.com/ZSabc123/") == ""
    assert (
        detect_tiktok_media_type(
            "https://vt.tiktok.com/ZSabc123/",
            resolved_url=video_url,
        )
        == TIKTOK_MEDIA_VIDEO
    )
    assert detect_tiktok_media_type(photo_url, resolved_url=video_url) == TIKTOK_MEDIA_PHOTO
    assert detect_tiktok_media_type(video_url, resolved_url=photo_url) == TIKTOK_MEDIA_VIDEO


def test_should_highlight_video_link_requires_readable_like_or_share_activity():
    video_url = "https://www.tiktok.com/@demo/video/123"
    photo_url = "https://www.tiktok.com/@demo/photo/123"
    short_url = "https://vt.tiktok.com/ZSdemo/"

    assert should_highlight_video_link(video_url, likes=1, shares=0) is True
    assert should_highlight_video_link(video_url, likes=0, shares=1) is True
    assert should_highlight_video_link(video_url, likes=0, shares=0) is False
    assert should_highlight_video_link(video_url, likes=1, shares=1, metrics_readable=False) is False
    assert should_highlight_video_link(photo_url, likes=1, shares=1) is False
    assert should_highlight_video_link(
        short_url,
        likes=1,
        shares=0,
        resolved_url=video_url,
    ) is True


def test_should_highlight_video_link_uses_latest_scan_status():
    video_url = "https://www.tiktok.com/@demo/video/123"

    assert should_highlight_video_link(
        video_url,
        likes=10,
        shares=4,
        scan_status="Success",
    ) is True
    assert should_highlight_video_link(
        video_url,
        likes=10,
        shares=4,
        scan_status="Error: Không đọc được số liệu",
    ) is False
    assert should_highlight_video_link(video_url, likes=10, shares=4, scan_status="") is True


def test_highlight_video_link_rows_marks_video_not_photo():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(
        [
            "https://www.tiktok.com/@ngkhangg.008/video/7635258581851360520",
            "Partner A",
            "Partner B",
            10,
            1,
            0,
        ]
    )
    sheet.append(
        [
            "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453",
            "Partner A",
            "Partner B",
            20,
            1,
            1,
        ]
    )

    count = highlight_video_link_rows(wb, "Tháng 6")

    assert count == 1
    video_fill = sheet.cell(row=2, column=1).fill
    photo_fill = sheet.cell(row=3, column=1).fill
    assert str(video_fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert photo_fill.fill_type is None or not str(photo_fill.fgColor.rgb or "").upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_highlight_video_link_rows_includes_single_partner_video():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Partner A", "", 10, 1, 0])

    count = highlight_video_link_rows(wb, "Tháng 6")

    assert count == 1
    assert str(sheet.cell(row=2, column=1).fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_highlight_video_link_rows_clears_stale_blue_from_zero_activity_video():
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Partner A", 10, 1, 0])
    sheet.append(["https://www.tiktok.com/@b/video/2", "Partner B", 20, 0, 1])
    sheet.append(["https://www.tiktok.com/@c/video/3", "Partner C", 30, 0, 0])
    stale_blue = PatternFill("solid", fgColor=VIDEO_LINK_FILL_COLOR)
    for column_index in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column_index).fill = stale_blue

    count = highlight_video_link_rows(wb, "Tháng 6")

    assert count == 2
    assert str(sheet.cell(row=2, column=1).fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert str(sheet.cell(row=3, column=1).fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert all(
        sheet.cell(row=4, column=column_index).fill.fill_type is None
        for column_index in range(1, sheet.max_column + 1)
    )


def test_inactive_single_partner_video_keeps_orange_warning():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Partner A", "", 10, 0, 0])

    orange_count = highlight_single_partner_link_rows(wb, "Tháng 6")
    blue_count = highlight_video_link_rows(wb, "Tháng 6")

    fill = sheet.cell(row=2, column=1).fill
    assert orange_count == 1
    assert blue_count == 0
    assert str(fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_highlight_single_partner_link_rows_skips_video_not_photo():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Partner A", "", 10, 1, 0])
    sheet.append(["https://www.tiktok.com/@b/photo/2", "Partner B", "", 20, 1, 1])
    sheet.append(["https://vt.tiktok.com/ZSabc123/", "Partner C", "", 30, 1, 1])

    count = highlight_single_partner_link_rows(wb, "Tháng 6")

    assert count == 2
    assert not str(sheet.cell(row=2, column=1).fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)
    assert str(sheet.cell(row=3, column=1).fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)
    assert str(sheet.cell(row=4, column=1).fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_highlight_video_link_rows_wrapper_after_single_partner():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Partner A", "", 10, 1, 0])
    sheet.append(["https://www.tiktok.com/@b/photo/2", "Partner B", "", 20, 1, 1])

    highlight_single_partner_link_rows(wb, "Tháng 6")
    count = highlight_video_link_rows(wb, "Tháng 6")

    assert count == 1
    fill = sheet.cell(row=2, column=1).fill
    assert str(fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert not str(fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_read_sheet_preview_flags_video_link_rows(tmp_path):
    import openpyxl

    file_path = tmp_path / "preview-video.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "TÊN KÊNH", "Đối tác", "Đối tác 2", "LƯỢT XEM", "TIM", "CHIA SẺ"])
    sheet.append(
        [
            "https://www.tiktok.com/@ngkhangg.008/video/7635258581851360520",
            "Video",
            "Partner A",
            "Partner B",
            10,
            1,
            0,
        ]
    )
    sheet.append(
        [
            "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453",
            "Photo",
            "Partner A",
            "Partner B",
            20,
            1,
            1,
        ]
    )
    sheet.append(
        [
            "https://www.tiktok.com/@inactive/video/3",
            "Inactive video",
            "Partner A",
            "Partner B",
            30,
            0,
            0,
        ]
    )
    wb.save(file_path)
    wb.close()

    preview = read_sheet_preview(str(file_path), sheet_name="Tháng 6")
    rows = preview["data"]
    assert rows[0].get("_videoLink") is True
    assert not rows[1].get("_videoLink")
    assert not rows[2].get("_videoLink")


def test_persisted_scan_metadata_drives_workbook_preview_and_is_hidden(tmp_path):
    import openpyxl

    file_path = tmp_path / "preview-scan-metadata.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data"
    sheet.append([
        "LINK AIR",
        "TÊN KÊNH",
        "Đối tác",
        "LƯỢT XEM",
        "TIM",
        "CHIA SẺ",
        "__TTBD_SCAN_STATUS",
        "__TTBD_RESOLVED_URL",
        "__TTBD_SOURCE_URL",
    ])
    sheet.append([
        "https://vt.tiktok.com/ZSactive/",
        "Active",
        "Partner A",
        100,
        10,
        4,
        "Success",
        "https://www.tiktok.com/@active/video/123",
        "https://vt.tiktok.com/ZSactive/",
    ])
    sheet.append([
        "https://www.tiktok.com/@stale/video/456",
        "Stale",
        "Partner B",
        200,
        20,
        5,
        "Error: Không đọc được số liệu",
        "https://www.tiktok.com/@stale/video/456",
        "https://www.tiktok.com/@stale/video/456",
    ])
    sheet.append([
        "https://vt.tiktok.com/ZSnew/",
        "Changed",
        "Partner C",
        300,
        30,
        6,
        "Success",
        "https://www.tiktok.com/@old/video/789",
        "https://vt.tiktok.com/ZSold/",
    ])

    blue_count = highlight_video_link_rows(wb, "Data")
    wb.save(file_path)
    wb.close()

    preview = read_sheet_preview(str(file_path), sheet_name="Data")

    assert blue_count == 1
    assert preview["data"][0].get("_videoLink") is True
    assert not preview["data"][1].get("_videoLink")
    assert not preview["data"][2].get("_videoLink")
    assert "__TTBD_SCAN_STATUS" not in preview["columns"]
    assert "__TTBD_RESOLVED_URL" not in preview["columns"]
    assert "__TTBD_SOURCE_URL" not in preview["columns"]


def test_build_workbook_rows_carries_internal_scan_metadata(tmp_path):
    import openpyxl

    file_path = tmp_path / "report-scan-metadata.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data"
    sheet.append([
        "LINK AIR",
        "TÊN KÊNH",
        "Đối tác",
        "LƯỢT XEM",
        "TIM",
        "BÌNH LUẬN",
        "LƯỢT LƯU",
        "CHIA SẺ",
        "__TTBD_SCAN_STATUS",
        "__TTBD_RESOLVED_URL",
        "__TTBD_SOURCE_URL",
    ])
    sheet.append([
        "https://vt.tiktok.com/ZSactive/",
        "Active",
        "Partner A",
        100,
        10,
        1,
        2,
        4,
        "Success",
        "https://www.tiktok.com/@active/video/123",
        "https://vt.tiktok.com/ZSactive/",
    ])
    wb.save(file_path)
    wb.close()

    rows = build_workbook_rows(str(file_path), sheet_name="Data")

    assert rows[0]["_scanStatus"] == "Success"
    assert rows[0]["_resolvedUrl"] == "https://www.tiktok.com/@active/video/123"
    assert rows[0]["_resolvedSourceUrl"] == "https://vt.tiktok.com/ZSactive/"


def test_highlight_single_partner_link_rows_marks_rows_with_exactly_one_partner():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM"])
    sheet.append(["https://vt.tiktok.com/ZSaaa111/", "Partner A", "", 10])
    sheet.append(["https://www.tiktok.com/@b/video/2", "Partner A", "Partner B", 20])
    sheet.append(["https://vt.tiktok.com/ZSbbb222/", "Partner B", "", 30])

    count = highlight_single_partner_link_rows(wb, "Tháng 6")

    assert count == 2
    single_fill_row2 = sheet.cell(row=2, column=1).fill
    multi_fill_row3 = sheet.cell(row=3, column=1).fill
    single_fill_row4 = sheet.cell(row=4, column=1).fill
    assert str(single_fill_row2.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)
    assert not str(multi_fill_row3.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)
    assert str(single_fill_row4.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_highlight_single_partner_link_rows_clears_stale_highlight_once_row_gains_partner():
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "Đối tác", "Đối tác 2", "LƯỢT XEM"])
    sheet.append(["https://vt.tiktok.com/ZSaaa111/", "Partner A", "", 10])

    first_count = highlight_single_partner_link_rows(wb, "Tháng 6")
    assert first_count == 1
    assert str(sheet.cell(row=2, column=1).fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)

    # Dòng vừa được gắn thêm 1 đối tác nữa -> highlight cũ phải được xóa.
    sheet.cell(row=2, column=3, value="Partner B")
    second_count = highlight_single_partner_link_rows(wb, "Tháng 6")

    assert second_count == 0
    assert not str(sheet.cell(row=2, column=1).fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_highlight_single_partner_link_rows_missing_sheet_returns_zero():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Tháng 6"
    assert highlight_single_partner_link_rows(wb, "Tháng 7") == 0


def test_read_sheet_preview_flags_single_partner_rows(tmp_path):
    import openpyxl

    file_path = tmp_path / "preview.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 6"
    sheet.append(["LINK AIR", "TÊN KÊNH", "Đối tác", "Đối tác 2", "LƯỢT XEM"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Kenh A", "Partner A", "", 10])
    sheet.append(["https://www.tiktok.com/@b/video/2", "Kenh B", "Partner A", "Partner B", 20])
    sheet.append(["", "Kenh C", "Partner A", "", 0])
    wb.save(file_path)
    wb.close()

    preview = read_sheet_preview(str(file_path), sheet_name="Tháng 6")

    rows = preview["data"]
    assert rows[0].get("_singlePartner") is True
    assert not rows[1].get("_singlePartner")
    assert not rows[2].get("_singlePartner")


def test_build_summary_totals_row_sums_partner_metrics():
    rows = [
        {
            "TỔNG LINK": 2,
            "TỔNG LƯỢT XEM": 100,
            "TỔNG TIM": 10,
            "TỔNG BÌNH LUẬN": 3,
            "TỔNG LƯỢT LƯU": 4,
            "TỔNG CHIA SẺ": 5,
        },
        {
            "TỔNG LINK": 1,
            "TỔNG LƯỢT XEM": 50,
            "TỔNG TIM": 5,
            "TỔNG BÌNH LUẬN": 1,
            "TỔNG LƯỢT LƯU": 2,
            "TỔNG CHIA SẺ": 1,
        },
    ]
    totals = build_summary_totals_row(rows)
    assert totals["ĐỐI TÁC"] == "TỔNG"
    assert totals["TỔNG LINK"] == 3
    assert totals["TỔNG LƯỢT XEM"] == 150
    assert totals["TỔNG TIM"] == 15
    assert totals["TỔNG CHIA SẺ"] == 6


def test_build_summary_totals_row_aligned_uses_actual_column_names():
    from workbook_utils import build_summary_totals_row_aligned

    columns = ["Stt", "Đối tác", "Tổng link", "Tổng lượt xem", "Tổng tim", "Cập nhật lần cuối"]
    rows = [
        {"Stt": 1, "Đối tác": "A", "Tổng link": 2, "Tổng lượt xem": 100, "Tổng tim": 10, "Cập nhật lần cuối": ""},
        {"Stt": 2, "Đối tác": "B", "Tổng link": 1, "Tổng lượt xem": 50, "Tổng tim": 5, "Cập nhật lần cuối": ""},
    ]
    aligned = build_summary_totals_row_aligned(columns, rows)
    assert aligned["Đối tác"] == "TỔNG"
    assert aligned["Tổng link"] == 3
    assert aligned["Tổng lượt xem"] == 150
    assert aligned["Tổng tim"] == 15


def test_is_scrapable_tiktok_url():
    assert is_scrapable_tiktok_url("tiktok.com/@a/video/1") is True
    assert is_scrapable_tiktok_url("https://www.tiktok.com/@a/video/1") is True
    assert is_scrapable_tiktok_url("TỔNG") is False
    assert is_scrapable_tiktok_url("") is False


def test_normalize_tiktok_url_adds_https():
    raw = "tiktok.com/@ngc.ngc.i.chill/photo/7646786750978854152"
    assert normalize_tiktok_url(raw) == "https://tiktok.com/@ngc.ngc.i.chill/photo/7646786750978854152"
    assert normalize_tiktok_url("https://www.tiktok.com/@a/video/1") == "https://www.tiktok.com/@a/video/1"
    assert normalize_tiktok_url("//www.tiktok.com/@a/video/1") == "https://www.tiktok.com/@a/video/1"


def test_safe_join_blocks_traversal(tmp_path):
    base = tmp_path / "root"
    base.mkdir()
    allowed = safe_join(str(base), "data/file.xlsx")
    assert allowed.endswith("file.xlsx")
    assert safe_join(str(base), "../outside.xlsx") == ""
    assert safe_join(str(base), "../../etc/passwd") == ""


def test_is_internal_workbook_filename():
    assert is_internal_workbook_filename("_test_export.xlsx") is True
    assert is_internal_workbook_filename("~$Report.xlsx") is True
    assert is_internal_workbook_filename("Report Seeding Tiktok.xlsx") is False


def test_workbook_file_entries_skips_internal_files(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (tmp_path / "Report.xlsx").write_bytes(b"x")
    (data_dir / "_test_export.xlsx").write_bytes(b"x")
    (data_dir / "Report Seeding Tiktok.xlsx").write_bytes(b"x")

    ids = {entry["id"] for entry in workbook_file_entries(str(tmp_path))}
    assert "data/_test_export.xlsx" not in ids
    assert "data/Report Seeding Tiktok.xlsx" in ids
    assert "Report.xlsx" in ids
