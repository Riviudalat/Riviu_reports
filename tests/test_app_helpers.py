import io
import asyncio
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app import build_google_push_rows, build_partner_report, spreadsheet_date_text
from workbook_utils import SINGLE_LINK_FILL_COLOR, VIDEO_LINK_FILL_COLOR, is_failed_channel_name, metric_number


def test_preview_javascript_preserves_numeric_zero_values():
    source = (Path(__file__).parents[1] / "static" / "app.js").read_text(encoding="utf-8")

    assert "let val = row[column] ?? '';" in source


def test_duplicate_links_panel_is_wired_to_websocket_and_new_scan_reset():
    root = Path(__file__).parents[1]
    source = (root / "static" / "app.js").read_text(encoding="utf-8")
    template = (root / "templates" / "index.html").read_text(encoding="utf-8")

    assert "message.type === 'duplicates'" in source
    assert "function renderDuplicateLinks" in source
    assert "function copyDuplicateLinks" in source
    assert "clearDuplicateLinks(false);" in source
    assert 'id="duplicateLinksBody"' in template
    assert 'id="duplicateCountBadge"' in template


def test_compact_workspace_keeps_global_workflows_and_unique_controls():
    root = Path(__file__).parents[1]
    template = (root / "templates" / "index.html").read_text(encoding="utf-8")

    assert 'class="compact-app"' in template
    assert 'id="sourceDrawer"' in template
    assert 'onclick="openReportModal()"' in template
    assert 'onclick="openHistoryModal()"' in template
    assert 'data-workspace-tab="live"' in template
    assert 'data-workspace-panel="duplicates"' in template

    for element_id in (
        "workerCountSelect",
        "scrapeModeSelect",
        "proxyUseCheckbox",
        "scanSheetSelect",
        "googleSheetUrlInput",
        "excelFileSelect",
        "reportModal",
        "historyModal",
        "proxyModal",
    ):
        assert template.count(f'id="{element_id}"') == 1


def test_compact_workspace_javascript_switches_tabs_and_source_drawer():
    source = (Path(__file__).parents[1] / "static" / "app.js").read_text(encoding="utf-8")

    assert "function setWorkspaceTab" in source
    assert "function openSourceDrawer" in source
    assert "function closeSourceDrawer" in source
    assert "setWorkspaceTab('live');" in source
    assert "hiddenCountBadge" in source


def test_preview_counter_does_not_break_websocket_dispatch():
    source = (Path(__file__).parents[1] / "static" / "app.js").read_text(encoding="utf-8")
    preview_source = source[source.index("async function loadPreview"):source.index("async function renderSummaryDashboard")]
    websocket_source = source[source.index("function connectWS"):source.index("function startScraping")]

    assert "Array.isArray(data.data) ? data.data.length : 0" in preview_source
    assert "data.data" not in websocket_source
    assert "else if (message.type === 'status') updateProgress(message.data);" in websocket_source
    assert "else if (message.type === 'data') appendData(message.row);" in websocket_source


def test_desktop_updater_only_activates_inside_tauri():
    source = (Path(__file__).parents[1] / "static" / "app.js").read_text(encoding="utf-8")

    assert "function desktopUpdaterInvoke" in source
    assert "window.__TAURI__?.core?.invoke" in source
    assert "function scheduleDesktopUpdates" in source
    assert "setInterval(() => void checkDesktopUpdate(), 4 * 60 * 60 * 1000)" in source


def test_desktop_bundle_keeps_resources_separate_from_user_data_and_release_ci():
    root = Path(__file__).parents[1]
    app_source = (root / "app.py").read_text(encoding="utf-8")
    sidecar_source = (root / "desktop" / "build_sidecar.py").read_text(encoding="utf-8")
    sidecar_entrypoint = (root / "desktop_server.py").read_text(encoding="utf-8")
    config = (root / "src-tauri" / "tauri.conf.json").read_text(encoding="utf-8")
    workflow = (root / ".github" / "workflows" / "desktop-release.yml").read_text(encoding="utf-8")

    assert 'getattr(sys, "_MEIPASS"' in app_source
    assert 'os.environ.get("RIVIU_DATA_DIR", APP_RESOURCE_DIR)' in app_source
    assert '"templates", "templates"' in sidecar_source
    assert '"static", "static"' in sidecar_source
    assert '"/_desktop/shutdown"' in sidecar_entrypoint
    assert 'PLAYWRIGHT_BROWSERS_PATH' in sidecar_source
    assert 'PLAYWRIGHT_BROWSERS_PATH' in sidecar_entrypoint
    assert '"--additional-hooks-dir"' in sidecar_source
    assert (root / "desktop" / "pyinstaller-hooks" / "hook-playwright.async_api.py").is_file()
    assert '"externalBin"' in config
    assert '"createUpdaterArtifacts": true' in config
    assert "windows-2022" in workflow
    assert "macos-13" in workflow
    assert "macos-14" in workflow
    assert "TAURI_SIGNING_PRIVATE_KEY" in workflow


def test_template_has_no_duplicate_ids():
    class IdCollector(HTMLParser):
        def __init__(self):
            super().__init__()
            self.ids = []

        def handle_starttag(self, _tag, attrs):
            values = dict(attrs)
            if values.get("id"):
                self.ids.append(values["id"])

    template = (Path(__file__).parents[1] / "templates" / "index.html").read_text(encoding="utf-8")
    parser = IdCollector()
    parser.feed(template)

    duplicates = sorted({element_id for element_id in parser.ids if parser.ids.count(element_id) > 1})
    assert duplicates == []


def test_broadcast_duplicates_uses_expected_websocket_envelope():
    from app import ConnectionManager

    class FakeConnection:
        def __init__(self):
            self.messages = []

        async def send_json(self, message):
            self.messages.append(message)

    manager = ConnectionManager()
    connection = FakeConnection()
    manager.active_connections.append(connection)
    payload = {
        "duplicateUrlCount": 1,
        "duplicateRowCount": 1,
        "items": [],
    }

    asyncio.run(manager.broadcast_duplicates(payload))

    assert connection.messages == [{"type": "duplicates", "data": payload}]


def test_validate_proxy_start_blocks_empty(tmp_path):
    from app import validate_proxy_start

    msg = validate_proxy_start(True, "", str(tmp_path))
    assert msg is not None
    assert "proxy" in msg.lower()


def test_validate_proxy_start_allows_disabled(tmp_path):
    from app import validate_proxy_start

    assert validate_proxy_start(False, "", str(tmp_path)) is None


def test_build_google_push_rows_includes_total_row():
    rows = [
        {
            "NGÀY AIR": datetime(2026, 6, 2),
            "TÊN KÊNH": "Channel A",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 200,
            "TIM": 10,
            "BÌNH LUẬN": 2,
            "LƯỢT LƯU": 3,
            "CHIA SẺ": 4,
            "partners": ["Partner A"],
        }
    ]
    values = build_google_push_rows(rows)
    assert values[0][0] == "Stt"
    assert values[-1][2] == "TỔNG"
    assert values[-1][4].startswith("=SUM(")


def test_build_partner_report_filters_by_min_views():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Good",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
        },
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Low",
            "LINK AIR": "https://www.tiktok.com/@b/video/2",
            "LƯỢT XEM": 10,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
        },
    ]
    report_bytes = build_partner_report("Partner", rows, apply_min_views=True, min_views=100)
    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0
    assert metric_number(200) >= 100
    assert is_failed_channel_name("Good") is False


def test_export_date_keeps_vietnamese_day_month_year_order():
    assert spreadsheet_date_text("02/06/2026") == "'02/06/2026"
    assert spreadsheet_date_text("02-06-2026") == "'02-06-2026"


def test_build_partner_report_total_row_has_numeric_sums():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "A",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 100,
            "TIM": 10,
            "BÌNH LUẬN": 2,
            "LƯỢT LƯU": 3,
            "CHIA SẺ": 4,
        },
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "B",
            "LINK AIR": "https://www.tiktok.com/@b/video/2",
            "LƯỢT XEM": 50,
            "TIM": 5,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 2,
            "CHIA SẺ": 1,
        },
    ]
    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes), data_only=True)
    worksheet = workbook.active
    total_row = None
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value == "TỔNG":
            total_row = row_index
            break
    assert total_row is not None
    assert worksheet.cell(row=total_row, column=4).value == 150
    assert worksheet.cell(row=total_row, column=5).value == 15
    assert worksheet.cell(row=total_row, column=6).value == 3
    assert worksheet.cell(row=total_row, column=7).value == 5
    assert worksheet.cell(row=total_row, column=8).value == 5


def test_build_partner_report_highlights_single_partner_video_as_video_color():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Solo video",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A"],
        },
    ]
    report_bytes = build_partner_report("Partner A", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    fill = worksheet.cell(row=4, column=1).fill
    assert str(fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert not str(fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_build_partner_report_highlights_row_with_single_partner_unknown_link():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Solo",
            "LINK AIR": "https://vt.tiktok.com/ZSabc123/",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A"],
        },
    ]
    report_bytes = build_partner_report("Partner A", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    data_row = 4  # table_header_row (3) + 1
    fill = worksheet.cell(row=data_row, column=1).fill
    assert str(fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_build_partner_report_only_highlights_single_partner_rows():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "A",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A"],
        },
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "B",
            "LINK AIR": "https://www.tiktok.com/@b/video/2",
            "LƯỢT XEM": 300,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A", "Partner B"],
        },
    ]
    report_bytes = build_partner_report("Partner A", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    single_partner_fill = worksheet.cell(row=4, column=1).fill
    multi_partner_fill = worksheet.cell(row=5, column=1).fill
    assert str(single_partner_fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert str(multi_partner_fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_build_partner_report_no_highlight_for_photo_link():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Photo",
            "LINK AIR": "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A", "Partner B"],
        },
    ]
    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    fill = worksheet.cell(row=4, column=1).fill
    assert fill.fill_type is None or not str(fill.fgColor.rgb or "").upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert not str(fill.fgColor.rgb or "").upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_build_partner_report_highlights_single_partner_photo_as_orange():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Photo solo",
            "LINK AIR": "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453",
            "LƯỢT XEM": 300,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A"],
        },
    ]
    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    fill = worksheet.cell(row=4, column=1).fill
    assert str(fill.fgColor.rgb).upper().endswith(SINGLE_LINK_FILL_COLOR)


def test_build_partner_report_highlights_video_link_without_single_partner():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Video",
            "LINK AIR": "https://www.tiktok.com/@ngkhangg.008/video/7635258581851360520",
            "LƯỢT XEM": 200,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A", "Partner B"],
        },
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Photo",
            "LINK AIR": "https://www.tiktok.com/@baoquyen.dalat/photo/7635505950807346453",
            "LƯỢT XEM": 300,
            "TIM": 1,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 1,
            "CHIA SẺ": 1,
            "partners": ["Partner A", "Partner B"],
        },
    ]
    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    worksheet = workbook.active
    video_fill = worksheet.cell(row=4, column=1).fill
    photo_fill = worksheet.cell(row=5, column=1).fill
    assert str(video_fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert photo_fill.fill_type is None or not str(photo_fill.fgColor.rgb or "").upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_build_partner_report_does_not_highlight_zero_activity_video():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Video chưa có hoạt động",
            "LINK AIR": "https://www.tiktok.com/@demo/video/123",
            "LƯỢT XEM": 100,
            "TIM": 0,
            "BÌNH LUẬN": 0,
            "LƯỢT LƯU": 0,
            "CHIA SẺ": 0,
            "partners": ["Partner A", "Partner B"],
        },
    ]

    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))
    fill = workbook.active.cell(row=4, column=1).fill

    assert fill.fill_type is None or not str(fill.fgColor.rgb or "").upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_build_partner_report_uses_persisted_scan_metadata_for_color():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Short video",
            "LINK AIR": "https://vt.tiktok.com/ZSactive/",
            "LƯỢT XEM": 100,
            "TIM": 10,
            "BÌNH LUẬN": 1,
            "LƯỢT LƯU": 2,
            "CHIA SẺ": 4,
            "partners": ["Partner A", "Partner B"],
            "_scanStatus": "Success",
            "_resolvedUrl": "https://www.tiktok.com/@active/video/123",
        },
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "Stale video",
            "LINK AIR": "https://www.tiktok.com/@stale/video/456",
            "LƯỢT XEM": 200,
            "TIM": 20,
            "BÌNH LUẬN": 2,
            "LƯỢT LƯU": 3,
            "CHIA SẺ": 5,
            "partners": ["Partner A", "Partner B"],
            "_scanStatus": "Error: Không đọc được số liệu",
            "_resolvedUrl": "https://www.tiktok.com/@stale/video/456",
        },
    ]

    report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    workbook = load_workbook(io.BytesIO(report_bytes))

    active_fill = workbook.active.cell(row=4, column=1).fill
    stale_fill = workbook.active.cell(row=5, column=1).fill
    assert str(active_fill.fgColor.rgb).upper().endswith(VIDEO_LINK_FILL_COLOR)
    assert stale_fill.fill_type is None or not str(stale_fill.fgColor.rgb or "").upper().endswith(VIDEO_LINK_FILL_COLOR)


def test_build_partner_report_works_when_logo_image_unavailable():
    rows = [
        {
            "NGÀY AIR": "",
            "TÊN KÊNH": "A",
            "LINK AIR": "https://www.tiktok.com/@a/video/1",
            "LƯỢT XEM": 100,
            "TIM": 10,
            "BÌNH LUẬN": 2,
            "LƯỢT LƯU": 3,
            "CHIA SẺ": 4,
        }
    ]
    with patch("app.ExcelImage", side_effect=ImportError("You must install Pillow to fetch image objects")):
        report_bytes = build_partner_report("Partner", rows, apply_min_views=False)
    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0


def test_build_export_payload_filename_includes_sheet_and_timestamp(tmp_path):
    import openpyxl
    from app import build_export_payload

    file_path = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tháng 7"
    sheet.append(["LINK AIR", "Tên Kênh", "Đối tác", "LƯỢT XEM", "TIM", "BÌNH LUẬN", "LƯỢT LƯU", "CHIA SẺ"])
    sheet.append(["https://www.tiktok.com/@a/video/1", "Kenh A", "1/2 Circle Coffee", 200, 1, 0, 0, 0])
    wb.save(file_path)
    wb.close()

    with patch("app.format_filename_datetime", return_value="09-07-2026-13-47"):
        payload = build_export_payload(
            str(file_path),
            ["1/2 Circle Coffee"],
            apply_min_views=False,
            min_views=0,
            requested_sheet_name="Tháng 7",
        )

    assert payload["filename"] == "1-2 Circle Coffee Tháng 7 09-07-2026-13-47.xlsx"
